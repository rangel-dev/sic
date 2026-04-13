"""
Auditor Engine – Double-Blind Price & Catalog Validation
Fiel à lógica do auditor.js v8.4 / V11.6.
"""
from __future__ import annotations

import re
import time
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import openpyxl
import pandas as pd
from lxml import etree

# ─── Namespaces ───────────────────────────────────────────────────────────────
PRICEBOOK_NS = "http://www.demandware.com/xml/impex/pricebook/2006-10-31"
CATALOG_NS   = "http://www.demandware.com/xml/impex/catalog/2006-10-31"

# ─── Regra dos 10 minutos (Auditor: arquivos SF devem ser RECENTES < 10 min) ─
MAX_FILE_AGE_SECONDS = 600

# ─── Categorias proibidas para promoção (Conflito de Margem) ─────────────────
PROHIBITED_CATEGORIES = {
    "Natura": {"promocao-da-semana", "LISTA_01"},
    "Avon":   {"desconto-progressivo", "lista-01"},
    "ML":     {"promocao-da-semana", "desconto-progressivo"},
}

# ─── Metadados de erro para a UI ─────────────────────────────────────────────
ERROR_META: dict[str, dict] = {
    "price":      {"title": "Divergência de Preço",         "impact": "Risco Financeiro Alto",       "icon": "💰", "desc": "Preço no Excel (DE/POR) não bate com o valor no Pricebook do Salesforce."},
    "list":       {"title": "Visibilidade em Listas",        "impact": "Baixa Conversão",             "icon": "📋", "desc": "SKU está na aba de lista (ex: LISTA_01) no Excel, mas não está na categoria equivalente no XML."},
    "ml":         {"title": "Conflito Canal ML",             "impact": "Conflito Estratégico",        "icon": "⚡", "desc": "Preço no catálogo 'Minha Loja' diverge do preço da marca principal (Natura/Avon)."},
    "margin":     {"title": "Margem de Segurança",           "impact": "Perda Operacional (Crítico)", "icon": "🚨", "desc": "Produto em promoção (POR < DE) dentro de uma categoria onde descontos são proibidos."},
    "logic":      {"title": "Erro Lógico (POR > DE)",        "impact": "Erro de Cadastro",            "icon": "⚠️", "desc": "O preço promocional (POR) está maior que o preço de lista (DE)."},
    "missing":    {"title": "Preço Ausente (DE/POR)",        "impact": "Perda Imediata de Venda",     "icon": "❌", "desc": "O produto não possui preço de lista (DE) ou preço promocional (POR) definido no Salesforce."},
    "primary":    {"title": "Categoria Primária",            "impact": "Impacto SEO e Filtros",       "icon": "🏷️", "desc": "O produto está no catálogo mas não possui uma categoria marcada como 'Primária'."},
    "bundle":     {"title": "Saúde de Bundles/Kits",         "impact": "Queda no Ticket Médio",       "icon": "📦", "desc": "O Kit/Bundle possui componentes que estão offline ou sem preço definido."},
    "cross":      {"title": "Cross-Brand (Invasão)",         "impact": "Erro Crítico de Governança",  "icon": "🔴", "desc": "Produto de uma marca (ex: Natura) possui preço no catálogo de outra marca (ex: Avon)."},
    "offline":    {"title": "Produto Indisponível",          "impact": "Risco de Receita",            "icon": "🔇", "desc": "O produto está na Grade de Ativação do Excel, mas está com a flag 'online=false' no SF."},
    "job":        {"title": "Falha de Sync (ML JOB)",        "impact": "Catálogo Desatualizado",      "icon": "🔄", "desc": "As categorias da Minha Loja (ML) estão divergentes das categorias espelho da marca original."},
    "searchable": {"title": "Flag Searchable",               "impact": "Perda de Fluxo Orgânico",     "icon": "🔍", "desc": "Divergência entre a coluna VISIBLE do Excel e a flag 'searchable' do Salesforce."},
}

SKU_RE = re.compile(r"^(NAT|AVN)BRA-", re.IGNORECASE)


# ─── Resultado ────────────────────────────────────────────────────────────────
@dataclass
class AuditResult:
    errors: dict[str, pd.DataFrame] = field(default_factory=dict)
    stats: dict = field(default_factory=dict)
    brands_found: list[str] = field(default_factory=list)
    total_excel_skus: int = 0
    error: Optional[str] = None
    preflight_error: Optional[str] = None


# ─── Engine ───────────────────────────────────────────────────────────────────
class AuditorEngine:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self._prog = progress_callback or (lambda p, m: None)

    # ── Entrada ───────────────────────────────────────────────────────────
    def run(self, excel_paths: list[str], pb_path: str, cat_paths: list[str]) -> AuditResult:
        result = AuditResult()
        try:
            # Pre-flight: arquivo SF deve ser recente (< 10 min) - DESATIVADO TEMPORARIAMENTE
            # self._prog(3, "Verificando antiguidade dos arquivos SF…")
            # expired = []
            # for p in [pb_path] + cat_paths:
            #     try:
            #         age = time.time() - os.path.getmtime(p)
            #         if age > MAX_FILE_AGE_SECONDS:
            #             expired.append(f"{Path(p).name} ({int(age/60)} min atrás)")
            #     except OSError:
            #         pass
            # if expired:
            #     result.preflight_error = (
            #         "Arquivos SF desatualizados (>10 min):\n\n"
            #         + "\n".join(expired)
            #         + "\n\nExporte-os novamente do Salesforce Business Manager."
            #     )
            #     return result

            # 1. Excel
            self._prog(10, "Lendo planilhas Excel…")
            excel_prices, excel_lists, brands_found, has_nat, has_avn = self._parse_excels(excel_paths)
            result.brands_found = brands_found
            result.total_excel_skus = len(excel_prices)

            if not excel_prices and not excel_lists:
                result.error = "Nenhum produto encontrado. Verifique a aba 'GRADE DE ATIVAÇÃO'."
                return result

            # Regra de Ouro (Gold Rule) V11.6
            cat_brands = set()
            for p in cat_paths:
                c_brand = self._get_catalog_brand_quick(p)
                cat_brands.add(c_brand)

            missing_xmls = []
            if has_nat and "Natura" not in cat_brands: missing_xmls.append("XML Natura")
            if has_avn and "Avon" not in cat_brands:   missing_xmls.append("XML Avon")
            if "ML" not in cat_brands:                 missing_xmls.append("XML Minha Loja (cbbrazil)")

            if missing_xmls:
                result.preflight_error = (
                    "Estrutura de Catálogos Incompleta\n\n"
                    "Para os Excels carregados, faltam os seguintes XMLs:\n\n• "
                    + "\n• ".join(missing_xmls)
                    + "\n\nAnexe-os para garantir a precisão do cruzamento Double-Blind."
                )
                return result

            # 2. Pricebook
            self._prog(35, "Descompactando Pricebook XML…")
            prices_xml = self._parse_pricebook(pb_path)

            # 3. Catálogos
            self._prog(60, "Varrendo Catálogos XML…")
            (online_status, searchable_status, technical_skus, xml_lists,
             prohibited_state, cat_missing_primary, bundles,
             variation_bases, category_assignments_map, ml_job_rules) = self._parse_catalogs(cat_paths)

            # 4. JOB errors (ML vs marca-mãe)
            self._prog(78, "Verificando sync de Jobs ML…")
            job_errors = self._calc_job_errors(category_assignments_map, ml_job_rules)

            # 5. Cruzamento analítico
            self._prog(85, "Cruzamento analítico — aplicando 12 regras…")
            errors, stats = self._cross_validate(
                excel_prices, excel_lists, prices_xml,
                online_status, searchable_status, technical_skus,
                xml_lists, prohibited_state, cat_missing_primary,
                bundles, variation_bases, job_errors,
                has_nat, has_avn,
            )
            result.errors = errors
            result.stats = stats

            self._prog(100, "Auditoria concluída!")
        except Exception as exc:
            result.error = str(exc)
        return result

    # ── Parsing Excel ─────────────────────────────────────────────────────
    def _parse_excels(self, paths):
        """
        Retorna:
          excel_prices : {sku: {DE, POR, VISIBLE}}
          excel_lists  : {list_id: set(skus)}   (LISTA_01 para Natura, lista-01 para Avon)
          brands_found : lista de marcas
          has_nat, has_avn : bool
        """
        excel_prices = {}
        excel_lists  = {}
        has_nat = has_avn = False

        for path in paths:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Detecta marca do arquivo
            file_brand = self._detect_brand_workbook(wb)
            if file_brand == "Natura": has_nat = True
            if file_brand == "Avon":   has_avn = True

            # Grade de Ativação → preços e visibilidade
            grade = self._find_grade_sheet(wb)
            if grade:
                self._parse_grade(grade, file_brand, excel_prices)

            # Listas LISTA_XX / lista-XX
            for name in wb.sheetnames:
                m = re.match(r"(?i)lista[-_\s]*0*(\d+)", name)
                if m:
                    num = m.group(1).zfill(2)
                    ws  = wb[name]
                    # Natura → LISTA_XX, Avon → lista-XX
                    list_key = f"LISTA_{num}" if file_brand == "Natura" else f"lista-{num}"
                    self._parse_lista(ws, file_brand, list_key, excel_lists)

            wb.close()

        brands = (["Natura"] if has_nat else []) + (["Avon"] if has_avn else [])
        return excel_prices, excel_lists, brands, has_nat, has_avn

    def _detect_brand_workbook(self, wb) -> str:
        """Varre toda a aba GRADE ou a primeira disponível e conta NATBRA/AVNBRA (Robust like JS)."""
        sheet = self._find_grade_sheet(wb) or (wb[wb.sheetnames[0]] if wb.sheetnames else None)
        if not sheet:
            return "Desconhecida"
        
        nat = avn = 0
        # Tenta ler um volume maior de dados para detecção precisa
        for row in sheet.iter_rows(max_row=500, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                v = str(cell).upper()
                nat += v.count("NATBRA-")
                avn += v.count("AVNBRA-")
        
        if nat == 0 and avn == 0:
            return "Desconhecida"
        return "Natura" if nat >= avn else "Avon"

    def _get_catalog_brand_quick(self, path: str) -> str:
        """Lê o início do XML para detectar a marca sem parsing total (Gold Rule)."""
        try:
            with open(path, "r", encoding="utf-8") as f:
                head = f.read(8192).lower()
                m = re.search(r'catalog-id="([^"]+)"', head)
                if m:
                    cid = m.group(1)
                    if any(x in cid for x in ["cb-br", "cbbrazil", "cbcom"]): return "ML"
                    if "natura" in cid: return "Natura"
                    if "avon" in cid: return "Avon"
        except:
            pass
        return "Desconhecida"

    def _find_grade_sheet(self, wb):
        for name in wb.sheetnames:
            n = name.upper()
            if "GRADE" in n or "ATIVA" in n:
                return wb[name]
        return None

    def _parse_grade(self, ws, file_brand: str, out: dict) -> None:
        rows = list(ws.iter_rows(max_row=10000, values_only=True))
        sku_col = de_col = por_col = vis_col = None
        sku_start = None

        for i, row in enumerate(rows[:60]):
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                v = str(cell).strip()
                vu = v.upper()
                if vu == "DE":
                    de_col = j
                elif vu == "POR":
                    por_col = j
                elif "VISIBLE" in vu or "VISIBILIDADE" in vu:
                    vis_col = j
                if sku_col is None and SKU_RE.match(v):
                    sku_col = j
                    sku_start = i

        if sku_col is None:
            return

        empty = 0
        for row in rows[sku_start:]:
            raw = row[sku_col] if sku_col < len(row) else None
            if not raw or not SKU_RE.match(str(raw).strip()):
                empty += 1
                if empty >= 50:
                    break
                continue
            empty = 0
            sku = str(raw).strip().upper()

            # Filtro de marca cruzada na leitura (igual ao JS)
            if file_brand == "Natura" and sku.startswith("AVNBRA-"):
                continue
            if file_brand == "Avon" and sku.startswith("NATBRA-"):
                continue

            de  = self._f(row[de_col]  if de_col  is not None and de_col  < len(row) else None)
            por = self._f(row[por_col] if por_col is not None and por_col < len(row) else None)
            vis_raw = row[vis_col] if vis_col is not None and vis_col < len(row) else None
            vis = str(vis_raw).strip().upper() if vis_raw else ""

            out[sku] = {"DE": de or 0.0, "POR": por or 0.0, "VISIBLE": vis}

    def _parse_lista(self, ws, file_brand: str, list_key: str, out: dict) -> None:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not cell or not SKU_RE.match(str(cell).strip()):
                    continue
                sku = str(cell).strip().upper()
                if file_brand == "Natura" and sku.startswith("AVNBRA-"):
                    continue
                if file_brand == "Avon" and sku.startswith("NATBRA-"):
                    continue
                out.setdefault(list_key, set()).add(sku)

    # ── Parsing Pricebook ─────────────────────────────────────────────────
    def _parse_pricebook(self, path: str) -> dict:
        """
        Retorna {sku: {Natura: {DE, POR}, Avon: {DE, POR}, ML: {DE, POR}}}
        Classificação por substring (igual ao JS), não por ID exato.
        """
        try:
            tree = etree.parse(path)
        except etree.XMLSyntaxError as exc:
            raise ValueError(f"Pricebook XML inválido: {exc}") from exc

        ns = {"pb": PRICEBOOK_NS}
        prices: dict = {}

        for pb_el in tree.findall(".//pb:pricebook", ns):
            header = pb_el.find("pb:header", ns)
            if header is None:
                continue
            pb_id = (header.get("pricebook-id") or "").lower()

            # Classificação por substring (igual ao JS)
            if "cb-br" in pb_id or "cbbrazil" in pb_id or "cbcom" in pb_id:
                pb_brand = "ML"
            elif "natura" in pb_id and ("brazil" in pb_id or "-br" in pb_id):
                pb_brand = "Natura"
            elif "avon" in pb_id and ("brazil" in pb_id or "-br" in pb_id):
                pb_brand = "Avon"
            else:
                continue

            price_type = "POR" if "sale" in pb_id else "DE"

            for pt in pb_el.findall(".//pb:price-table", ns):
                sku = pt.get("product-id", "").upper()
                amt_el = pt.find("pb:amount[@quantity='1']", ns)
                if amt_el is None or not amt_el.text:
                    continue
                try:
                    amt = float(amt_el.text)
                except ValueError:
                    continue

                prices.setdefault(sku, {"Natura": {}, "Avon": {}, "ML": {}})
                prices[sku][pb_brand][price_type] = amt

        return prices

    # ── Parsing Catálogos ─────────────────────────────────────────────────
    def _parse_catalogs(self, paths: list[str]):
        """
        Retorna:
          online_status            : {sku: bool}
          searchable_status        : {sku: bool}
          technical_skus           : {sku: bool}  (nomes em CAIXA ALTA = incompleto)
          xml_lists                : {list_id: set(skus)}   só Natura/Avon, não ML
          prohibited_state         : {brand: set(skus)}
          cat_missing_primary      : {sku: [brands]}
          bundles                  : {sku: [comp_skus]}
          variation_bases          : {sku: bool}
          category_assignments_map : {brand: {cat_id: set(skus)}}
          ml_job_rules             : list[dict]
        """
        online_status:   dict[str, bool]        = {}
        searchable_status: dict[str, bool]      = {}
        technical_skus:  dict[str, bool]        = {}
        xml_lists:       dict[str, set]         = {}
        prohibited_state = {"Natura": set(), "Avon": set(), "ML": set()}
        cat_missing_primary: dict[str, list]    = {}
        bundles:         dict[str, list]        = {}
        variation_bases: dict[str, bool]        = {}
        category_assignments_map = {"Natura": {}, "Avon": {}, "ML": {}}
        ml_job_rules:    list[dict]             = []

        for path in paths:
            try:
                tree = etree.parse(path)
            except etree.XMLSyntaxError as exc:
                raise ValueError(f"Catálogo XML inválido ({Path(path).name}): {exc}") from exc

            ns   = {"c": CATALOG_NS}
            root = tree.getroot()
            cat_id_str = (root.get("catalog-id") or "").lower()

            # Classificação do catálogo por substring
            if "cb-br" in cat_id_str or "cbbrazil" in cat_id_str or "cbcom" in cat_id_str:
                brand_cat = "ML"
            elif "natura" in cat_id_str:
                brand_cat = "Natura"
            elif "avon" in cat_id_str:
                brand_cat = "Avon"
            else:
                brand_cat = "Desconhecido"

            skus_in_file:  set[str] = set()
            primary_skus:  set[str] = set()

            # Regras de Job (Mirroring) - Só no catálogo ML
            if brand_cat == "ML":
                for cat in root.findall(".//c:category", ns):
                    ml_cid = cat.get("category-id")
                    if not ml_cid: continue
                    rules = cat.findall("c:category-assignment-rule", ns)
                    for rule in rules:
                        conds = rule.findall("c:category-condition", ns)
                        for cond in conds:
                            mother_cid = cond.get("category-id")
                            if mother_cid:
                                ml_job_rules.append({"mlCatId": ml_cid, "motherCatId": mother_cid})
                                break

            # Produtos
            for prod in root.findall(".//c:product", ns):
                sku = (prod.get("product-id") or "").upper()
                if not sku:
                    continue
                
                skus_in_file.add(sku)

                # online-flag (true prevalece)
                o = prod.find("c:online-flag", ns)
                if o is not None:
                    flag = (o.text or "").lower() == "true"
                    if flag:
                        online_status[sku] = True
                    elif online_status.get(sku) is not True:
                        online_status[sku] = False

                # searchable-flag
                s = prod.find("c:searchable-flag", ns)
                if s is not None:
                    flag = (s.text or "").lower() == "true"
                    if flag:
                        searchable_status[sku] = True
                    elif searchable_status.get(sku) is not True:
                        searchable_status[sku] = False

                # Detecta SKU técnico (nome todo em CAIXA ALTA) - v10.4 logic parity
                names_to_test = []
                for tname in ["display-name", "name", "friendly-name"]:
                    for name_el in prod.findall(f".//c:{tname}", ns):
                        if name_el.text: names_to_test.append(name_el.text.strip())
                
                if any(n == n.upper() and re.search(r"[A-Z]", n) for n in names_to_test):
                    technical_skus[sku] = True

                # Variation base product
                var_marker = (prod.get("variation-base-product") or
                              prod.get("is-variation-base") or "").lower()
                prod_type = (prod.get("type") or "").lower()
                has_variants = len(prod.findall("c:variants/c:variant", ns)) > 0
                var_flag_el = prod.find("c:variation-base-product", ns)
                if (var_marker == "true" or "variation" in prod_type or has_variants
                        or (var_flag_el is not None and (var_flag_el.text or "").lower() == "true")):
                    variation_bases[sku] = True

                # Bundles
                bp_el = prod.find("c:bundled-products", ns)
                if bp_el is not None:
                    comps = [c.get("product-id", "").upper()
                             for c in bp_el.findall("c:bundled-product", ns)
                             if c.get("product-id")]
                    if comps:
                        bundles[sku] = comps

            # Category-assignments
            for asgn in root.findall("c:category-assignment", ns):
                sku    = (asgn.get("product-id") or "").upper()
                cat_id = (asgn.get("category-id") or "")
                if not sku or not cat_id:
                    continue

                # Mapa para o check de JOB
                category_assignments_map[brand_cat].setdefault(cat_id, set()).add(sku)

                # Categorias proibidas (Conflito de Margem)
                if brand_cat in PROHIBITED_CATEGORIES and cat_id in PROHIBITED_CATEGORIES[brand_cat]:
                    prohibited_state[brand_cat].add(sku)

                # Listas de vitrine (só Natura/Avon, não ML)
                if brand_cat != "ML" and re.match(r"(?i)^(LISTA_|lista-)\d+", cat_id):
                    xml_lists.setdefault(cat_id, set()).add(sku)

                # primary-flag
                pf = asgn.find("c:primary-flag", ns)
                # Às vezes é atributo, não elemento
                pf_attr = asgn.get("primary-flag", "")
                if (pf is not None and (pf.text or "").lower() == "true") or pf_attr.lower() == "true":
                    primary_skus.add(sku)

            # SKUs sem categoria primária (Audit Rule: todo SKU no catálogo deve ter category-primary)
            for sku in skus_in_file:
                # Ignoramos SKUs técnicos (geralmente não precisam de navegação/SEO)
                if technical_skus.get(sku):
                    continue
                if sku not in primary_skus:
                    cat_missing_primary.setdefault(sku, []).append(brand_cat)

        return (online_status, searchable_status, technical_skus, xml_lists,
                prohibited_state, cat_missing_primary, bundles, variation_bases,
                category_assignments_map, ml_job_rules)

    # ── JOB errors ────────────────────────────────────────────────────────
    def _calc_job_errors(self, cam: dict, ml_job_rules: list[dict]) -> dict[str, list[str]]:
        """Compara categorias ML com marcas-mãe seguindo as regras de mirror do XML."""
        job_errors: dict[str, list[str]] = {}
        ml_cats = cam.get("ML", {})
        nat_cats = cam.get("Natura", {})
        avn_cats = cam.get("Avon", {})

        for rule in ml_job_rules:
            ml_cid = rule["mlCatId"]
            mother_cid = rule["motherCatId"]
            
            ml_skus = ml_cats.get(ml_cid, set())
            # Tenta achar a categoria na Natura ou Avon
            mother_skus = nat_cats.get(mother_cid) or avn_cats.get(mother_cid)
            
            if mother_skus is None:
                continue
                
            diff = ml_skus.symmetric_difference(mother_skus)
            for sku in diff:
                job_errors.setdefault(sku, []).append(
                    f"JOB NÃO RODOU (A Categoria ML {ml_cid} divergiu do espelho da Marca Mãe {mother_cid})"
                )
        return job_errors

    # ── Cruzamento analítico ──────────────────────────────────────────────
    def _cross_validate(
        self,
        excel_prices, excel_lists, prices_xml,
        online_status, searchable_status, technical_skus,
        xml_lists, prohibited_state, cat_missing_primary,
        bundles, variation_bases, job_errors,
        has_nat: bool, has_avn: bool,
    ):
        # Universo de SKUs = união de tudo (igual ao JS)
        all_skus: set[str] = set()
        all_skus.update(prices_xml.keys())
        all_skus.update(excel_prices.keys())
        all_skus.update(cat_missing_primary.keys())
        all_skus.update(bundles.keys())
        all_skus.update(job_errors.keys())
        for skus in xml_lists.values():
            all_skus.update(skus)

        errors: dict[str, list[dict]] = {k: [] for k in ERROR_META}
        stats = {k: {"total": 0, "natura": 0, "avon": 0} for k in ERROR_META}

        def bump(code, brand):
            stats[code]["total"] += 1
            if brand == "Natura":
                stats[code]["natura"] += 1
            else:
                stats[code]["avon"] += 1

        for sku in all_skus:
            if not SKU_RE.match(sku):
                continue

            brand = "Natura" if sku.startswith("NATBRA-") else "Avon"
            pE = excel_prices.get(sku)
            is_offline      = online_status.get(sku, True) is not True
            is_on_grade     = pE is not None

            # Regra JS: se offline E não está no Excel, ignora completamente
            if is_offline and not is_on_grade:
                continue

            px     = (prices_xml.get(sku) or {}).get(brand, {})
            px_ml  = (prices_xml.get(sku) or {}).get("ML", {})
            px_nat = (prices_xml.get(sku) or {}).get("Natura", {})
            px_avn = (prices_xml.get(sku) or {}).get("Avon", {})

            px_de  = px.get("DE", 0) or 0
            px_por = px.get("POR", 0) or 0

            # ML com herança: se ML não tem preço próprio, herda da marca
            sf_ml_de  = px_ml["DE"]  if "DE"  in px_ml else px_de
            sf_ml_por = px_ml["POR"] if "POR" in px_ml else px_por

            row_base = {"sku": sku, "brand": brand}

            # ── Check #1: PRODUTO OFFLINE (estava no Excel) ───────────────
            if is_offline and is_on_grade:
                errors["offline"].append({**row_base, "detail": "PRODUTO OFFLINE (Ação Comercial Exigida)"})
                bump("offline", brand)

            if not is_offline:
                # ── Check #2: BUNDLE QUEBRADO ─────────────────────────────
                if sku in bundles:
                    offline_comps = []
                    missing_price_comps = []
                    for comp in bundles[sku]:
                        if variation_bases.get(comp):
                            continue
                        if online_status.get(comp, True) is not True:
                            offline_comps.append(comp)
                        comp_px = (prices_xml.get(comp) or {}).get(brand, {})
                        if (comp_px.get("DE") or 0) <= 0 or (comp_px.get("POR") or 0) <= 0:
                            missing_price_comps.append(comp)

                    if offline_comps or missing_price_comps:
                        parts = []
                        if offline_comps:
                            parts.append(f"Comp. Offline: {', '.join(offline_comps)}")
                        if missing_price_comps:
                            parts.append(f"Comp. sem preço: {', '.join(missing_price_comps)}")
                        errors["bundle"].append({**row_base, "detail": f"BUNDLE QUEBRADO ({' ; '.join(parts)})"})
                        bump("bundle", brand)

                # ── Check #3: CROSS-BRAND ─────────────────────────────────
                has_nat_price = bool(px_nat.get("DE") or px_nat.get("POR"))
                has_avn_price = bool(px_avn.get("DE") or px_avn.get("POR"))
                if brand == "Natura" and has_avn_price:
                    errors["cross"].append({**row_base, "detail": "MARCA CRUZADA (NAT no Pricebook AVN)"})
                    bump("cross", brand)
                if brand == "Avon" and has_nat_price:
                    errors["cross"].append({**row_base, "detail": "MARCA CRUZADA (AVN no Pricebook NAT)"})
                    bump("cross", brand)

                # Os checks seguintes só correm se temos Excel para esta marca
                should_compare = (brand == "Natura" and has_nat) or (brand == "Avon" and has_avn)
                if should_compare:
                    if pE:
                        e_de  = pE.get("DE", 0) or 0
                        e_por = pE.get("POR", 0) or 0

                        # ── Check #4: PREÇO AUSENTE NO SF ────────────────
                        if not px_de and not px_por:
                            errors["price"].append({**row_base, "de_excel": e_de, "de_sf": 0,
                                                    "por_excel": e_por, "por_sf": 0,
                                                    "detail": "FALTA NO SF (PREÇO)"})
                            bump("price", brand)
                        else:
                            # ── Check #5: DIVERGÊNCIA DE PREÇO ───────────
                            if e_de > 0 and abs(e_de - px_de) > 0.01:
                                errors["price"].append({**row_base, "de_excel": e_de, "de_sf": px_de,
                                                        "por_excel": e_por, "por_sf": px_por,
                                                        "detail": f"DIVERGE SF (DE: R${e_de:.2f} vs R${px_de:.2f})"})
                                bump("price", brand)
                            if e_por > 0 and abs(e_por - px_por) > 0.01:
                                errors["price"].append({**row_base, "de_excel": e_de, "de_sf": px_de,
                                                        "por_excel": e_por, "por_sf": px_por,
                                                        "detail": f"DIVERGE SF (POR: R${e_por:.2f} vs R${px_por:.2f})"})
                                bump("price", brand)

                        # ── Check #6: SEARCHABLE vs VISIBLE ──────────────
                        vis = pE.get("VISIBLE", "")
                        is_visible_excel = (vis == "SIM")
                        is_hidden_excel  = (vis in ("NÃO", "NAO"))
                        is_searchable_sf = searchable_status.get(sku) is True
                        if not technical_skus.get(sku):
                            if is_visible_excel and not is_searchable_sf:
                                errors["searchable"].append({**row_base,
                                    "detail": "DIVERGE SEARCHABLE (Excel SIM vs SF false)"})
                                bump("searchable", brand)
                            elif is_hidden_excel and is_searchable_sf:
                                errors["searchable"].append({**row_base,
                                    "detail": "DIVERGE SEARCHABLE (Excel NÃO vs SF true)"})
                                bump("searchable", brand)

                        # ── Check #7: LISTAS DE VITRINE ───────────────────
                        for list_id, ex_skus in excel_lists.items():
                            list_is_mine = (
                                (brand == "Natura" and list_id.startswith("LISTA_")) or
                                (brand == "Avon"   and list_id.startswith("lista-"))
                            )
                            if not list_is_mine:
                                continue
                            if sku not in ex_skus:
                                continue
                            if list_id not in xml_lists:
                                errors["list"].append({**row_base,
                                    "detail": f"LISTA INEXISTENTE NO SF ({list_id})"})
                                bump("list", brand)
                            elif sku not in xml_lists[list_id]:
                                errors["list"].append({**row_base,
                                    "detail": f"FALTA NO SF ({list_id})"})
                                bump("list", brand)

                # Checks que dependem apenas de preços SF (sem obrigar Excel)
                if px_de or px_por or sf_ml_de or sf_ml_por or sku in cat_missing_primary:
                    # ── Check #8: FALTA PREÇO DE / POR ───────────────────
                    if px_por == 0 and px_de != 0:
                        errors["missing"].append({**row_base, "ausente": "POR",
                                                  "detail": "FALTA PREÇO POR"})
                        bump("missing", brand)
                    if px_de == 0 and px_por != 0:
                        errors["missing"].append({**row_base, "ausente": "DE",
                                                  "detail": "FALTA PREÇO DE"})
                        bump("missing", brand)

                    # ── Check #9: POR > DE ────────────────────────────────
                    if px_por > px_de > 0:
                        errors["logic"].append({**row_base, "de": px_de, "por": px_por,
                                                "detail": f"POR > DE (R${px_por:.2f} > R${px_de:.2f})"})
                        bump("logic", brand)

                    # ── Check #10: CONFLITO DE MARGEM ─────────────────────
                    is_promo_brand = px_por < px_de and px_por > 0
                    is_promo_ml    = sf_ml_por < sf_ml_de and sf_ml_por > 0
                    if is_promo_brand and sku in prohibited_state.get(brand, set()):
                        errors["margin"].append({**row_base,
                            "detail": f"CONFLITO PROG ({brand})"})
                        bump("margin", brand)
                    if is_promo_ml and sku in prohibited_state.get("ML", set()):
                        errors["margin"].append({**row_base,
                            "detail": "CONFLITO PROG (Minha Loja)"})
                        bump("margin", brand)

                    # ── Check #11: DIVERGÊNCIA ML ─────────────────────────
                    if (px_por > 0 or sf_ml_por > 0) and abs(px_por - sf_ml_por) > 0.01:
                        errors["ml"].append({**row_base,
                            "detail": f"DIVERGE ML (POR: Marca R${px_por:.2f} vs ML R${sf_ml_por:.2f})"})
                        bump("ml", brand)
                    if (px_de > 0 or sf_ml_de > 0) and abs(px_de - sf_ml_de) > 0.01:
                        errors["ml"].append({**row_base,
                            "detail": f"DIVERGE ML (DE: Marca R${px_de:.2f} vs ML R${sf_ml_de:.2f})"})
                        bump("ml", brand)

                    # ── Check #12: CATEGORIA PRIMÁRIA ─────────────────────
                    for cat_brand in (cat_missing_primary.get(sku) or []):
                        errors["primary"].append({**row_base,
                            "detail": f"FALTA CAT PRIMÁRIA ({cat_brand})"})
                        bump("primary", brand)

                # ── Check JOB ─────────────────────────────────────────────
                for msg in (job_errors.get(sku) or []):
                    errors["job"].append({**row_base, "detail": msg})
                    bump("job", brand)

        # Converte listas em DataFrames
        error_dfs = {code: pd.DataFrame(rows) if rows else pd.DataFrame()
                     for code, rows in errors.items()}

        total_stats = {
            "total": sum(s["total"] for s in stats.values()),
            "by_type": stats,
            "by_brand": {
                "natura": sum(s["natura"] for s in stats.values()),
                "avon":   sum(s["avon"]   for s in stats.values()),
            },
        }
        return error_dfs, total_stats

    # ── Helper ────────────────────────────────────────────────────────────
    @staticmethod
    def _f(val) -> Optional[float]:
        if val is None:
            return None
        try:
            return float(str(val).replace(",", ".").replace("R$", "").strip())
        except (ValueError, TypeError):
            return None
