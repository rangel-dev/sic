"""BRD-010 — Auditor: LISTA_XX visível (mesmo vazia) registra chave;
oculta/ausente não. Espelha tests/test_sync_engine_parse_excel.py."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from src.core.auditor_engine import AuditorEngine


def _make_workbook(tmp_path: Path, brand_marker: str, sheets: list[tuple[str, list[list], bool]]) -> Path:
    """brand_marker: célula colocada numa aba GRADE DE ATIVAÇÃO só para a
    detecção de marca (_detect_brand_workbook) funcionar de forma estável,
    independente do conteúdo das abas LISTA_XX testadas.
    sheets: lista de (nome_aba, linhas, oculta)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    grade = wb.create_sheet("GRADE DE ATIVAÇÃO")
    grade.append([brand_marker])
    for name, rows, hidden in sheets:
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
        if hidden:
            ws.sheet_state = "hidden"
    path = tmp_path / "grade.xlsx"
    wb.save(path)
    return path


class TestParseExcelsListaVaziaEOcultaBRD010:
    def test_lista_natura_visivel_vazia_registra_chave_com_set_vazio(self, tmp_path):
        path = _make_workbook(tmp_path, "NATBRA-000", [("LISTA_07", [["SKU"]], False)])
        _, excel_lists, *_ = AuditorEngine()._parse_excels([str(path)])
        assert "LISTA_07" in excel_lists
        assert excel_lists["LISTA_07"] == set()

    def test_lista_avon_visivel_vazia_registra_chave_com_set_vazio(self, tmp_path):
        path = _make_workbook(tmp_path, "AVNBRA-000", [("lista-07", [["SKU"]], False)])
        _, excel_lists, *_ = AuditorEngine()._parse_excels([str(path)])
        assert "lista-07" in excel_lists
        assert excel_lists["lista-07"] == set()

    def test_lista_oculta_com_skus_nao_gera_chave(self, tmp_path):
        path = _make_workbook(tmp_path, "NATBRA-000", [("LISTA_02", [["NATBRA-001"]], True)])
        _, excel_lists, *_ = AuditorEngine()._parse_excels([str(path)])
        assert "LISTA_02" not in excel_lists

    def test_lista_com_skus_continua_funcionando(self, tmp_path):
        path = _make_workbook(tmp_path, "NATBRA-000", [
            ("LISTA_01", [["NATBRA-001"], ["NATBRA-002"]], False),
        ])
        _, excel_lists, *_ = AuditorEngine()._parse_excels([str(path)])
        assert excel_lists["LISTA_01"] == {"NATBRA-001", "NATBRA-002"}

    def test_lista_visivel_e_oculta_juntas_nao_interferem(self, tmp_path):
        path = _make_workbook(tmp_path, "NATBRA-000", [
            ("LISTA_01", [["NATBRA-001"]], False),
            ("LISTA_02", [["NATBRA-999"]], True),
        ])
        _, excel_lists, *_ = AuditorEngine()._parse_excels([str(path)])
        assert excel_lists["LISTA_01"] == {"NATBRA-001"}
        assert "LISTA_02" not in excel_lists
