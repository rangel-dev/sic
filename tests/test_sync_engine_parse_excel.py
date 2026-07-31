"""BRD-009 — LISTA_XX vazia gera remoção; abas ocultas são ignoradas."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from src.core.sync_engine import SyncEngine


def _make_workbook(tmp_path: Path, sheets: list[tuple[str, list[list], bool]]) -> Path:
    """sheets: lista de (nome_aba, linhas, oculta)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba padrão "Sheet"
    for name, rows, hidden in sheets:
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
        if hidden:
            ws.sheet_state = "hidden"
    if all(hidden for _, _, hidden in sheets):
        # openpyxl não permite salvar um workbook cuja ÚNICA aba está oculta
        # ("The only worksheet of a workbook cannot be hidden") — aba extra
        # visível e irrelevante (nome não bate GRADE/LISTA), só para o
        # arquivo ser salvável; não interfere nas asserções do teste.
        wb.create_sheet("_dummy_visible")
    path = tmp_path / "grade.xlsx"
    wb.save(path)
    return path


class TestParseExcelFilesListaVaziaEAbasOcultas:
    def test_lista_visivel_vazia_registra_chave_com_set_vazio(self, tmp_path):
        path = _make_workbook(tmp_path, [("LISTA_07", [["SKU"]], False)])
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert "LISTA_07" in excel_lists
        assert excel_lists["LISTA_07"] == set()

    def test_lista_visivel_com_skus_continua_funcionando(self, tmp_path):
        path = _make_workbook(tmp_path, [
            ("LISTA_01", [["NATBRA-001"], ["NATBRA-002"]], False),
        ])
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert excel_lists["LISTA_01"] == {"NATBRA-001", "NATBRA-002"}

    def test_lista_oculta_com_skus_nao_gera_chave(self, tmp_path):
        path = _make_workbook(tmp_path, [
            ("LISTA_02", [["NATBRA-001"]], True),
        ])
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert "LISTA_02" not in excel_lists

    def test_lista_veryhidden_tambem_e_ignorada(self, tmp_path):
        path = _make_workbook(tmp_path, [("LISTA_03", [["NATBRA-001"]], False)])
        wb = openpyxl.load_workbook(path)
        wb.create_sheet("_dummy_visible")  # evita "única aba não pode ficar oculta"
        wb["LISTA_03"].sheet_state = "veryHidden"
        wb.save(path)
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert "LISTA_03" not in excel_lists

    def test_grade_oculta_nao_captura_nenhum_dado(self, tmp_path):
        path = _make_workbook(tmp_path, [
            ("GRADE DE ATIVAÇÃO", [
                ["VISIBLE", "SELO", "POR", "CATEGORIA PLANEJAMENTO"],
                ["SIM", "", "30.00", "Presente"],
                ["NATBRA-001"],
            ], True),
        ])
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert grade_map == {}
        assert presente_cols_ok is False

    def test_lista_e_grade_ocultas_juntas_nao_interferem_em_abas_visiveis(self, tmp_path):
        """Uma aba oculta não deve afetar o processamento normal de outras
        abas visíveis no mesmo arquivo."""
        path = _make_workbook(tmp_path, [
            ("LISTA_01", [["NATBRA-001"]], False),
            ("LISTA_02", [["NATBRA-999"]], True),
        ])
        excel_lists, grade_map, brand, presente_cols_ok = SyncEngine()._parse_excel_files([str(path)])
        assert excel_lists["LISTA_01"] == {"NATBRA-001"}
        assert "LISTA_02" not in excel_lists
