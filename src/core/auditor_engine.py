"""
Auditor Engine – Double-Blind Price & Catalog Validation
Fiel à lógica do auditor.js v8.4 / V11.6.
"""
from __future__ import annotations

import re
import time
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import openpyxl
import pandas as pd
from lxml import etree

from src.core.auditor.integrity import verify_core_integrity
from src.core.auditor.parity_rules_v11 import execute_parity_rules

# ─── Namespaces ───────────────────────────────────────────────────────────────
PRICEBOOK_NS = "http://www.demandware.com/xml/impex/pricebook/2006-10-31"
CATALOG_NS   = "http://www.demandware.com/xml/impex/catalog/2006-10-31"

# ─── Regra dos 15 minutos (Auditor: arquivos SF devem ser RECENTES < 15 min) ─
MAX_FILE_AGE_SECONDS = 900

# ─── Categorias proibidas para promoção (Conflito de Margem) ─────────────────
PROHIBITED_CATEGORIES = {
    "Natura": {"promocao-da-semana", "LISTA_01", "monte-seu-kit", "LISTA_02"},
    "Avon":   {"desconto-progressivo", "lista-01"},
    "ML":     {"promocao-da-semana", "desconto-progressivo", "monte-seu-kit"},
}

# ─── Metadados de erro para a UI ─────────────────────────────────────────────
ERROR_META: dict[str, dict] = {
    "price":      {"title": "Divergência de Preço",         "impact": "Risco Financeiro Alto",       "icon": "💰", "desc": "Preço no Excel (DE/POR) não bate com o valor no Pricebook do Salesforce."},
    "list":       {"title": "Visibilidade em Listas",        "impact": "Baixa Conversão",             "icon": "📋", "desc": "SKU está na aba de lista (ex: LISTA_01) no Excel, mas não está na categoria equivalente no XML."},
    "ml":         {"title": "Conflito Canal ML",             "impact": "Conflito Estratégico",        "icon": "⚡", "desc": "Preço no catálogo 'Minha Loja' diverge do preço da marca principal (Natura/Avon)."},
    "margin":     {"title": "Margem de Segurança",           "impact": "Perda Operacional (Crítico)", "icon": "🚨", "desc": "Produto em promoção (POR < DE) dentro de uma categoria onde descontos são proibidos."},
    "logic":      {"title": "Erro Lógico (POR > DE)",        "impact": "Erro de Cadastro",            "icon": "⚠️", "desc": "O preço promocional (POR) está maior que o preço de lista (DE)."},
    "missing":    {"title": "Preço Ausente (DE/POR)",        "impact": "Perda Imediata de Venda",     "icon": "❌", "desc": "O produto não possui preço de lista (DE) ou preço promocional (POR) definido no Salesforce."},
    "primary":    {"title": "Categoria Primária",            "impact": "Impacto SEO e Filtros",       "icon": "🏷️", "desc": "O produto está no catálogo mas não possui uma categoria marcada como 'Primária'."},
    "bundle":     {"title": "Saúde de Bundles/Kits",         "impact": "Queda no Ticket Médio",       "icon": "📦", "desc": "O Kit/Bundle possui componentes que estão offline ou sem preço definido."},
    "cross":      {"title": "Cross-Brand (Invasão)",         "impact": "Erro Crítico de Governança",  "icon": "🔴", "desc": "Produto de uma marca (ex: Natura) possui preço no catálogo de outra marca (ex: Avon)."},
    "offline":    {"title": "Produto Indisponível",          "impact": "Risco de Receita",            "icon": "🔇", "desc": "O produto está na Grade de Ativação do Excel, mas está com a flag 'online=false' no SF."},
    "job":        {"title": "Falha de Sync (ML JOB)",        "impact": "Catálogo Desatualizado",      "icon": "🔄", "desc": "As categorias da Minha Loja (ML) estão divergentes das categorias espelho da marca original."},
    "searchable": {"title": "Flag Searchable",               "impact": "Perda de Fluxo Orgânico",     "icon": "🔍", "desc": "Divergência entre a coluna VISIBLE do Excel e a flag 'searchable' do Salesforce."},
}

SKU_RE = re.compile(r"^(NAT|AVN)BRA-", re.IGNORECASE)


# ─── Resultado ────────────────────────────────────────────────────────────────
@dataclass
class AuditResult:
    errors: dict[str, pd.DataFrame] = field(default_factory=dict)
    stats: dict = field(default_factory=dict)
    brands_found: list[str] = field(default_factory=list)
    total_excel_skus: int = 0
    acertos: pd.DataFrame = field(default_factory=pd.DataFrame)
    error: Optional[str] = None
    preflight_error: Optional[str] = None
    integrity_error: bool = False


# ─── Engine ───────────────────────────────────────────────────────────────────
class AuditorEngine:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self._prog = progress_callback or (lambda p, m: None)

    # ── Entrada ───────────────────────────────────────────────────────────
    def run(self, excel_paths: list[str], pb_path: str, cat_paths: list[str]) -> AuditResult:
        result = AuditResult()
        try:
            # 0. Verificação de Lacre de Paridade
            if not verify_core_integrity():
                print("⚠️ [Integrity Check] O arquivo parity_rules_v11.py foi modificado, mas a execução prosseguirá.")

            # Trava 5: Pricebook e Catálogos devem ter sido exportados há menos de 15 min
            self._prog(3, "Verificando antiguidade dos arquivos SF…")
            expired = []
            for p in [pb_path] + cat_paths:
                try:
                    age = time.time() - os.path.getmtime(p)
                    if age > MAX_FILE_AGE_SECONDS:
                        expired.append(f"{Path(p).name} ({int(age/60)} min atrás)")
                except OSError:
                    pass
            if expired:
                result.preflight_error = (
                    "Arquivos SF desatualizados (>15 min):\n\n"
                    + "\n".join(expired)
                    + "\n\nExporte-os novamente do Salesforce Business Manager."
                )
                return result

            # 1. Excel (Opcional)
            self._prog(10, "Lendo planilhas Excel (se houver)…")
            excel_prices, excel_lists, excel_brands, has_nat, has_avn = self._parse_excels(excel_paths)
            result.total_excel_skus = len(excel_prices)

            # Regra de Ouro (Gold Rule) V11.6 & Detecção de Marcas
            cat_brands = set()
            for p in cat_paths:
                c_brand = self._get_catalog_brand_quick(p)
                cat_brands.add(c_brand)

            # Define as marcas encontradas no processo (Excel primeiro, depois XML)
            result.brands_found = list(set(excel_brands) | (cat_brands - {"ML", "Desconhecida"}))

            missing_xmls = []
            if has_nat and "Natura" not in cat_brands: missing_xmls.append("XML Natura")
            if has_avn and "Avon" not in cat_brands:   missing_xmls.append("XML Avon")
            if "ML" not in cat_brands:                 missing_xmls.append("XML Minha Loja (cbbrazil)")

            if missing_xmls:
                result.preflight_error = (
                    "Estrutura de Catálogos Incompleta\n\n"
                    "Para os Excels ou Catálogos carregados, faltam os seguintes XMLs:\n\n• "
                    + "\n• ".join(missing_xmls)
                    + "\n\nAnexe-os para garantir a precisão do cruzamento Double-Blind."
                )
                return result

            # 2. Pricebook
            self._prog(35, "Descompactando Pricebook XML…")
            prices_xml = self._parse_pricebook(pb_path)

            # Trava 4: Pricebook deve conter DE e POR para as 3 operações
            pb_coverage: dict[str, set] = {"Natura": set(), "Avon": set(), "ML": set()}
            for sku_data in prices_xml.values():
                for brand, type_map in sku_data.items():
                    pb_coverage[brand].update(type_map.keys())

            missing_pb = []
            brand_labels = {"Natura": "Natura", "Avon": "Avon", "ML": "Minha Loja"}
            for brand, label in brand_labels.items():
                for price_type in ("DE", "POR"):
                    if price_type not in pb_coverage[brand]:
                        missing_pb.append(f"{label}: sem Preço {price_type}")

            if missing_pb:
                result.preflight_error = (
                    "Dados Incompletos: O Pricebook não contém todos os preços necessários "
                    "para as 3 operações:\n\n• "
                    + "\n• ".join(missing_pb)
                    + "\n\nExporte o Pricebook completo do Salesforce Business Manager "
                    "com todos os pricebooks das operações Natura, Avon e Minha Loja."
                )
                return result

            # 3. Catálogos
            self._prog(60, "Varrendo Catálogos XML…")
            (online_status, searchable_status, technical_skus, xml_lists,
             prohibited_state, cat_missing_primary, bundles,
             variation_bases, category_assignments_map, ml_job_rules) = self._parse_catalogs(cat_paths)

            # 4. JOB errors (ML vs marca-mãe)
            self._prog(78, "Verificando sync de Jobs ML…")
            job_errors = self._calc_job_errors(category_assignments_map, ml_job_rules)

            # 5. Cruzamento analítico
            self._prog(85, "Cruzamento analítico — aplicando 12 regras…")
            errors, stats, acertos_df = self._cross_validate(
                excel_prices, excel_lists, prices_xml,
                online_status, searchable_status, technical_skus,
                xml_lists, prohibited_state, cat_missing_primary,
                bundles, variation_bases, job_errors,
                has_nat, has_avn,
            )
            result.errors   = errors
            result.stats    = stats
            result.acertos  = acertos_df

            self._prog(100, "Auditoria concluída!")
        except Exception as exc:
            result.error = str(exc)
        return result

    # ── Parsing Excel ─────────────────────────────────────────────────────
    def _parse_excels(self, paths):
        """
        Retorna:
          excel_prices : {sku: {DE, POR, VISIBLE}}
          excel_lists  : {list_id: set(skus)}   (LISTA_01 para Natura, lista-01 para Avon)
          brands_found : lista de marcas
          has_nat, has_avn : bool
        """
        excel_prices = {}
        excel_lists  = {}
        has_nat = has_avn = False

        for path in paths:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Detecta marca do arquivo
            file_brand = self._detect_brand_workbook(wb)
            if file_brand == "Natura": has_nat = True
            if file_brand == "Avon":   has_avn = True

            # Grade de Ativação → preços e visibilidade
            grade = self._find_grade_sheet(wb)
            if grade:
                self._parse_grade(grade, file_brand, excel_prices)

            # Listas LISTA_XX / lista-XX
            for name in wb.sheetnames:
                ws = wb[name]
                if ws.sheet_state != 'visible':
                    continue
                    
                m = re.match(r"(?i)lista[-_\s]*0*(\d+)", name)
                if m:
                    num = m.group(1).zfill(2)
                    # Natura → LISTA_XX, Avon → lista-XX
                    list_key = f"LISTA_{num}" if file_brand == "Natura" else f"lista-{num}"
                    self._parse_lista(ws, file_brand, list_key, excel_lists)

            wb.close()

        brands = (["Natura"] if has_nat else []) + (["Avon"] if has_avn else [])
        return excel_prices, excel_lists, brands, has_nat, has_avn

    def _detect_brand_workbook(self, wb) -> str:
        """Varre toda a aba GRADE ou a primeira disponível e conta NATBRA/AVNBRA (Robust like JS)."""
        sheet = self._find_grade_sheet(wb) or (wb[wb.sheetnames[0]] if wb.sheetnames else None)
        if not sheet:
            return "Desconhecida"
        
        nat = avn = 0
        # Tenta ler um volume maior de dados para detecção precisa
        for row in sheet.iter_rows(max_row=500, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                v = str(cell).upper()
                nat += v.count("NATBRA-")
                avn += v.count("AVNBRA-")
        
        if nat == 0 and avn == 0:
            return "Desconhecida"
        return "Natura" if nat >= avn else "Avon"

    def _get_catalog_brand_quick(self, path: str) -> str:
        """Lê o início do XML para detectar a marca sem parsing total (Gold Rule)."""
        try:
            with open(path, "r", encoding="utf-8") as f:
                head = f.read(8192).lower()
                m = re.search(r'catalog-id="([^"]+)"', head)
                if m:
                    cid = m.group(1)
                    if any(x in cid for x in ["cb-br", "cbbrazil", "cbcom"]): return "ML"
                    if "natura" in cid: return "Natura"
                    if "avon" in cid: return "Avon"
        except:
            pass
        return "Desconhecida"

    def _find_grade_sheet(self, wb):
        for name in wb.sheetnames:
            n = name.upper()
            if "GRADE" in n or "ATIVA" in n:
                if wb[name].sheet_state == 'visible':
                    return wb[name]
        return None

    def _parse_grade(self, ws, file_brand: str, out: dict) -> None:
        rows = list(ws.iter_rows(max_row=10000, values_only=True))
        sku_col = de_col = por_col = vis_col = None
        sku_start = None

        for i, row in enumerate(rows[:60]):
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                v = str(cell).strip()
                vu = v.upper()
                if vu == "DE":
                    de_col = j
                elif vu == "POR":
                    por_col = j
                elif "VISIBLE" in vu or "VISIBILIDADE" in vu:
                    vis_col = j
                if sku_col is None and SKU_RE.match(v):
                    sku_col = j
                    sku_start = i

        if sku_col is None:
            return

        empty = 0
        for row in rows[sku_start:]:
            raw = row[sku_col] if sku_col < len(row) else None
            if not raw or not SKU_RE.match(str(raw).strip()):
                empty += 1
                if empty >= 50:
                    break
                continue
            empty = 0
            sku = str(raw).strip().upper()

            # Filtro de marca cruzada na leitura (igual ao JS)
            if file_brand == "Natura" and sku.startswith("AVNBRA-"):
                continue
            if file_brand == "Avon" and sku.startswith("NATBRA-"):
                continue

            de  = self._f(row[de_col]  if de_col  is not None and de_col  < len(row) else None)
            por = self._f(row[por_col] if por_col is not None and por_col < len(row) else None)
            vis_raw = row[vis_col] if vis_col is not None and vis_col < len(row) else None
            vis = str(vis_raw).strip().upper() if vis_raw else ""

            out[sku] = {"DE": de or 0.0, "POR": por or 0.0, "VISIBLE": vis}

    def _parse_lista(self, ws, file_brand: str, list_key: str, out: dict) -> None:
        sku_col = -1
        rows = list(ws.iter_rows(max_row=10000, values_only=True))
        
        # Scanner Dinâmico (Legacy JS findSkuColumnOnly)
        for i, row in enumerate(rows[:50]):
            for j, cell in enumerate(row):
                if cell:
                    val = str(cell).strip().upper()
                    if val.startswith("NATBRA-") or val.startswith("AVNBRA-"):
                        sku_col = j
                        break
            if sku_col != -1:
                break
                
        if sku_col == -1:
            sku_col = 1 # Fallback para coluna B (index 1) - equivalente ao 2 no JS
            
        for row in rows:
            if sku_col < len(row):
                cell = row[sku_col]
                if not cell:
                    continue
                val = str(cell).strip().upper()
                if not SKU_RE.match(val):
                    continue
                if file_brand == "Natura" and val.startswith("AVNBRA-"):
                    continue
                if file_brand == "Avon" and val.startswith("NATBRA-"):
                    continue
                out.setdefault(list_key, set()).add(val)

    # ── Parsing Pricebook ─────────────────────────────────────────────────
    def _parse_pricebook(self, path: str) -> dict:
        """
        Retorna {sku: {Natura: {DE, POR}, Avon: {DE, POR}, ML: {DE, POR}}}
        Classificação por substring (igual ao JS), não por ID exato.
        """
        try:
            tree = etree.parse(path)
        except etree.XMLSyntaxError as exc:
            raise ValueError(f"Pricebook XML inválido: {exc}") from exc

        ns = {"pb": PRICEBOOK_NS}
        prices: dict = {}

        for pb_el in tree.findall(".//pb:pricebook", ns):
            header = pb_el.find("pb:header", ns)
            if header is None:
                continue
            pb_id = (header.get("pricebook-id") or "").lower()

            # Classificação por substring (igual ao JS)
            if "cb-br" in pb_id or "cbbrazil" in pb_id or "cbcom" in pb_id:
                pb_brand = "ML"
            elif "natura" in pb_id and ("brazil" in pb_id or "-br" in pb_id):
                pb_brand = "Natura"
            elif "avon" in pb_id and ("brazil" in pb_id or "-br" in pb_id):
                pb_brand = "Avon"
            else:
                continue

            # Identificação do Price Type = Legacy Rules Mismatch
            # Legado JS: pbId.includes("lista") || pbId.includes("list") ? "DE" : "POR";
            price_type = "DE" if "lista" in pb_id or "list" in pb_id else "POR"

            for pt in pb_el.findall(".//pb:price-table", ns):
                sku = pt.get("product-id", "").upper()
                amt_el = pt.find("pb:amount[@quantity='1']", ns)
                if amt_el is None or not amt_el.text:
                    continue
                try:
                    amt = float(amt_el.text)
                except ValueError:
                    continue

                prices.setdefault(sku, {"Natura": {}, "Avon": {}, "ML": {}})
                prices[sku][pb_brand][price_type] = amt

        return prices

    # ── Parsing Catálogos ─────────────────────────────────────────────────
    def _parse_catalogs(self, paths: list[str]):
        """
        Retorna:
          online_status            : {sku: bool}
          searchable_status        : {sku: bool}
          technical_skus           : {sku: bool}  (nomes em CAIXA ALTA = incompleto)
          xml_lists                : {list_id: set(skus)}   só Natura/Avon, não ML
          prohibited_state         : {brand: set(skus)}
          cat_missing_primary      : {sku: [brands]}
          bundles                  : {sku: [comp_skus]}
          variation_bases          : {sku: bool}
          category_assignments_map : {brand: {cat_id: set(skus)}}
          ml_job_rules             : list[dict]
        """
        online_status:   dict[str, bool]        = {}
        searchable_status: dict[str, bool]      = {}
        technical_skus:  dict[str, bool]        = {}
        xml_lists:       dict[str, set]         = {}
        prohibited_state = {"Natura": set(), "Avon": set(), "ML": set()}
        cat_missing_primary: dict[str, list]    = {}
        bundles:         dict[str, list]        = {}
        variation_bases: dict[str, bool]        = {}
        category_assignments_map = {"Natura": {}, "Avon": {}, "ML": {}}
        ml_job_rules:    list[dict]             = []

        for path in paths:
            try:
                tree = etree.parse(path)
            except etree.XMLSyntaxError as exc:
                raise ValueError(f"Catálogo XML inválido ({Path(path).name}): {exc}") from exc

            ns   = {"c": CATALOG_NS}
            root = tree.getroot()
            cat_id_str = (root.get("catalog-id") or "").lower()

            # Classificação do catálogo por substring
            if "cb-br" in cat_id_str or "cbbrazil" in cat_id_str or "cbcom" in cat_id_str:
                brand_cat = "ML"
            elif "natura" in cat_id_str:
                brand_cat = "Natura"
            elif "avon" in cat_id_str:
                brand_cat = "Avon"
            else:
                brand_cat = "Desconhecido"

            # Função helper espelho da getTagText (JS) - Busca profunda e ignorando strict namespaces
            def get_tag_text(node, tag: str) -> Optional[str]:
                for child in node.iter():
                    if not isinstance(child.tag, str):
                        continue
                    # Ignora a URL de namespace ou tag crua
                    if child.tag.endswith(f"}}{tag}") or child.tag == tag:
                        if child.text and child.text.strip():
                            return child.text.strip()
                return None

            skus_in_file:  set[str] = set()
            primary_skus:  set[str] = set()
            assigned_skus: set[str] = set()

            # Regras de Job (Mirroring) - Só no catálogo ML
            if brand_cat == "ML":
                for cat in root.findall(".//c:category", ns):
                    ml_cid = cat.get("category-id")
                    if not ml_cid: continue
                    rules = cat.findall("c:category-assignment-rule", ns)
                    for rule in rules:
                        conds = rule.findall("c:category-condition", ns)
                        for cond in conds:
                            mother_cid = cond.get("category-id")
                            if mother_cid:
                                ml_job_rules.append({"mlCatId": ml_cid, "motherCatId": mother_cid})
                                break

            # Produtos
            for prod in root.findall(".//c:product", ns):
                sku = (prod.get("product-id") or "").upper()
                if not sku:
                    continue
                
                skus_in_file.add(sku)

                # online-flag (true prevalece)
                o_flag = get_tag_text(prod, "online-flag")
                if o_flag is not None:
                    if o_flag.lower() == "true":
                        online_status[sku] = True
                    elif online_status.get(sku) is not True:
                        online_status[sku] = False

                # searchable-flag
                s_flag = get_tag_text(prod, "searchable-flag")
                if s_flag is not None:
                    if s_flag.lower() == "true":
                        searchable_status[sku] = True
                    elif searchable_status.get(sku) is not True:
                        searchable_status[sku] = False

                # Detecta SKU técnico (nome todo em CAIXA ALTA) - v10.4 logic parity
                names_to_test = []
                for tname in ["display-name", "name"]:
                    for name_el in prod.findall(f".//c:{tname}", ns):
                        if name_el.text: names_to_test.append(name_el.text.strip())
                
                f_name = get_tag_text(prod, "friendly-name")
                if f_name:
                    names_to_test.append(f_name)
                
                if any(n == n.upper() and re.search(r"[A-Z]", n) for n in names_to_test):
                    technical_skus[sku] = True

                # Variation base product
                var_marker = (prod.get("variation-base-product") or
                              prod.get("is-variation-base") or "").lower()
                
                prod_type_attr = (prod.get("type") or "").lower()
                prod_type_el = (get_tag_text(prod, "product-type") or "").lower()
                prod_type = prod_type_attr if prod_type_attr else prod_type_el
                
                has_variants = len(prod.findall("c:variants/c:variant", ns)) > 0
                
                var_flag_text = (get_tag_text(prod, "variation-base-product") or "").lower()
                is_var_flag_text = (get_tag_text(prod, "is-variation-base") or "").lower()
                
                if (var_marker == "true" or "variation" in prod_type or has_variants
                        or var_flag_text == "true" or is_var_flag_text == "true"):
                    variation_bases[sku] = True

                # Bundles
                bp_el = prod.find("c:bundled-products", ns)
                if bp_el is not None:
                    comps = [c.get("product-id", "").upper()
                             for c in bp_el.findall("c:bundled-product", ns)
                             if c.get("product-id")]
                    if comps:
                        bundles[sku] = comps

            # Category-assignments
            for asgn in root.findall(".//c:category-assignment", ns):
                sku    = (asgn.get("product-id") or "").upper()
                cat_id = (asgn.get("category-id") or "")
                if not sku or not cat_id:
                    continue
                
                assigned_skus.add(sku)

                # Mapa para o check de JOB
                category_assignments_map[brand_cat].setdefault(cat_id, set()).add(sku)

                # Categorias proibidas (Conflito de Margem)
                if brand_cat in PROHIBITED_CATEGORIES and cat_id in PROHIBITED_CATEGORIES[brand_cat]:
                    prohibited_state[brand_cat].add(sku)

                # Listas de vitrine (só Natura/Avon, não ML)
                if brand_cat != "ML" and re.match(r"(?i)^(LISTA_|lista-)\d+", cat_id):
                    xml_lists.setdefault(cat_id, set()).add(sku)

                # primary-flag
                pf = asgn.find("c:primary-flag", ns)
                # Às vezes é atributo, não elemento
                pf_attr = asgn.get("primary-flag", "")
                if (pf is not None and (pf.text or "").lower() == "true") or pf_attr.lower() == "true":
                    primary_skus.add(sku)

            # SKUs sem categoria primária (Audit Rule: todo SKU no catálogo deve ter category-primary)
            # Requisito do Legado JS: Apenas varre os SKUs que foram marcados com Assigned (assigned_skus), não todos os 'skus_in_file'
            for sku in assigned_skus:
                # Ignoramos SKUs técnicos (geralmente não precisam de navegação/SEO)
                if technical_skus.get(sku):
                    continue
                if sku not in primary_skus:
                    cat_missing_primary.setdefault(sku, []).append(brand_cat)

        return (online_status, searchable_status, technical_skus, xml_lists,
                prohibited_state, cat_missing_primary, bundles, variation_bases,
                category_assignments_map, ml_job_rules)

    # ── JOB errors ────────────────────────────────────────────────────────
    def _calc_job_errors(self, cam: dict, ml_job_rules: list[dict]) -> dict[str, list[str]]:
        """Compara categorias ML com marcas-mãe seguindo as regras de mirror do XML."""
        job_errors: dict[str, list[str]] = {}
        ml_cats = cam.get("ML", {})
        nat_cats = cam.get("Natura", {})
        avn_cats = cam.get("Avon", {})

        for rule in ml_job_rules:
            ml_cid = rule["mlCatId"]
            mother_cid = rule["motherCatId"]
            
            ml_skus = ml_cats.get(ml_cid, set())
            # Tenta achar a categoria na Natura ou Avon
            mother_skus = nat_cats.get(mother_cid) or avn_cats.get(mother_cid)
            
            if mother_skus is None:
                continue
                
            diff = ml_skus.symmetric_difference(mother_skus)
            for sku in diff:
                job_errors.setdefault(sku, []).append(
                    f"JOB NÃO RODOU (A Categoria ML {ml_cid} divergiu do espelho da Marca Mãe {mother_cid})"
                )
        return job_errors

    # ── Cruzamento analítico ──────────────────────────────────────────────
    def _cross_validate(
        self,
        excel_prices, excel_lists, prices_xml,
        online_status, searchable_status, technical_skus,
        xml_lists, prohibited_state, cat_missing_primary,
        bundles, variation_bases, job_errors,
        has_nat: bool, has_avn: bool,
    ):
        # Universo de SKUs = união de tudo (igual ao JS)
        all_skus: set[str] = set()
        all_skus.update(prices_xml.keys())
        all_skus.update(excel_prices.keys())
        all_skus.update(cat_missing_primary.keys())
        all_skus.update(bundles.keys())
        all_skus.update(job_errors.keys())
        for skus in xml_lists.values():
            all_skus.update(skus)

        errors: dict[str, list[dict]] = {k: [] for k in ERROR_META}
        stats = {k: {"total": 0, "natura": 0, "avon": 0} for k in ERROR_META}

        def bump(code, brand):
            stats[code]["total"] += 1
            if brand == "Natura":
                stats[code]["natura"] += 1
            else:
                stats[code]["avon"] += 1

        # Delega processamento para o arquivo lacrado
        execute_parity_rules(
            all_skus, excel_prices, online_status, prices_xml,
            bundles, variation_bases, searchable_status, technical_skus,
            excel_lists, xml_lists, cat_missing_primary, prohibited_state,
            job_errors, has_nat, has_avn, errors, bump
        )

        # Converte listas em DataFrames
        error_dfs = {code: pd.DataFrame(rows) if rows else pd.DataFrame()
                     for code, rows in errors.items()}

        # ─── RECONCILIAÇÃO LEGADA V11.6 (Deduplicação de SKUs) ────────────
        # O sistema em Javascript validava cada SKU apenas 1 vez por categoria no Dashboard,
        # gerando os números da UI via 'dynamicStats' agrupados por SKU_ID únicos.
        # Caso um Produto tenha 2 erros de Listas (ex LISTA_1, LISTA_2), ele contava como 1.
        dedup_stats = {k: {"total": 0, "natura": 0, "avon": 0} for k in ERROR_META}
        
        for code, rows in errors.items():
            unique_skus = {}
            for row in rows:
                sku = row["sku"]
                if sku not in unique_skus:
                    unique_skus[sku] = row["brand"]
            
            dedup_stats[code]["total"] = len(unique_skus)
            for brand_str in unique_skus.values():
                if brand_str == "Natura":
                    dedup_stats[code]["natura"] += 1
                else:
                    dedup_stats[code]["avon"] += 1

        total_stats = {
            "total": sum(s["total"] for s in dedup_stats.values()),
            "by_type": dedup_stats,
            "by_brand": {
                "natura": sum(s["natura"] for s in dedup_stats.values()),
                "avon":   sum(s["avon"]   for s in dedup_stats.values()),
            },
        }

        # ─── ACERTOS: SKUs que passaram em todas as verificações ──────────────
        skus_com_erro: set[str] = set()
        for rows in errors.values():
            for row in rows:
                skus_com_erro.add(row["sku"])

        acertos_rows = []
        for sku in sorted(all_skus):
            if not SKU_RE.match(sku):
                continue
            pE = excel_prices.get(sku)
            is_offline = online_status.get(sku) is not True
            if is_offline and pE is None:
                continue
            if sku in skus_com_erro:
                continue
            brand = "Natura" if sku.upper().startswith("NATBRA-") else "Avon"
            px = (prices_xml.get(sku) or {}).get(brand, {})
            acertos_rows.append({
                "sku":        sku,
                "brand":      brand,
                "de_sf":      px.get("DE", 0) or 0,
                "por_sf":     px.get("POR", 0) or 0,
                "online":     online_status.get(sku, False),
                "searchable": searchable_status.get(sku, False),
            })

        acertos_df = (
            pd.DataFrame(acertos_rows)
            if acertos_rows
            else pd.DataFrame(columns=["sku", "brand", "de_sf", "por_sf", "online", "searchable"])
        )

        return error_dfs, total_stats, acertos_df

    # ── Helper ────────────────────────────────────────────────────────────
    @staticmethod
    def _f(val) -> Optional[float]:
        if val is None:
            return None
        # Emula estritamente o comportamento do parseFloat() do JavaScript Legado.
        # Ele lê apenas a parte inicial que se parece com número, truncando na vírgula 
        # (causando o aumento e identificação do Falso Positivo/Erro de Typping em 10,50 vs 10.5).
        v_str = str(val).replace("R$", "").strip()
        m = re.match(r'^[+-]?\d+(?:\.\d+)?', v_str)
        if m:
            return float(m.group(0))
        return None
