"""
Cupom Engine – Lista de códigos → Salesforce Demandware Coupon XML
Porta fiel da lógica do SFCC Universal Coupon Generator v1.6.1 (HTML/JS).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from lxml import etree

COUPON_NS = "http://www.demandware.com/xml/impex/coupon/2008-06-17"

_SPECIAL_RE = re.compile(r"[^A-Za-z0-9_\-]")
_LOWER_RE   = re.compile(r"[a-z]")


@dataclass
class CupomResult:
    xml_content:  Optional[bytes]  = None
    log_workbook: Optional[object] = None   # openpyxl.Workbook | None
    stats:        dict             = field(default_factory=dict)
    error:        Optional[str]    = None


class CupomEngine:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self._progress = progress_callback or (lambda p, msg: None)

    # ── Ponto de entrada ──────────────────────────────────────────────────
    def run(
        self,
        coupon_id:   str,
        manual_text: str,
        xlsx_paths:  list[str],
    ) -> CupomResult:
        result = CupomResult()

        try:
            unique_codes: list[str]              = []
            seen:         dict[str, str]         = {}
            duplicates:   list[tuple[str, str]]  = []
            invalids:     list[tuple[str, str, str]] = []

            total_deleted   = 0
            total_corrected = 0

            def add_code(raw, source: str) -> None:
                nonlocal total_deleted, total_corrected
                if not raw:
                    return
                code = str(raw).strip()
                if not code:
                    return

                has_lower   = bool(_LOWER_RE.search(code))
                has_special = bool(_SPECIAL_RE.search(code))

                if has_special:
                    reason = "Caracteres especiais/acentos (DELETADO do XML)"
                    if has_lower:
                        reason = "Minúsculas e caracteres especiais (DELETADO do XML)"
                    invalids.append((code, source, reason))
                    total_deleted += 1
                    return

                if has_lower:
                    invalids.append((
                        code, source,
                        "Letras minúsculas (CORRIGIDO para maiúsculas no XML)",
                    ))
                    code = code.upper()
                    total_corrected += 1

                if code in seen:
                    duplicates.append((code, source))
                else:
                    seen[code] = source
                    unique_codes.append(code)

            self._progress(10, "Processando entrada manual…")
            if manual_text:
                for line in manual_text.splitlines():
                    add_code(line.strip(), "Manual")

            total_files = len(xlsx_paths)
            for i, path in enumerate(xlsx_paths):
                pct = 20 + int(i * (40 / max(total_files, 1)))
                self._progress(pct, f"Lendo {Path(path).name}…")
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(values_only=True):
                        if row:
                            add_code(row[0], sheet_name)
                wb.close()

            if not unique_codes:
                result.error = "Nenhum código válido encontrado após validação."
                return result

            self._progress(70, "Gerando XML SFCC…")
            result.xml_content = self._generate_xml(coupon_id, unique_codes)

            self._progress(85, "Gerando log de inconsistências…")
            if duplicates or invalids:
                result.log_workbook = self._generate_log(duplicates, invalids)

            result.stats = {
                "total":      len(unique_codes),
                "deleted":    total_deleted,
                "corrected":  total_corrected,
                "duplicates": len(duplicates),
            }

            self._progress(100, "Concluído!")

        except Exception as exc:
            import traceback
            traceback.print_exc()
            result.error = str(exc)

        return result

    # ── Geração de XML ────────────────────────────────────────────────────
    def _generate_xml(self, coupon_id: str, codes: list[str]) -> bytes:
        root = etree.Element("coupons", xmlns=COUPON_NS)

        coupon_el = etree.SubElement(root, "coupon")
        coupon_el.set("coupon-id", coupon_id)
        etree.SubElement(coupon_el, "enabled-flag").text = "true"
        etree.SubElement(coupon_el, "multiple-codes")

        codes_el = etree.SubElement(root, "coupon-codes")
        codes_el.set("coupon-id", coupon_id)
        for code in codes:
            etree.SubElement(codes_el, "code").text = code

        return etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", pretty_print=True
        )

    # ── Geração de log Excel ──────────────────────────────────────────────
    def _generate_log(
        self,
        duplicates: list[tuple[str, str]],
        invalids:   list[tuple[str, str, str]],
    ) -> object:
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        if duplicates:
            ws_dup = wb.create_sheet("Duplicatas")
            ws_dup.append(["Cupom Duplicado", "Aba / Fonte de Origem"])
            for code, source in duplicates:
                ws_dup.append([code, source])

        if invalids:
            ws_inv = wb.create_sheet("Cupons Inválidos")
            ws_inv.append([
                "Cupom Lido",
                "Aba / Fonte de Origem",
                "Motivo do Erro / Ação Tomada",
            ])
            for code, source, reason in invalids:
                ws_inv.append([code, source, reason])

        return wb
