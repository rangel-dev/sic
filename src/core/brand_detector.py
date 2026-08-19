"""
Brand Detector Utility – Smart Content Identification
Identifies if a file (XML/Excel) belongs to Natura, Avon, or CB (Minha Loja).
Supports multi-brand detection within a single file or across multiple files.
"""

import re
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Set


# ── Brazil ID guard (BRD — incidente Chile) ─────────────────────────────────
# Allowlist exata dos catalog-id Brasil, confirmada em produção (tela de
# Catálogos do Salesforce Business Manager). Usar match exato para catálogo:
# é dado real, mais seguro que qualquer heurística por substring.
BRAZIL_CATALOG_IDS: dict[str, str] = {
    "natura-br-storefront-catalog": "natura",
    "avon-br-storefront-catalog": "avon",
    "cb-br-storefront-catalog": "ml",
}

# Marcas Natura/Avon/CB fora do Brasil observadas na mesma instância
# Salesforce (grupo Natura&Co) — usadas só para popular `country_hint` na
# mensagem de erro, nunca para decidir aprovação/rejeição.
_COUNTRY_HINTS: dict[str, str] = {
    "cl": "Chile", "ar": "Argentina", "mx": "México",
    "co": "Colômbia", "pe": "Peru",
    "chile": "Chile", "argentina": "Argentina", "mexico": "México",
    "colombia": "Colômbia", "peru": "Peru",
}

_BRAND_MARKERS: dict[str, tuple[str, ...]] = {
    "natura": ("natura",),
    "avon": ("avon",),
    "ml": ("cb-br", "cbbrazil", "cbcom", "br-cb"),
}


@dataclass(frozen=True)
class BrazilIdCheck:
    """Resultado da checagem 'este ID pertence a uma loja Brasil?'."""
    raw_id: str
    brand: Optional[str]           # "natura" | "avon" | "ml" | None
    is_brazil: bool
    country_hint: Optional[str] = None  # ex. "Chile", best-effort

    @property
    def ok(self) -> bool:
        return self.brand is not None and self.is_brazil


class BrandDetector:
    @staticmethod
    def detect_single(file_path: str) -> set[str]:
        """
        Detects all brands within a SINGLE file.

        Parameters
        ----------
        file_path : str
            Path to a single file (XML or Excel)

        Returns
        -------
        set[str]
            Set of brand strings {"natura", "avon", "ml"} found in the file

        Example
        -------
        >>> BrandDetector.detect_single("/path/to/catalog_natura.xml")
        {"natura"}
        """
        if not file_path or not os.path.exists(file_path):
            return set()

        ext = Path(file_path).suffix.lower()

        if ext == ".xml":
            return BrandDetector._detect_xml_set(file_path)
        elif ext in [".xlsx", ".xlsm", ".xls"]:
            return BrandDetector._detect_excel_set(file_path)

        return set()

    @staticmethod
    def detect(file_paths: list[str]) -> set[str]:
        """
        Main entry point. Detects all brands across a list of files.

        Parameters
        ----------
        file_paths : list[str]
            List of file paths to scan

        Returns
        -------
        set[str]
            Union of all brands found across all files {"natura", "avon", "ml"}
        """
        all_brands = set()
        for path in file_paths:
            all_brands.update(BrandDetector.detect_single(path))
        return all_brands

    @staticmethod
    def _detect_xml_set(path: str) -> set[str]:
        """
        Detects brands in an XML file by reading its opening tag.

        Strategy:
        - Read first 1KB (enough to find the root element and its attributes)
        - If root tag has catalog-id  → it's a catalog file  → detect by catalog-id value
        - If root tag has pricebook-id → it's a pricebook file → detect by pricebook-id values
        """
        brands = set()
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                # 1KB is enough to read the opening tag regardless of file size
                header = f.read(1024).lower()

            # --- CATALOG detection ---
            # <catalog ... catalog-id="natura-br-storefront-catalog" ...>
            catalog_id_match = re.search(r'catalog-id=["\']([^"\']+)["\']', header)
            if catalog_id_match:
                cid = catalog_id_match.group(1)
                if "natura" in cid:
                    brands.add("natura")
                if "avon" in cid:
                    brands.add("avon")
                if "cb" in cid:
                    brands.add("ml")
                return brands  # catalog-id found → done, no need to keep searching

            # --- PRICEBOOK detection ---
            # Pricebook doesn't have catalog-id in root tag.
            # It has multiple <header pricebook-id="..."> entries — need full file scan.
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read().lower()

            for m in re.finditer(r'pricebook-id=["\']([^"\']+)["\']', content):
                pid = m.group(1)
                if "natura" in pid:
                    brands.add("natura")
                if "avon" in pid:
                    brands.add("avon")
                if "cb" in pid:
                    brands.add("ml")

        except Exception as e:
            import sys
            print(f"Warning: XML detection failed for {path}: {e}", file=sys.stderr)
        return brands

    @staticmethod
    def _detect_excel_set(path: str) -> set[str]:
        """Lightweight scan of an Excel file to find brand signatures."""
        brands = set()
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Detect by sheet names
            sheet_names_joined = " ".join(wb.sheetnames).upper()
            if "NATURA" in sheet_names_joined:
                brands.add("natura")
            if "AVON" in sheet_names_joined:
                brands.add("avon")

            # Detect by content peeking in the first visible sheet
            sheet = None
            for name in wb.sheetnames:
                try:
                    if wb[name].sheet_state == "visible":
                        sheet = wb[name]
                        break
                except:
                    # Some sheets may not have sheet_state, try anyway
                    sheet = wb[name]
                    break

            if sheet:
                nat = avn = 0
                # Scan first 300 rows (or entire sheet if smaller)
                for row in sheet.iter_rows(max_row=300, values_only=True):
                    for cell in row:
                        if cell:
                            v = str(cell)
                            # Match NATBRA prefix (case-insensitive)
                            if re.search(r"natbra", v, re.IGNORECASE):
                                nat += 1
                            # Match AVNBRA prefix (case-insensitive)
                            if re.search(r"avnbra", v, re.IGNORECASE):
                                avn += 1
                    if nat > 0 and avn > 0:
                        break  # Found both

                if nat > 0:
                    brands.add("natura")
                if avn > 0:
                    brands.add("avon")

            # If no brands found yet, scan ALL visible sheets (not just first)
            if not brands:
                for sheet_name in wb.sheetnames:
                    try:
                        if wb[sheet_name].sheet_state == "visible":
                            sheet = wb[sheet_name]
                            nat = avn = 0
                            for row in sheet.iter_rows(max_row=300, values_only=True):
                                for cell in row:
                                    if cell:
                                        v = str(cell)
                                        if re.search(r"natbra", v, re.IGNORECASE):
                                            nat += 1
                                        if re.search(r"avnbra", v, re.IGNORECASE):
                                            avn += 1
                                if nat > 0 and avn > 0:
                                    break
                            if nat > 0:
                                brands.add("natura")
                            if avn > 0:
                                brands.add("avon")
                            if brands:
                                break
                    except:
                        pass

            wb.close()
        except Exception as e:
            # Log error for debugging, but don't break the app
            import sys
            print(f"Warning: Excel detection failed for {path}: {e}", file=sys.stderr)
        return brands

    @staticmethod
    def get_combined_display_name(brands) -> str:
        if isinstance(brands, str):
            brands = {brands}
        if not brands:
            return "Desconhecida"

        names = []
        if "natura" in brands:
            names.append("Natura")
        if "avon" in brands:
            names.append("Avon")
        if "ml" in brands:
            names.append("Minha Loja")

        return " + ".join(names)

    @staticmethod
    def get_brand_qss_state(brands: set[str]) -> str:
        """Returns a string to be used as 'brand' property in QSS."""
        if not brands:
            return ""
        if len(brands) >= 3:
            return "all"

        # Sort to ensure consistent property values for QSS matching
        sorted_brands = sorted(list(brands))
        return "_".join(sorted_brands)

    # ── Brazil ID guard (BRD — incidente Chile) ─────────────────────────────
    # Funções aditivas: não alteram detect_single/_detect_xml_set/etc. acima,
    # que continuam sendo usadas só para exibição/agrupamento cosmético. Estas
    # aqui são a checagem estrita usada para BLOQUEAR execução.
    @staticmethod
    def check_brazil_id(raw_id: str, *, strict_catalog: bool = False) -> "BrazilIdCheck":
        """
        Checa se `raw_id` (catalog-id ou pricebook-id) pertence a uma loja
        Brasil (Natura/Avon/CB).

        strict_catalog=True: match exato contra o allowlist real de
        catalog-id de produção (BRAZIL_CATALOG_IDS) — usar para catalog-id.
        strict_catalog=False (default): heurística marca + marcador de
        Brasil ("brazil" ou token "-br") — usar para pricebook-id, que ainda
        não tem um allowlist exato confirmado.
        """
        cid = (raw_id or "").strip().lower()
        country_hint = None
        for token, name in _COUNTRY_HINTS.items():
            if re.search(rf"(^|-){re.escape(token)}(-|$)", cid) or token in cid:
                country_hint = name
                break

        if strict_catalog:
            brand = BRAZIL_CATALOG_IDS.get(cid)
            return BrazilIdCheck(raw_id=raw_id, brand=brand, is_brazil=brand is not None,
                                  country_hint=None if brand else country_hint)

        brand = None
        for b, markers in _BRAND_MARKERS.items():
            if any(m in cid for m in markers):
                brand = b
                break

        is_brazil = bool(re.search(r"(^|-)br(-|$)", cid)) or "brazil" in cid
        return BrazilIdCheck(raw_id=raw_id, brand=brand, is_brazil=is_brazil,
                              country_hint=None if (brand and is_brazil) else country_hint)

    @staticmethod
    def check_catalog_file(path: str) -> "BrazilIdCheck":
        """Lê o catalog-id do elemento raiz e valida contra o allowlist Brasil."""
        raw_id = ""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                header = f.read(1024)
            m = re.search(r'catalog-id=["\']([^"\']+)["\']', header, re.IGNORECASE)
            if m:
                raw_id = m.group(1)
        except Exception:
            pass
        return BrandDetector.check_brazil_id(raw_id, strict_catalog=True)

    @staticmethod
    def check_pricebook_file(path: str) -> list["BrazilIdCheck"]:
        """Um BrazilIdCheck por cada pricebook-id distinto encontrado no arquivo."""
        results: list[BrazilIdCheck] = []
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()
            seen: set[str] = set()
            for m in re.finditer(r'pricebook-id=["\']([^"\']+)["\']', content, re.IGNORECASE):
                raw_id = m.group(1)
                if raw_id in seen:
                    continue
                seen.add(raw_id)
                results.append(BrandDetector.check_brazil_id(raw_id, strict_catalog=False))
        except Exception:
            pass
        return results
