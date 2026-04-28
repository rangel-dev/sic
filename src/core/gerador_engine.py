"""
Gerador Engine – Excel → Salesforce Demandware Pricebook XML
Migrated from gerador.js using lxml for schema-faithful generation.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from lxml import etree

PRICEBOOK_NS = "http://www.demandware.com/xml/impex/pricebook/2006-10-31"

# Targets válidos: "natura", "avon", "ambas", "ml"
# Cada entrada define: pricebook-id (DE/POR), display-name, description e parent (POR)
PRICEBOOK_DEFS: dict[str, dict] = {
    "natura": {
        "de": {
            "id":           "br-natura-brazil-list-prices",
            "display":      "01 - Preço DE (Natura Brasil)",
            "description":  "Pricebook designada para alocar os preços cheio dos produtos Natura Brasil",
            "parent":       None,
        },
        "por": {
            "id":           "br-natura-brazil-sale-prices",
            "display":      "02 - Preço POR (Natura Brasil)",
            "description":  "Pricebook designada para alocar os preços promocionais para Natura Brasil",
            "parent":       "br-natura-brazil-list-prices",
        },
        "sku_prefix": "NATBRA-",
    },
    "avon": {
        "de": {
            "id":           "brl-avon-brazil-list-prices",
            "display":      "Avon Brazil BRL List Prices",
            "description":  "Avon Brazil BRL Sale Prices",
            "parent":       None,
        },
        "por": {
            "id":           "brl-avon-brazil-sale-prices",
            "display":      "Avon Brazil BRL Sale Prices",
            "description":  None,
            "parent":       "brl-avon-brazil-list-prices",
        },
        "sku_prefix": "AVNBRA-",
    },
    "ml": {
        "de": {
            "id":           "br-cb-brazil-list-prices",
            "display":      "CB Brazil BRL List Prices",
            "description":  "CB Brazil BRL Sale Prices",
            "parent":       None,
        },
        "por": {
            "id":           "br-cb-brazil-sale-prices",
            "display":      "CB Brazil BRL Sale Prices",
            "description":  "Pricebook designada para alocar os preços promocionais para CB Brasil",
            "parent":       "br-cb-brazil-list-prices",
        },
        "sku_prefix": None,  # ML uses all SKUs regardless of prefix
    },
}

# Legacy IDs for delta base-XML parsing
PRICEBOOK_IDS: dict[str, dict[str, str]] = {
    b: {"de": PRICEBOOK_DEFS[b]["de"]["id"], "por": PRICEBOOK_DEFS[b]["por"]["id"]}
    for b in PRICEBOOK_DEFS
}

SKU_PATTERN = re.compile(r"^(NAT|AVN)BRA-", re.IGNORECASE)


class GeradorEngine:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self._progress = progress_callback or (lambda p, msg: None)

    # ── Public entry point ────────────────────────────────────────────────
    def run(
        self,
        excel_paths: list[str],
        mode: str = "full",
        base_xml_path: Optional[str] = None,
        target: str = "ambas",
    ) -> dict:
        """
        Args:
            excel_paths:   list of Excel file paths to process
            mode:          "full" | "delta"
            base_xml_path: path to the base pricebook XML (delta mode only)
            target:        "natura" | "avon" | "ambas" | "ml"

        Returns a dict with keys: xml_content, brand, stats, error
        """
        result: dict = {"xml_content": None, "brand": "unknown", "stats": {}, "error": None}

        try:
            base_prices: dict = {}
            if mode == "delta" and base_xml_path:
                self._progress(5, "Carregando preços base para Delta…")
                base_prices = self._parse_base_prices(base_xml_path)

            self._progress(20, "Lendo planilhas Excel…")

            # Use a dict to group by SKU — multi-file overlaps: last file wins
            consolidated: dict[str, dict] = {}
            brands_detected: set[str] = set()

            for i, path in enumerate(excel_paths):
                pct = 20 + int(i * (40 / len(excel_paths)))
                self._progress(pct, f"Processando arquivo {i+1}/{len(excel_paths)}…")

                file_products = self._parse_excel(path)
                for p in file_products:
                    sku = p["sku"]
                    consolidated[sku] = p
                    if sku.startswith("NATBRA-"):
                        brands_detected.add("natura")
                    if sku.startswith("AVNBRA-"):
                        brands_detected.add("avon")

            if not consolidated:
                result["error"] = "Nenhum produto válido encontrado na aba 'GRADE DE ATIVAÇÃO'."
                return result

            all_products = list(consolidated.values())

            # Dominant brand label for history/logging
            nat_count = sum(1 for p in all_products if p["sku"].startswith("NATBRA-"))
            avn_count = sum(1 for p in all_products if p["sku"].startswith("AVNBRA-"))
            brand_label = "natura" if nat_count >= avn_count else "avon"
            if len(brands_detected) > 1:
                brand_label = "multibrand"
            result["brand"] = brand_label

            if mode == "delta" and base_prices:
                self._progress(70, "Calculando Delta…")
                before = len(all_products)
                all_products = self._apply_delta(all_products, base_prices)
                result["stats"]["delta_removed"] = before - len(all_products)

            if not all_products:
                result["error"] = "Modo Delta: Nenhuma alteração de preço detectada em relação ao XML base."
                return result

            self._progress(80, "Gerando XML Pricebook…")
            xml_bytes = self._generate_xml(all_products, target)
            result["xml_content"] = xml_bytes
            result["stats"].update(
                {
                    "total":  len(all_products),
                    "natura": nat_count,
                    "avon":   avn_count,
                    "mode":   mode,
                    "target": target,
                }
            )
            self._progress(100, "XML gerado com sucesso!")

        except Exception as exc:
            import traceback
            print(traceback.format_exc())
            result["error"] = str(exc)

        return result

    # ── Excel parsing ─────────────────────────────────────────────────────
    def _parse_excel(self, path: str) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_name = "GRADE DE ATIVAÇÃO"

        if sheet_name not in wb.sheetnames:
            found = False
            for name in wb.sheetnames:
                if "GRADE" in name.upper() and "ATIVA" in name.upper():
                    sheet_name = name
                    found = True
                    break
            if not found:
                wb.close()
                return []

        ws = wb[sheet_name]

        if ws.sheet_state == "hidden":
            wb.close()
            raise ValueError(f"A aba '{sheet_name}' está oculta no arquivo {Path(path).name}")

        products = self._parse_grade(ws)
        wb.close()
        return products

    def _parse_grade(self, ws) -> list[dict]:
        rows = list(ws.iter_rows(max_row=10000, values_only=True))
        if not rows:
            return []

        sku_col = de_col = por_col = None
        sku_start_row = 0

        for i, row in enumerate(rows[:50]):
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                val = str(cell).strip().upper()

                if val == "DE":
                    de_col = j
                if val == "POR":
                    por_col = j

                if sku_col is None and (val.startswith("NATBRA-") or val.startswith("AVNBRA-")):
                    sku_col = j
                    sku_start_row = i

        if sku_col is None: sku_col = 2
        if de_col is None:  de_col  = 26
        if por_col is None: por_col = 27

        # Detect dominant brand in this file to filter cross-brand contamination
        nat_count = avn_count = 0
        for row in rows[sku_start_row:]:
            sku_raw = str(row[sku_col]).strip().upper() if sku_col < len(row) and row[sku_col] else ""
            if not sku_raw:
                continue
            if sku_raw.startswith("NATBRA-"): nat_count += 1
            if sku_raw.startswith("AVNBRA-"): avn_count += 1

        if nat_count == 0 and avn_count == 0:
            return []

        file_brand = "natura" if nat_count >= avn_count else "avon"

        products: list[dict] = []
        empty_streak = 0

        for row in rows[sku_start_row:]:
            sku_row = row[sku_col] if sku_col < len(row) else None
            sku_raw = str(sku_row).strip().upper() if sku_row else ""

            if not sku_raw:
                empty_streak += 1
                if empty_streak >= 50:
                    break
                continue

            empty_streak = 0

            # Contamination filter
            if file_brand == "natura" and sku_raw.startswith("AVNBRA-"): continue
            if file_brand == "avon"   and sku_raw.startswith("NATBRA-"): continue

            if not (sku_raw.startswith("NATBRA-") or sku_raw.startswith("AVNBRA-")):
                continue

            de  = self._to_float(row[de_col]  if de_col  < len(row) else None)
            por = self._to_float(row[por_col] if por_col < len(row) else None)

            if (de and de > 0) or (por and por > 0):
                products.append({"sku": sku_raw, "de": de or 0.0, "por": por or 0.0})

        return products

    # ── Base price parsing (for delta) ────────────────────────────────────
    def _parse_base_prices(self, xml_path: str) -> dict:
        tree = etree.parse(xml_path)
        ns = {"pb": PRICEBOOK_NS}
        prices: dict = {}

        for pb_el in tree.findall(".//pb:pricebook", ns):
            header = pb_el.find("pb:header", ns)
            if header is None:
                continue

            pb_id = header.get("pricebook-id", "")
            price_type = None

            for b_ids in PRICEBOOK_IDS.values():
                if pb_id == b_ids["de"]:
                    price_type = "de"
                    break
                if pb_id == b_ids["por"]:
                    price_type = "por"
                    break

            if price_type is None:
                continue

            for pt in pb_el.findall(".//pb:price-table", ns):
                pid = pt.get("product-id", "").upper()
                amt = pt.find("pb:amount[@quantity='1']", ns)
                if amt is not None and amt.text:
                    prices.setdefault(pid, {})[price_type] = float(amt.text)

        return prices

    # ── Delta filter ──────────────────────────────────────────────────────
    def _apply_delta(self, products: list[dict], base: dict) -> list[dict]:
        delta = []
        for p in products:
            sku = p["sku"]
            if sku not in base:
                delta.append(p)
                continue

            base_de  = base[sku].get("de", 0.0)
            base_por = base[sku].get("por", 0.0)

            de_changed  = abs(p["de"]  - base_de)  > 0.01
            por_changed = abs(p["por"] - base_por) > 0.01

            if de_changed or por_changed:
                delta.append(p)
        return delta

    # ── XML generation ────────────────────────────────────────────────────
    def _generate_xml(self, products: list[dict], target: str) -> bytes:
        """
        Generate the pricebook XML based on the chosen target.

        target:
          "natura" → Natura DE + Natura POR  (only NATBRA- SKUs)
          "avon"   → Avon DE   + Avon POR    (only AVNBRA- SKUs)
          "ambas"  → Natura DE + Natura POR + Avon DE + Avon POR
          "ml"     → CB DE     + CB POR      (all SKUs from both brands)
        """
        root = etree.Element("pricebooks", xmlns=PRICEBOOK_NS)

        if target in ("natura", "ambas"):
            defs = PRICEBOOK_DEFS["natura"]
            nat_skus = [p for p in products if p["sku"].startswith("NATBRA-")]
            self._add_pricebook_pair(root, defs, nat_skus)

        if target in ("avon", "ambas"):
            defs = PRICEBOOK_DEFS["avon"]
            avn_skus = [p for p in products if p["sku"].startswith("AVNBRA-")]
            self._add_pricebook_pair(root, defs, avn_skus)

        if target == "ml":
            defs = PRICEBOOK_DEFS["ml"]
            # ML uses all products from both brands
            self._add_pricebook_pair(root, defs, products)

        return etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", pretty_print=True
        )

    def _add_pricebook_pair(self, root: etree._Element, defs: dict, prods: list[dict]):
        """Add the DE and POR pricebook blocks for a given brand definition."""
        de_list  = [p for p in prods if p["de"]  > 0]
        por_list = [p for p in prods if p["por"] > 0]

        if de_list:
            self._build_pricebook(root, defs["de"], de_list, "de")
        if por_list:
            self._build_pricebook(root, defs["por"], por_list, "por")

    def _build_pricebook(
        self,
        root: etree._Element,
        pb_def: dict,
        products: list[dict],
        price_key: str,
    ) -> None:
        """
        Build a complete <pricebook> element with the correct header fields.

        pb_def keys: id, display, description, parent
        """
        NS = PRICEBOOK_NS

        pb_el  = etree.SubElement(root, "pricebook")
        header = etree.SubElement(pb_el, f"{{{NS}}}header")
        header.set("pricebook-id", pb_def["id"])

        # <currency>
        curr = etree.SubElement(header, f"{{{NS}}}currency")
        curr.text = "BRL"

        # <display-name xml:lang="x-default">
        disp = etree.SubElement(header, f"{{{NS}}}display-name")
        disp.set("{http://www.w3.org/XML/1998/namespace}lang", "x-default")
        disp.text = pb_def["display"]

        # <description xml:lang="x-default">  (optional per brand)
        if pb_def.get("description"):
            desc = etree.SubElement(header, f"{{{NS}}}description")
            desc.set("{http://www.w3.org/XML/1998/namespace}lang", "x-default")
            desc.text = pb_def["description"]

        # <online-flag>
        online = etree.SubElement(header, f"{{{NS}}}online-flag")
        online.text = "true"

        # <parent>  (POR pricebooks only)
        if pb_def.get("parent"):
            parent_el = etree.SubElement(header, f"{{{NS}}}parent")
            parent_el.text = pb_def["parent"]

        # <price-tables>
        tables = etree.SubElement(pb_el, f"{{{NS}}}price-tables")
        for p in products:
            pt = etree.SubElement(tables, f"{{{NS}}}price-table")
            pt.set("product-id", p["sku"])
            amt = etree.SubElement(pt, f"{{{NS}}}amount")
            amt.set("quantity", "1")
            amt.text = f"{p[price_key]:.2f}"

    # ── Helper ────────────────────────────────────────────────────────────
    @staticmethod
    def _to_float(val) -> Optional[float]:
        if val is None or str(val).strip() == "":
            return None
        try:
            s = str(val).replace("R$", "").replace("\xa0", "").strip()
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            return float(s)
        except (ValueError, TypeError):
            return None
