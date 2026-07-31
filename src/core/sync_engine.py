"""
Sync Engine – Merchandising Sync: store-list assignments + product attributes.
Rebuilt from legacy JS V14.0/V11.1 to include full Business Rules.
"""
from __future__ import annotations

import json
import os
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional, Dict, Set, Any

import openpyxl
from lxml import etree

from src.core.excel_reader import dominant_brand, parse_price

CATALOG_NS = "http://www.demandware.com/xml/impex/catalog/2006-10-31"
SKU_PATTERN = re.compile(r"^(NAT|AVN)BRA-", re.IGNORECASE)
MIN_FILE_AGE_SECONDS = 600  # 10-minute golden rule

# BRD-008: categorias SFCC de faixa de preço para produtos "Presente" — Natura
PRESENTE_CATEGORY_IDS = (
    "presentes-faixa-de-preco-agradecer",
    "presentes-faixa-de-preco-encantar",
    "presentes-faixa-de-preco-surpreender",
    "presentes-faixa-de-preco-impressionar",
)

# BRD-009: mesma lógica, faixas e IDs de categoria próprios da Avon.
PRESENTE_CATEGORY_IDS_AVON = (
    "presentes-faixa-de-preco-ate-19",
    "presentes-faixa-de-preco-de-20-ate-49",
    "presentes-faixa-de-preco-de-50-ate-99",
    "presentes-faixa-de-preco-acima-de-150",
)


def _price_bucket(price: float) -> str:
    """Mapeia o "Preço POR" para uma das 4 faixas de presente Natura
    (BRD-008 §4.2). Limites superiores inclusivos."""
    if price <= 50.00:
        return "presentes-faixa-de-preco-agradecer"
    if price <= 100.00:
        return "presentes-faixa-de-preco-encantar"
    if price <= 150.00:
        return "presentes-faixa-de-preco-surpreender"
    return "presentes-faixa-de-preco-impressionar"


def _price_bucket_avon(price: float) -> Optional[str]:
    """Mapeia o "Preço POR" para uma das 4 faixas de presente Avon (BRD-009).
    Lacuna PROPOSITAL entre R$100,00 e R$150,00 (ambos inclusive): SKU Avon
    "Presente" nessa faixa não recebe nenhuma das 4 categorias — retorna
    None e o chamador trata como "sem categoria" (mesmo mecanismo do preço
    inválido). "Acima de 150" é estritamente > 150.00 (150.00 exato cai na
    lacuna)."""
    if price <= 19.99:
        return "presentes-faixa-de-preco-ate-19"
    if price <= 49.99:
        return "presentes-faixa-de-preco-de-20-ate-49"
    if price <= 99.99:
        return "presentes-faixa-de-preco-de-50-ate-99"
    if price <= 150.00:
        return None  # lacuna proposital 100,00–150,00 (BRD-009)
    return "presentes-faixa-de-preco-acima-de-150"


# Marca -> (tupla de IDs de categoria, função de bucket de preço). Fonte
# única de verdade consultada por _compute_presente_targets.
_PRESENTE_RULES_BY_BRAND: dict[str, tuple[tuple[str, ...], Callable[[float], Optional[str]]]] = {
    "natura": (PRESENTE_CATEGORY_IDS, _price_bucket),
    "avon": (PRESENTE_CATEGORY_IDS_AVON, _price_bucket_avon),
}


@dataclass
class SyncResult:
    xml_content: Optional[bytes] = None
    stats: dict = field(default_factory=dict)
    brand: str = "unknown"
    catalog_id: str = ""
    error: Optional[str] = None
    warnings: list[str] = field(default_factory=list)
    report: list[dict] = field(default_factory=list)


class SyncEngine:
    def __init__(self, progress_callback: Optional[Callable] = None):
        self._progress = progress_callback or (lambda p, msg: None)

    # ── Entry point ───────────────────────────────────────────────────────
    def run(
        self,
        excel_paths: list[str],
        catalog_xml_paths: list[str],
    ) -> SyncResult:
        result = SyncResult()

        try:
            # File-age safety check
            self._progress(5, "Validando antiguidade dos XMLs…")
            for path in catalog_xml_paths:
                warning = self._check_file_age(path)
                if warning:
                    result.warnings.append(warning)

            # Parse Excel: lists + grade (visibility + seals)
            self._progress(15, "Lendo planilhas Excel (Extração Bruta)…")
            excel_lists, grade_map, brand, presente_cols_ok = self._parse_excel_files(excel_paths)
            result.brand = brand

            if not excel_lists and not grade_map:
                result.error = "Nenhum dado encontrado nas planilhas (Listas ou Grade)."
                return result

            # Parse current catalog XMLs
            self._progress(40, "Analisando Catálogos Master Data…")
            catalog_state, catalog_id = self._parse_catalogs(catalog_xml_paths)
            result.catalog_id = catalog_id

            # Consolidate Brand accurately
            result.brand = self._consolidate_brand(catalog_state, grade_map)

            # BRD-008: presentes por faixa de preço — calculado uma vez, usado
            # nas estatísticas (_execute_rules) e na geração do XML
            presente_targets = self._compute_presente_targets(catalog_state, grade_map, result.brand, presente_cols_ok)

            # Execute Business Rules (V11.1 Motor)
            self._progress(60, "Aplicando Regras de Governança V11.1…")
            delta, metrics, report = self._execute_rules(catalog_state, grade_map, excel_lists, result.brand, presente_targets)

            # Generate XML
            self._progress(80, "Gerando XML Catálogo Delta…")
            result.xml_content = self._generate_catalog_xml(delta, catalog_id, excel_lists, catalog_state, result.brand, presente_targets)

            result.stats = metrics
            result.report = report
            self._progress(100, "Sincronização concluída!")

        except Exception as exc:
            result.error = f"Erro crítico: {str(exc)}"

        return result

    # ── File age check ────────────────────────────────────────────────────
    def _check_file_age(self, path: str) -> Optional[str]:
        try:
            age = time.time() - os.path.getmtime(path)
            if age < MIN_FILE_AGE_SECONDS:
                minutes = int(age / 60)
                return (
                    f"{Path(path).name}: arquivo tem apenas {minutes}m "
                    f"(regra: ≥ 10 min). Pode estar incompleto."
                )
        except OSError:
            pass
        return None

    # ── Brand Consolidation ───────────────────────────────────────────────
    def _consolidate_brand(self, catalog_state: dict, grade_map: dict) -> str:
        nat_count = 0
        avn_count = 0
        for pid in catalog_state["products"].keys():
            if "NATBRA-" in pid: nat_count += 1
            if "AVNBRA-" in pid: avn_count += 1
        for pid in grade_map.keys():
            if "NATBRA-" in pid: nat_count += 1
            if "AVNBRA-" in pid: avn_count += 1
        
        return dominant_brand(nat_count, avn_count)

    # ── Presente / Faixa de Preço (BRD-008 Natura, BRD-009 Avon) ────────────
    def _compute_presente_targets(
        self, catalog_state: dict, grade_map: dict, brand: str, presente_cols_ok: bool
    ) -> Optional[dict[str, set[str]]]:
        """SKUs (Natura ou Avon), variação (não-master), com "CATEGORIA
        PLANEJAMENTO" == "Presente"/"PRESENTES" e preço válido na Grade,
        mapeados para a categoria SFCC de faixa de preço da marca (tabelas de
        IDs e faixas distintas por marca, ver _PRESENTE_RULES_BY_BRAND).
        Sempre retorna todas as chaves da tabela da marca (mesmo vazias) para
        que o diff add/remove detecte quando uma faixa esvazia.

        Retorna None quando a regra não deve rodar (marca sem tabela de
        faixas definida, ou a Grade não tem as colunas "CATEGORIA
        PLANEJAMENTO"/"POR") — nesse caso o chamador não deve gerar NENHUM
        delta de presente (nem add, nem remove), para não apagar
        categorização existente por engano.
        """
        rules = _PRESENTE_RULES_BY_BRAND.get(brand)
        if rules is None:
            return None  # marca sem tabela de faixas definida

        if not presente_cols_ok:
            return None  # Grade sem as colunas necessárias: regra não roda

        category_ids, bucket_fn = rules
        targets: dict[str, set[str]] = {cat_id: set() for cat_id in category_ids}

        for pid, prod in catalog_state["products"].items():
            if prod["isMaster"]:
                continue  # CA-06: produto-pai/mestre é isento

            g = grade_map.get(pid)
            if not g:
                continue

            # Dados reais da Grade usam "PRESENTES" (plural); o BRD-008 documenta
            # "Presente" (singular). Aceita as duas formas por segurança entre ciclos.
            if str(g.get("planning_cat", "")).strip().upper() not in ("PRESENTE", "PRESENTES"):
                continue

            price = g.get("price")
            if not price:  # None ou 0.0 -> CA-04 (sem preço válido = não categorizado)
                continue

            bucket = bucket_fn(price)
            if bucket is None:  # Avon: lacuna proposital 100,00–150,00 (BRD-009)
                continue

            targets[bucket].add(pid)

        return targets

    # ── Parse Excel Files ─────────────────────────────────────────────────
    def _parse_excel_files(self, paths: list[str]) -> tuple[dict[str, set[str]], dict[str, dict], str, bool]:
        excel_lists: dict[str, set[str]] = {}
        grade_map: dict[str, dict] = {}
        nat = avn = 0
        presente_cols_ok = False

        self._progress(15, "> Iniciando Varredura Bruta V14.0...")
        self._progress(16, "> Mapeando cores de fundo do Excel (Bruto OpenXML)...")

        for path in paths:
            wb = openpyxl.load_workbook(path, data_only=True)
            self._progress(17, f"> Excel carregado (Modo Recuperação). Abas detectadas: {len(wb.sheetnames)}")

            for name in wb.sheetnames:
                ws = wb[name]
                if ws.sheet_state != 'visible':
                    # Aba oculta (hidden/veryHidden) é tratada como se não
                    # existisse no arquivo — mesmo padrão de
                    # auditor_engine.py:299 (sheet_state != 'visible' cobre
                    # os dois estados ocultos do openpyxl de uma vez). BRD-009.
                    continue

                n = name.upper()
                is_grade = "GRADE" in n or "ATIVA" in n
                list_match = re.search(r"LISTA[\s_-]*([0-9.]+)", n)
                is_list = bool(list_match)

                if not is_grade and not is_list:
                    continue

                if is_grade:
                    self._progress(18, f"> Aba [{name}] -> Identificada como: GRADE")
                if is_list:
                    clean_id = f"LISTA_{list_match.group(1)}"
                    self._progress(18, f"> Aba [{name}] -> Identificada como: LISTA ({clean_id})")
                    # BRD-009: registra a chave já aqui, mesmo com 0 SKUs, para
                    # que o diff downstream enxergue a lista esvaziada e gere
                    # delta de remoção total (old_set - set() = old_set), em
                    # vez de simplesmente ignorá-la por ausência no dict.
                    excel_lists.setdefault(clean_id, set())

                # values_only=False allows accessing Cell properties like .fill color style
                rows = list(ws.iter_rows(max_row=5000, values_only=False))
                if not rows:
                    continue

                sku_col = vis_col = selo_col = por_col = plan_col = None

                if is_grade:
                    # Scan headers
                    for i, row in enumerate(rows[:50]):
                        for j, cell in enumerate(row):
                            if cell.value is None: continue
                            val = str(cell.value).strip().upper()
                            if "VISIBLE" in val or "VITRINE" in val or "VISIBILIDADE" in val:
                                vis_col = j
                            if val in ("SELO", "MARKETING"):
                                selo_col = j
                            if val == "POR":  # BRD-008: "Preço POR" — cabeçalho real é só "POR"
                                por_col = j
                            if val == "CATEGORIA PLANEJAMENTO":  # BRD-008
                                plan_col = j
                            if sku_col is None and SKU_PATTERN.match(str(cell.value).strip()):
                                sku_col = j
                                break
                    if por_col is not None and plan_col is not None:
                        presente_cols_ok = True
                    self._progress(19, "> Chaves de dados carregadas: GRADE")
                
                # Scan entire sheet
                extracted_skus = 0
                for rr, row in enumerate(rows):
                    for cc, cell in enumerate(row):
                        if cell.value is None: continue
                        val = str(cell.value).strip().upper()
                        
                        if ("NATBRA-" in val or "AVNBRA-" in val):
                            if "NATBRA-" in val: nat += 1
                            if "AVNBRA-" in val: avn += 1
                            pid = val.replace("-(0+)", "-").strip()
                            extracted_skus += 1
                            if is_grade:
                                vis = False
                                if vis_col is not None and vis_col < len(row):
                                    vc = row[vis_col].value
                                    if vc is not None: vis = str(vc).strip().upper() in ("SIM", "S", "TRUE", "1", "YES", "X", "✓")
                                
                                selo = ""
                                selo_color = "#edf0ff"
                                if selo_col is not None and selo_col < len(row):
                                    sc_cell = row[selo_col]
                                    if sc_cell.value: 
                                        selo = str(sc_cell.value).strip()
                                        # Get fill color directly from cell
                                        if sc_cell.fill and sc_cell.fill.start_color and sc_cell.fill.start_color.rgb:
                                            raw_c = str(sc_cell.fill.start_color.rgb)
                                            if raw_c not in ("00000000", "00000000000000", "000000"):
                                                selo_color = "#" + raw_c[-6:]
                                
                                price = None
                                if por_col is not None and por_col < len(row):
                                    price = parse_price(row[por_col].value)

                                planning_cat = ""
                                if plan_col is not None and plan_col < len(row):
                                    pc = row[plan_col].value
                                    if pc is not None:
                                        planning_cat = str(pc).strip()

                                grade_map[pid] = {
                                    "visible": vis, "seal": selo, "color": selo_color,
                                    "price": price, "planning_cat": planning_cat,
                                }
                            
                            if is_list:
                                clean_id = f"LISTA_{list_match.group(1)}"
                                excel_lists.setdefault(clean_id, set()).add(pid)
                
                self._progress(19, f"> Aba [{name}]: {extracted_skus} SKUs extraídos.")
            wb.close()

        brand = dominant_brand(nat, avn)
        return excel_lists, grade_map, brand, presente_cols_ok

    # ── Parse Catalog XML ─────────────────────────────────────────────────
    def _parse_catalogs(self, paths: list[str]) -> tuple[dict, str]:
        state: dict = {
            "products": {},       # product-id -> info
            "masters": {},        # master-id -> set(variants)
            "child_to_master": {},# variant-id -> master-id
            "assignments": {}     # category-id -> set(product-ids)
        }
        catalog_id = ""

        for path in paths:
            try:
                tree = etree.parse(path)
            except etree.XMLSyntaxError as exc:
                raise ValueError(f"XML inválido ({Path(path).name}): {exc}") from exc

            ns = {"c": CATALOG_NS}
            root = tree.getroot()
            if not catalog_id:
                catalog_id = root.get("catalog-id", "")

            # Assignments
            for ca in root.findall("c:category-assignment", ns):
                cat_id = ca.get("category-id", "").upper()
                pid    = ca.get("product-id",  "").upper()
                state["assignments"].setdefault(cat_id, set()).add(pid)

            # Products
            for prod_el in root.findall(".//c:product", ns):
                pid = prod_el.get("product-id", "").upper()
                
                is_master = prod_el.find("c:variations", ns) is not None
                
                online_el = prod_el.find("c:online-flag", ns)
                srch_el   = prod_el.find("c:searchable-flag", ns)
                seo_el    = prod_el.find("c:searchable-if-unavailable-flag", ns)
                dn_el = prod_el.find("c:display-name", ns)
                
                online = (online_el.text or "").lower() == "true" if online_el is not None else False
                searchable = (srch_el.text or "").lower() == "true" if srch_el is not None else False
                seo_flag = (seo_el.text or "").lower() == "true" if seo_el is not None else False
                dn = (dn_el.text or "").strip() if dn_el is not None else ""
                
                # Custom Attributes
                fn = ""
                sobj = ""
                cas = prod_el.find("c:custom-attributes", ns)
                if cas is not None:
                    for ca in cas.findall("c:custom-attribute", ns):
                        aid = ca.get("attribute-id", "")
                        if aid == "natg_friendlyName":
                            fn = (ca.text or "").strip()
                        elif aid == "natg_preferencialProductSlot":
                            sobj = (ca.text or "").strip()

                state["products"][pid] = {
                    "id": pid,
                    "isMaster": is_master,
                    "online": online,
                    "searchable": searchable,
                    "seoFlag": seo_flag,
                    "dn": dn,
                    "fn": fn,
                    "sobj": sobj
                }

                if is_master:
                    kids = set()
                    vars_el = prod_el.find("c:variations", ns)
                    if vars_el is not None:
                        # Assuming variations structure: <variations><variants><variant product-id="..."/>
                        variants_tag = vars_el.find(".//c:variant", ns)
                        if variants_tag is not None:
                            for var_el in vars_el.findall(".//c:variant", ns):
                                vid = var_el.get("product-id", "").upper()
                                kids.add(vid)
                                state["child_to_master"][vid] = pid
                    state["masters"][pid] = kids

        return state, catalog_id

    # ── Engine Business Rules V11.1 ───────────────────────────────────────
    def _execute_rules(self, catalog_state: dict, grade_map: dict, excel_lists: dict, brand: str, presente_targets: Optional[dict[str, set[str]]]) -> tuple[dict, dict, list]:
        deltas = []
        report = []
        active_masters = set()
        
        matched = cut = seal_c = on_c = vis_c = 0
        
        is_avon = brand == "avon"
        
        for pid, prod in catalog_state["products"].items():
            if prod["isMaster"]: continue
            
            g = grade_map.get(pid)
            if g: matched += 1
            
            n_on = bool(g)
            n_vis = g["visible"] if g else False
            reason = "Ativo na Grade" if n_on else "Fora da Grade (Desativar)"
            
            # FACÃO (Regra 4)
            is_dn_upper = prod["dn"] and prod["dn"] == prod["dn"].upper()
            is_fn_upper = prod["fn"] and prod["fn"] == prod["fn"].upper()
            is_fn_empty = not prod["fn"]
            
            if n_vis and is_dn_upper and (is_fn_upper or is_fn_empty):
                n_vis = False
                cut += 1
                reason = "Corte Facão (Caixa Alta)"
                
            if n_on:
                on_c += 1
                if pid in catalog_state["child_to_master"]:
                    active_masters.add(catalog_state["child_to_master"][pid])
            if n_vis: vis_c += 1
            
            up = {}
            if prod["online"] != n_on: up["online-flag"] = "true" if n_on else "false"
            if prod["searchable"] != n_vis: up["searchable-flag"] = "true" if n_vis else "false"
            
            s_act = "-"
            # LÓGICA DE SELO COM LIMPEZA DE LIXO (V14.0 sync_app.html linhas 478-494)
            # O V14 usa JSON.stringify com lowercase e COMPARA com prod.sObj
            if n_on and g and g.get("seal"):
                clean = g["seal"].lower()
                # JSON.stringify no JS sem argumentos extras gera: {"Name":"val","Color":"#edf0ff","Border":"radius"}
                # Python json.dumps com separators=(',',':') gera o mesmo formato
                s_o = json.dumps({"Name": clean, "Color": g["color"], "Border": "radius"}, separators=(',', ':'), ensure_ascii=False)
                s_a = json.dumps([{"tag": clean, "type": "Product", "color": g["color"], "border": "radius", "date_start": "", "date_end": "", "priority": "1", "values": clean}], separators=(',', ':'), ensure_ascii=False)
                if prod["sobj"] != s_o:
                    up["natg_preferencialProductSlot"] = s_o
                    up["natg_preferencialProductSlotCbSite"] = s_o
                    up["natg_productSlotsRules"] = s_a
                    up["natg_productSlotsRulesCbSite"] = s_a
                    s_act = clean
                    seal_c += 1
                else:
                    s_act = f"{clean} (Mantido)"
            elif prod["sobj"] and prod["sobj"] != "" and prod["sobj"] != "[]":
                # SANITIZAÇÃO: lixo no XML mas sem selo na grade -> limpa tudo
                up["natg_preferencialProductSlot"] = ""
                up["natg_preferencialProductSlotCbSite"] = ""
                up["natg_productSlotsRules"] = ""
                up["natg_productSlotsRulesCbSite"] = ""
                s_act = "Limpeza (Lixo)"
                seal_c += 1
                
            if up:
                deltas.append({"id": pid, "up": up})
                
            # Formatting lists for report
            prod_lists = []
            for cat_id, skus in excel_lists.items():
                if pid in skus:
                    fmt_id = cat_id.lower().replace("_", "-") if is_avon else cat_id.upper().replace("-", "_")
                    prod_lists.append(fmt_id)
            
            report.append({
                "SKU": pid,
                "Nome": prod["dn"],
                "Na Grade": "SIM" if g else "NÃO",
                "Online Antigo": "SIM" if prod["online"] else "NÃO",
                "Online Novo": "SIM" if n_on else "NÃO",
                "Vitrine Antiga": "SIM" if prod["searchable"] else "NÃO",
                "Vitrine Nova": "SIM" if n_vis else "NÃO",
                "Ação Selo": s_act,
                "Listas": ", ".join(prod_lists),
                "Motivo/Regra": reason
            })
            
        # Mestres (Regra 7)
        for m_id, kids in catalog_state["masters"].items():
            if m_id not in catalog_state["products"]: continue
            m_prod = catalog_state["products"][m_id]
            act = m_id in active_masters
            
            up = {}
            if m_prod["online"] != act: up["online-flag"] = "true" if act else "false"
            if m_prod["searchable"] != act: up["searchable-flag"] = "true" if act else "false"
            
            if up:
                deltas.append({"id": m_id, "up": up})
                
            report.append({
                "SKU": m_id,
                "Nome": m_prod["dn"],
                "Na Grade": "MESTRE",
                "Online Antigo": "SIM" if m_prod["online"] else "NÃO",
                "Online Novo": "SIM" if act else "NÃO",
                "Vitrine Antiga": "SIM" if m_prod["searchable"] else "NÃO",
                "Vitrine Nova": "SIM" if act else "NÃO",
                "Ação Selo": "-",
                "Listas": "-",
                "Motivo/Regra": "Mestre Ativo" if act else "Mestre Sem Filhos (OFF)"
            })
            
        # Calculate list stats and deltas — JS only counts add+rem of category assignments
        list_add = 0
        list_rem = 0
        lists_details = []
        for cat_id, skus in excel_lists.items():
            fmt_id = cat_id.lower().replace("_", "-") if is_avon else cat_id.upper().replace("-", "_")
            old_set = catalog_state["assignments"].get(fmt_id.upper(), set())
            
            added = len(skus - old_set)
            removed = len(old_set - skus)
            list_add += added
            list_rem += removed
            
            status = "Sincronizado ✓"
            if added > 0 or removed > 0:
                status = f"Delta: +{added} / -{removed}"
            if len(skus) == 0:
                status = "Lista Vazia"
            
            lists_details.append({
                "id": fmt_id,
                "excel": len(skus),
                "xml": len(old_set),
                "status": status
            })

        # BRD-008: categorias de faixa de preço de presente — cat_id JÁ é o ID
        # final (ex. "presentes-faixa-de-preco-agradecer"), sem a transformação
        # de case usada para LISTA_N (por isso o loop é separado do de cima).
        # presente_targets é None quando a regra não deve rodar nesta execução
        # (marca != Natura, ou Grade sem as colunas necessárias) — nesse caso
        # nenhuma linha/estatística de presente é gerada.
        if presente_targets is not None:
            for cat_id, skus in presente_targets.items():
                old_set = catalog_state["assignments"].get(cat_id.upper(), set())

                added = len(skus - old_set)
                removed = len(old_set - skus)
                list_add += added
                list_rem += removed

                status = "Sincronizado ✓"
                if added > 0 or removed > 0:
                    status = f"Delta: +{added} / -{removed}"
                if len(skus) == 0:
                    status = "Lista Vazia"

                lists_details.append({
                    "id": cat_id,
                    "excel": len(skus),
                    "xml": len(old_set),
                    "status": status
                })

        metrics = {
            "xml": len(catalog_state["products"]),
            "grade": len(grade_map),
            "match": matched,
            "onC": on_c,
            "visC": vis_c,
            "hidden": on_c - vis_c,
            "cut": cut,
            "mOff": len(catalog_state["masters"]) - len(active_masters),
            "sealC": seal_c,
            # JS Total Deltas = db.deltas.length (V14.0 sync_app.html linha 528)
            "deltas": len(deltas),
            "lists": len(excel_lists),
            "presente_categories": len(presente_targets) if presente_targets is not None else 0,
            "lists_details": lists_details
        }
        
        return {"products": deltas}, metrics, report

    # ── XML Generation ────────────────────────────────────────────────────
    def _generate_catalog_xml(self, delta: dict, catalog_id: str, excel_lists: dict, catalog_state: dict, brand: str, presente_targets: Optional[dict[str, set[str]]]) -> bytes:
        root = etree.Element("catalog", xmlns=CATALOG_NS)
        root.set("catalog-id", catalog_id or "storefront-catalog")
        
        is_avon = brand == "avon"
        
        # Product Deltas
        for d in delta["products"]:
            pid = d["id"]
            up = d["up"]
            
            prod_el = etree.SubElement(root, "product")
            prod_el.set("product-id", pid)
            
            if "online-flag" in up:
                el = etree.SubElement(prod_el, "online-flag")
                el.text = up["online-flag"]
            if "searchable-flag" in up:
                el = etree.SubElement(prod_el, "searchable-flag")
                el.text = up["searchable-flag"]
            if "searchable-if-unavailable-flag" in up:
                el = etree.SubElement(prod_el, "searchable-if-unavailable-flag")
                el.text = up["searchable-if-unavailable-flag"]
                
            custom_attrs = {k: v for k, v in up.items() if not k.endswith("flag")}
            if custom_attrs:
                cas_el = etree.SubElement(prod_el, "custom-attributes")
                for attr_id, val in custom_attrs.items():
                    ca_el = etree.SubElement(cas_el, "custom-attribute")
                    ca_el.set("attribute-id", attr_id)
                    ca_el.text = val

        # Category Assignments (Lists)
        for cat_id, skus_in_excel in excel_lists.items():
            fmt_id = cat_id.lower().replace("_", "-") if is_avon else cat_id.upper().replace("-", "_")
            
            # Additions
            for sku_id in skus_in_excel:
                p = catalog_state["products"].get(sku_id)
                if p:
                    # Check if already in xml for this cat
                    cats = catalog_state["assignments"].get(fmt_id.upper(), set())
                    if sku_id not in cats:
                        ca = etree.SubElement(root, "category-assignment")
                        ca.set("category-id", fmt_id)
                        ca.set("product-id", sku_id)
            
            # Removals
            cats = catalog_state["assignments"].get(fmt_id.upper(), set())
            for pid_in_xml in cats:
                if pid_in_xml not in skus_in_excel:
                    ca = etree.SubElement(root, "category-assignment")
                    ca.set("category-id", fmt_id)
                    ca.set("product-id", pid_in_xml)
                    ca.set("mode", "delete")

        # BRD-008: category-assignment de faixa de preço de presente. cat_id
        # já é o ID final da categoria, sem transformação de case.
        # presente_targets é None quando a regra não deve rodar nesta execução
        # (marca != Natura, ou Grade sem as colunas necessárias) — nesse caso
        # NENHUM category-assignment de presente é emitido, nem add nem
        # delete, para não apagar categorização existente por engano.
        if presente_targets is not None:
            for cat_id, skus_in_target in presente_targets.items():
                # Additions
                for sku_id in skus_in_target:
                    p = catalog_state["products"].get(sku_id)
                    if p:
                        cats = catalog_state["assignments"].get(cat_id.upper(), set())
                        if sku_id not in cats:
                            ca = etree.SubElement(root, "category-assignment")
                            ca.set("category-id", cat_id)
                            ca.set("product-id", sku_id)

                # Removals
                cats = catalog_state["assignments"].get(cat_id.upper(), set())
                for pid_in_xml in cats:
                    if pid_in_xml not in skus_in_target:
                        ca = etree.SubElement(root, "category-assignment")
                        ca.set("category-id", cat_id)
                        ca.set("product-id", pid_in_xml)
                        ca.set("mode", "delete")

        return etree.tostring(
            root, xml_declaration=True, encoding="UTF-8", pretty_print=True
        )
