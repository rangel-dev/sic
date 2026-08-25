"""
Auditor View – Double-Blind Price & Catalog Validation Engine.
The most feature-rich view of the suite:
  • 3 DropZone inputs  (Pricebook XML, Catalog XMLs, Excel files)
  • Run button + QProgressBar
  • 12 clickable ErrorCard tiles
  • Filtered QTableWidget (max 500 rows)
  • AI Diagnostic panel (HTML rendered in QTextBrowser)
  • Export to Excel + Send to Google Chat
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

import pandas as pd
from PySide6.QtCore import Qt, QSettings, QTimer
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

from src.core.auditor_engine import AuditResult, ERROR_META
from src.core.ai_agent import AiAgent
from src.core.brand_detector import BrandDetector
from src.core.auditor.kit_validation import KIT_ERROR_META
from src.ui.components.base_widgets import Divider, DropZone, ErrorCard, SectionHeader, show_rejection_dialog
from src.workers.worker_auditor import AuditorWorker
from src.core.history_engine import HistoryEngine
from src.core.utils import get_unique_path
from src.core.certificate_engine import CertificateEngine


MAX_TABLE_ROWS = 500

# Layout da tabela no scroll único: cresce até MAX linhas visíveis; acima
# disso a própria tabela rola, mantendo a página em tamanho gerenciável.
MIN_TABLE_VISIBLE_ROWS = 6
MAX_TABLE_VISIBLE_ROWS = 25


class AuditorView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._worker: Optional[AuditorWorker] = None
        self._result: Optional[AuditResult]   = None
        self._active_filters: set[str]        = set()  # error codes
        self._brand_filter:  str = "all"       # "all", "natura", "avon"
        self._last_kit_only: bool = False      # última execução foi modo só-kit (BRD-007)
        self._error_cards: dict[str, ErrorCard] = {}
        self._kit_active_filters: set[str] = set()      # kit "kind" codes (independente de _active_filters)
        self._kit_error_cards: dict[str, ErrorCard] = {}
        self._kit_brand_filter: str = "all"    # marca da seção de kits ("all"/"natura"/"avon")
        self._audit_collapsed: bool = False    # capítulo "Auditoria Tradicional" recolhido?
        self._kit_collapsed: bool   = False    # capítulo "Validação de Kits (BO)" recolhido?
        self._settings = QSettings("SIC", "SIC_Suite")
        self._setup_ui()

    # ── Layout ────────────────────────────────────────────────────────────
    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(SectionHeader(
            "✓  Auditor  — Motor de Auditoria de Catálogo"
        ))
        root.addWidget(Divider())

        # ── Página de rolagem única (padrão Exportador) ───────────────────
        # Tudo vive empilhado dentro de um único QScrollArea para que cada
        # bloco respire em altura natural, sem ser espremido por splitters.
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        root.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        # ── Bloco 1: Arquivos de Entrada ──────────────────────────────────
        inputs_box = QGroupBox("Arquivos de Entrada")
        inputs_box_layout = QVBoxLayout(inputs_box)
        inputs_box_layout.setContentsMargins(16, 18, 16, 14)
        inputs_box_layout.setSpacing(12)

        # File inputs row
        inputs_row = QHBoxLayout()
        inputs_row.setSpacing(16)

        # 1 – Pricebook XML
        pb_col = QVBoxLayout()
        pb_lbl = QLabel("Pricebook XML")
        pb_lbl.setObjectName("label_section")
        pb_col.addWidget(pb_lbl)
        self._dz_pb = DropZone(
            "Pricebook XML\n(br-natura / brl-avon)",
            "XML (*.xml)",
        )
        self._dz_pb.setToolTip(
            "XML de Pricebook exportado do Salesforce Business Manager.\n"
            "Contém os pricebooks de Lista (DE) e Promocional (POR)."
        )
        pb_col.addWidget(self._dz_pb)
        inputs_row.addLayout(pb_col, 1)

        # 2 – Catalog XMLs
        cat_col = QVBoxLayout()
        cat_lbl = QLabel("Catálogo(s) XML  (Natura + Avon + ML)")
        cat_lbl.setObjectName("label_section")
        cat_col.addWidget(cat_lbl)
        self._dz_cat = DropZone(
            "XMLs de Catálogo\n(múltiplos permitidos)",
            "XML (*.xml)",
            multiple=True,
        )
        self._dz_cat.setToolTip(
            "XMLs de Catálogo do Salesforce: natura-br, avon-br, cbbrazil.\n"
            "Contêm category-assignments, online-flag e searchable-flag."
        )
        cat_col.addWidget(self._dz_cat)
        inputs_row.addLayout(cat_col, 1)

        # 3 – Excel files
        excel_col = QVBoxLayout()
        excel_lbl = QLabel("Planilha(s) Excel  (GRADE DE ATIVAÇÃO)")
        excel_lbl.setObjectName("label_section")
        excel_col.addWidget(excel_lbl)
        self._dz_excel = DropZone(
            "Excel(s) comerciais\n(múltiplos permitidos)",
            "Excel (*.xlsx *.xlsm *.xls)",
            multiple=True,
        )
        self._dz_excel.setToolTip(
            "Planilha comercial com aba 'GRADE DE ATIVAÇÃO'.\n"
            "Colunas: SKU, DE (preço lista), POR (preço promo), VISIBLE, SELO."
        )
        excel_col.addWidget(self._dz_excel)
        inputs_row.addLayout(excel_col, 1)

        inputs_box_layout.addLayout(inputs_row)

        # 4 – Planilha BO (Kits) — opcional (BRD-007)
        bo_col = QVBoxLayout()
        bo_lbl = QLabel("Planilha BO (Kits)  —  opcional")
        bo_lbl.setObjectName("label_section")
        bo_col.addWidget(bo_lbl)
        self._dz_bo = DropZone(
            "Excel do BO (opcional)\n(COD_VENDA_PAI / FILHO / QUANTIDADE)",
            "Excel (*.xlsx *.xlsm *.xls)",
            default_icon="📦",  # planilha do BO não tem marca (NATBRA-/AVNBRA-);
                                # sem isto, BrandDetector cai em "unknown" (❓ vermelho)
        )
        self._dz_bo.setToolTip(
            "Opcional. Se anexada, o Auditor valida a composição dos Kits\n"
            "(Pai/Filho/Quantidade) do Salesforce contra a planilha do BO.\n"
            "Sem ela, a auditoria roda normalmente, apenas sem o check de Kits."
        )
        bo_col.addWidget(self._dz_bo)
        inputs_box_layout.addLayout(bo_col)

        layout.addWidget(inputs_box)

        # ── Bloco 2: Barra de Ações (rola junto com a página) ─────────────
        action_row = QHBoxLayout()
        action_row.setSpacing(12)

        self._btn_run = QPushButton("✓  Executar Auditoria")
        self._btn_run.setObjectName("btn_primary")
        self._btn_run.setFixedWidth(180)
        self._btn_run.clicked.connect(self._run)
        action_row.addWidget(self._btn_run)

        self._btn_clear = QPushButton("Limpar")
        self._btn_clear.setObjectName("btn_ghost")
        self._btn_clear.clicked.connect(self._clear)
        action_row.addWidget(self._btn_clear)

        action_row.addSpacing(20)

        # Brand filter pills
        self._btn_all    = self._make_filter_pill("Todos",   "all",    True)
        self._btn_natura = self._make_filter_pill("Natura",  "natura", False)
        self._btn_avon   = self._make_filter_pill("Avon",    "avon",   False)
        self._btn_all.setToolTip("Mostrar erros de todas as marcas")
        self._btn_natura.setToolTip("Filtrar apenas erros da marca Natura")
        self._btn_avon.setToolTip("Filtrar apenas erros da marca Avon")
        for b in (self._btn_all, self._btn_natura, self._btn_avon):
            action_row.addWidget(b)

        action_row.addStretch()

        self._btn_export = QPushButton("⬇  Relatório")
        self._btn_export.setObjectName("btn_secondary")
        self._btn_export.clicked.connect(self._export_excel)
        self._btn_export.setEnabled(False)
        self._btn_export.setToolTip("Exportar apenas divergências detectadas")
        action_row.addWidget(self._btn_export)

        self._btn_export_full = QPushButton("📊  Evidências (Full)")
        self._btn_export_full.setObjectName("btn_secondary")
        self._btn_export_full.clicked.connect(self._export_evidence)
        self._btn_export_full.setEnabled(False)
        self._btn_export_full.setToolTip("Exportar relatório completo de evidências (OK + ERRO)")
        action_row.addWidget(self._btn_export_full)

        self._btn_master_cert = QPushButton("🏅  Certificado Mestre")
        self._btn_master_cert.setObjectName("btn_cert")
        self._btn_master_cert.clicked.connect(self._export_master_audit)
        self._btn_master_cert.setEnabled(False)
        self._btn_master_cert.setToolTip(
            "Disponível apenas quando a auditoria retorna ZERO divergências.\n"
            "Gera Certificado de Conformidade (PDF) + Relatório de Evidências Master (Excel)."
        )
        action_row.addWidget(self._btn_master_cert)

        self._cert_status_lbl = QLabel("")
        self._cert_status_lbl.setObjectName("cert_status_waiting")
        self._cert_status_lbl.hide()
        action_row.addWidget(self._cert_status_lbl)

        self._btn_webhook = QPushButton("⊕  Enviar ao Google Chat")
        self._btn_webhook.setObjectName("btn_ghost")
        self._btn_webhook.clicked.connect(self._send_webhook)
        self._btn_webhook.setEnabled(False)
        action_row.addWidget(self._btn_webhook)

        layout.addLayout(action_row)

        # ── Capítulo 1: Auditoria Tradicional ─────────────────────────────
        # Cards (13 regras de negócio) + tabela detalhada + diagnóstico IA.
        # Vem sempre primeiro; a Validação de Kits (BO) fica em capítulo
        # separado ao final da página (ver mais abaixo). O corpo inteiro fica
        # em self._audit_body, recolhível via botão no cabeçalho — facilita
        # navegar entre os dois capítulos em telas menores.
        chapter_audit_row = QHBoxLayout()
        chapter_audit_header = QLabel("📋  AUDITORIA TRADICIONAL")
        chapter_audit_header.setStyleSheet(
            "font-size:13px;font-weight:800;color:#3b82f6;letter-spacing:0.5px;"
            "text-transform:uppercase;padding-top:4px;"
        )
        chapter_audit_row.addWidget(chapter_audit_header)
        chapter_audit_row.addStretch()
        self._btn_audit_collapse = self._make_collapse_button()
        self._btn_audit_collapse.clicked.connect(self._toggle_audit_chapter)
        chapter_audit_row.addWidget(self._btn_audit_collapse)
        layout.addLayout(chapter_audit_row)

        self._audit_body = QWidget()
        audit_body_layout = QVBoxLayout(self._audit_body)
        audit_body_layout.setContentsMargins(0, 0, 0, 0)
        audit_body_layout.setSpacing(20)
        layout.addWidget(self._audit_body)

        # ── Bloco 3: Painel de Divergências (cards) ───────────────────────
        cards_header = QLabel("Painel de Divergências")
        cards_header.setStyleSheet(
            "color:#888;font-size:11px;font-weight:700;text-transform:uppercase;"
        )
        audit_body_layout.addWidget(cards_header)

        # Error card dashboard (Flow-like grid)
        self._cards_container = QWidget()
        self._cards_grid = QGridLayout(self._cards_container)
        self._cards_grid.setContentsMargins(0, 0, 0, 0)
        self._cards_grid.setSpacing(10)

        # Pre-instantiate cards...
        for code, meta in ERROR_META.items():
            card = ErrorCard(
                code,
                icon=meta.get("icon", "·"),
                title=meta.get("title", code),
                impact=meta.get("impact", ""),
                desc=meta.get("desc", ""),
            )
            card.clicked_code.connect(self._on_card_clicked)
            self._error_cards[code] = card
            card.hide()

        audit_body_layout.addWidget(self._cards_container)

        # Empty-state (no divergences) — celebratory message shown when the
        # audit completes with zero errors. Hidden by default; toggled in
        # `_refresh_cards` based on the result stats.
        self._empty_state = QLabel()
        self._empty_state.setObjectName("auditor_empty_state")
        self._empty_state.setAlignment(Qt.AlignCenter)
        self._empty_state.setWordWrap(True)
        self._empty_state.setTextFormat(Qt.RichText)
        self._empty_state.setText(self._empty_state_html_ok())
        self._empty_state.hide()
        audit_body_layout.addWidget(self._empty_state)

        # ── Bloco 4: Tabela de Divergências Detalhadas ────────────────────
        table_header = QHBoxLayout()
        self._table_title = QLabel("Divergências Detalhadas")
        self._table_title.setStyleSheet("color:#888;font-size:11px;font-weight:700;text-transform:uppercase;")
        table_header.addWidget(self._table_title)
        table_header.addStretch()
        self._table_count_lbl = QLabel("")
        self._table_count_lbl.setObjectName("label_muted")
        table_header.addWidget(self._table_count_lbl)
        audit_body_layout.addLayout(table_header)

        self._table = QTableWidget(0, 5)
        self._table.setHorizontalHeaderLabels(["SKU", "Marca", "Tipo", "Detalhe", "Impt."])
        # Configure resize modes: all interactive, last column stretches
        for col in range(4):
            self._table.horizontalHeader().setSectionResizeMode(col, QHeaderView.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.verticalHeader().setVisible(False)
        # Cresce até MAX_TABLE_VISIBLE_ROWS; acima disso usa scroll próprio
        # (o resto da página rola no QScrollArea externo).
        self._table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        audit_body_layout.addWidget(self._table)

        # ── Bloco 5: Diagnóstico Estratégico — IA ─────────────────────────
        ai_header = QLabel("Diagnóstico Estratégico — IA")
        ai_header.setStyleSheet("font-size:11px;font-weight:700;color:#888;text-transform:uppercase;")
        audit_body_layout.addWidget(ai_header)

        self._ai_browser = QTextBrowser()
        self._ai_browser.setObjectName("ai_panel")
        self._ai_browser.setOpenExternalLinks(False)
        self._ai_browser.setPlaceholderText("Diagnóstico estratégico aparecerá aqui…")
        self._ai_browser.setMinimumHeight(360)
        self._ai_browser.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        audit_body_layout.addWidget(self._ai_browser)

        # ── Capítulo 2: Validação de Kits (BO) — BRD-007 ──────────────────
        # Reproduz a experiência da antiga tela Cadastro → Validação de Kits:
        # KPIs (analisados/corretos/divergências), cards clicáveis por
        # subtipo de divergência, tabela SKU Pai/Status/Detalhe e download
        # do XML de Correção. Só aparece quando o BO é anexado. Fica DEPOIS
        # do capítulo de Auditoria Tradicional inteiro (cards + tabela + IA),
        # como um segundo capítulo separado da mesma página de rolagem —
        # filtros e tabela são independentes dos do capítulo tradicional.
        # self._kit_chapter agrupa divisor+cabeçalho+corpo: some por inteiro
        # quando não há planilha BO anexada (_refresh_kit_panel), e o corpo
        # (self._kit_panel) é recolhível via botão, independente do capítulo
        # tradicional.
        self._kit_chapter = QWidget()
        kit_chapter_layout = QVBoxLayout(self._kit_chapter)
        kit_chapter_layout.setContentsMargins(0, 0, 0, 0)
        kit_chapter_layout.setSpacing(20)

        kit_chapter_sep = Divider()
        kit_chapter_sep.setFixedHeight(2)
        kit_chapter_sep.setStyleSheet("background-color:#3b82f6;")
        kit_chapter_layout.addWidget(kit_chapter_sep)

        chapter_kit_row = QHBoxLayout()
        chapter_kit_header = QLabel("📦  VALIDAÇÃO DE KITS (BO)")
        chapter_kit_header.setStyleSheet(
            "font-size:13px;font-weight:800;color:#3b82f6;letter-spacing:0.5px;"
            "text-transform:uppercase;padding-top:4px;"
        )
        chapter_kit_row.addWidget(chapter_kit_header)
        chapter_kit_row.addStretch()
        self._btn_kit_collapse = self._make_collapse_button()
        self._btn_kit_collapse.clicked.connect(self._toggle_kit_chapter)
        chapter_kit_row.addWidget(self._btn_kit_collapse)
        kit_chapter_layout.addLayout(chapter_kit_row)

        self._kit_panel = QWidget()
        kit_layout = QVBoxLayout(self._kit_panel)
        kit_layout.setContentsMargins(0, 0, 0, 0)
        kit_layout.setSpacing(10)

        kit_header = QLabel("Composição de Kits (BO)")
        kit_header.setStyleSheet(
            "font-size:11px;font-weight:700;color:#888;text-transform:uppercase;"
        )
        kit_layout.addWidget(kit_header)

        kit_stats_row = QHBoxLayout()
        kit_stats_row.setSpacing(12)
        self._kit_stat_total = self._make_stat_card("0", "Kits Analisados", "#00a1e0")
        self._kit_stat_ok    = self._make_stat_card("0", "Kits Corretos",   "#2ecc71")
        self._kit_stat_erro  = self._make_stat_card("0", "Divergências",    "#e74c3c")
        for card in (self._kit_stat_total, self._kit_stat_ok, self._kit_stat_erro):
            kit_stats_row.addWidget(card)
        kit_layout.addLayout(kit_stats_row)

        # Cards clicáveis por subtipo de divergência de kit (BRD-007). Espelha
        # o dashboard tradicional (ErrorCard + toggle de filtro), mas com
        # estado e alvo de filtro PRÓPRIOS (self._kit_active_filters /
        # self._kit_table) — nunca cruza com self._active_filters/self._table.
        kit_cards_row = QHBoxLayout()
        kit_cards_header = QLabel("Divergências por Subtipo")
        kit_cards_header.setStyleSheet(
            "font-size:11px;font-weight:700;color:#888;text-transform:uppercase;"
        )
        kit_cards_row.addWidget(kit_cards_header)
        kit_cards_row.addStretch()
        # Pills de marca PRÓPRIAS da seção de kits — não mexem no filtro de
        # marca do dashboard tradicional (self._brand_filter).
        self._btn_kit_all    = self._make_kit_filter_pill("Todos",  "all",    True)
        self._btn_kit_natura = self._make_kit_filter_pill("Natura", "natura", False)
        self._btn_kit_avon   = self._make_kit_filter_pill("Avon",   "avon",   False)
        self._btn_kit_all.setToolTip("Mostrar kits de todas as marcas")
        self._btn_kit_natura.setToolTip("Filtrar apenas kits da marca Natura")
        self._btn_kit_avon.setToolTip("Filtrar apenas kits da marca Avon")
        for b in (self._btn_kit_all, self._btn_kit_natura, self._btn_kit_avon):
            kit_cards_row.addWidget(b)
        kit_layout.addLayout(kit_cards_row)

        self._kit_cards_container = QWidget()
        self._kit_cards_grid = QGridLayout(self._kit_cards_container)
        self._kit_cards_grid.setContentsMargins(0, 0, 0, 0)
        self._kit_cards_grid.setSpacing(10)

        for kind, meta in KIT_ERROR_META.items():
            card = ErrorCard(
                kind,
                icon=meta.get("icon", "·"),
                title=meta.get("title", kind),
                impact=meta.get("impact", ""),
                desc=meta.get("desc", ""),
            )
            card.clicked_code.connect(self._on_kit_card_clicked)
            self._kit_error_cards[kind] = card
            card.hide()

        kit_layout.addWidget(self._kit_cards_container)

        self._kit_table = QTableWidget(0, 4)
        self._kit_table.setHorizontalHeaderLabels(["SKU Pai", "Marca", "Status", "Detalhe"])
        self._kit_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self._kit_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self._kit_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self._kit_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self._kit_table.setAlternatingRowColors(True)
        self._kit_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._kit_table.setSelectionBehavior(QTableWidget.SelectRows)
        self._kit_table.verticalHeader().setVisible(False)
        # Viewport próprio com scroll interno — evita que 100+ linhas (agora
        # de altura variável, por causa do word wrap) empurrem os botões de
        # export pra muito longe ou façam a página inteira rolar demais.
        self._kit_table.setMinimumHeight(320)
        self._kit_table.setMaximumHeight(480)
        self._kit_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        # Detalhe agora traz textos didáticos mais longos — sem quebra de linha
        # eles ficavam cortados na largura fixa da coluna. Word wrap + altura
        # de linha recalculada por conteúdo (_refresh_kit_table) resolve.
        self._kit_table.setWordWrap(True)
        kit_layout.addWidget(self._kit_table)

        kit_btn_row = QHBoxLayout()
        kit_btn_row.setSpacing(12)
        self._btn_kit_report = QPushButton("⬇  Relatório de Kits")
        self._btn_kit_report.setObjectName("btn_secondary")
        self._btn_kit_report.clicked.connect(self._export_kit_report)
        kit_btn_row.addWidget(self._btn_kit_report)
        self._btn_kit_correction = QPushButton("⬇  XML de Correção")
        self._btn_kit_correction.setObjectName("btn_secondary")
        self._btn_kit_correction.clicked.connect(self._export_kit_correction)
        kit_btn_row.addWidget(self._btn_kit_correction)
        kit_btn_row.addStretch()
        kit_layout.addLayout(kit_btn_row)

        kit_chapter_layout.addWidget(self._kit_panel)
        self._kit_chapter.hide()
        layout.addWidget(self._kit_chapter)

        # Folga ao final: mantém os blocos no topo quando a página é curta.
        layout.addStretch(1)

        # Altura inicial da tabela (vazia)
        self._adjust_table_height()

        # ── Barra de progresso fixa (fora do scroll, sempre visível) ──────
        progress_container = QWidget()
        progress_layout = QVBoxLayout(progress_container)
        progress_layout.setContentsMargins(28, 4, 28, 8)
        progress_layout.setSpacing(4)

        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, 100)
        self._progress_bar.hide()
        self._progress_bar.setFixedHeight(6)  # Thinner, more modern
        progress_layout.addWidget(self._progress_bar)

        self._status_lbl = QLabel("")
        self._status_lbl.setObjectName("label_muted")
        self._status_lbl.hide()
        progress_layout.addWidget(self._status_lbl)

        root.addWidget(progress_container)

        # Conecta validadores de bloqueio imediato nos DropZones
        self._setup_dropzone_validators()

    # ── DropZone validators (bloqueio imediato no drag-and-drop) ──────────
    def _setup_dropzone_validators(self) -> None:
        """Registra funções validadoras e conecta o sinal de rejeição em cada DropZone."""
        self._dz_cat.set_validator(self._validate_cat_paths)
        self._dz_excel.set_validator(self._validate_excel_paths)
        self._dz_pb.set_validator(self._validate_pb_paths)
        self._dz_cat.file_rejected.connect(self._on_file_rejected)
        self._dz_excel.file_rejected.connect(self._on_file_rejected)
        self._dz_pb.file_rejected.connect(self._on_file_rejected)

    @staticmethod
    def _format_country_rejection(path: str, check) -> str:
        """Mensagem de rejeição por país incorreto (BRD — incidente Chile)."""
        name = Path(path).name
        if check.country_hint:
            local = f'Parece pertencer a: <b>{check.country_hint}</b>.'
        else:
            local = ('País/região não identificado, mas o ID não contém '
                      'marcador de Brasil ("-br" ou "brazil").')
        return (
            "Este arquivo não corresponde a uma loja Brasil.\n\n"
            f"<b>Arquivo:</b> {name}\n"
            f'<b>ID encontrado:</b> "{check.raw_id}"\n'
            f"{local}\n\n"
            "<b>Esperado:</b> um Catálogo/Pricebook Natura BR, Avon BR ou Minha Loja BR "
            "(ex.: natura-br-storefront-catalog, avon-br-storefront-catalog, "
            "cb-br-storefront-catalog).\n\n"
            "Verifique se você selecionou o arquivo correto antes de tentar novamente."
        )

    def _validate_cat_paths(self, paths: list[str]) -> "Optional[str]":
        """
        Valida a lista proposta de catálogos antes de qualquer commit no DropZone.
        Retorna mensagem de erro se o país estiver incorreto ou houver
        duplicidade de marca; None se ok.
        (A checagem de quantidade exata de 3 fica reservada para o _run().)
        """
        for path in paths:
            check = BrandDetector.check_catalog_file(path)
            if not check.ok:
                return self._format_country_rejection(path, check)

        brand_map: dict[str, str] = {}
        for path in paths:
            brands = BrandDetector.detect_single(path)
            brand_key = next(iter(brands)) if brands else "desconhecida"
            if brand_key in brand_map:
                brand_display = BrandDetector.get_combined_display_name({brand_key})
                return (
                    f"Erro de Unicidade: Foram detectados dois arquivos da mesma marca "
                    f"({brand_display}).\n\n"
                    f"• {Path(brand_map[brand_key]).name}\n"
                    f"• {Path(path).name}\n\n"
                    f"Por favor, verifique os catálogos."
                )
            brand_map[brand_key] = path
        return None

    def _validate_pb_paths(self, paths: list[str]) -> "Optional[str]":
        """Valida o Pricebook proposto (BRD — incidente Chile): rejeita se algum
        pricebook-id do arquivo não pertencer a uma loja Brasil."""
        for path in paths:
            for check in BrandDetector.check_pricebook_file(path):
                if not check.ok:
                    return self._format_country_rejection(path, check)
        return None

    def _validate_excel_paths(self, paths: list[str]) -> "Optional[str]":
        """
        Valida a lista proposta de grades Excel antes de qualquer commit no DropZone.
        Retorna mensagem de erro se houver duplicidade de marca; None se ok.
        """
        brand_map: dict[str, str] = {}
        for path in paths:
            brands = BrandDetector.detect_single(path)
            brand_key = next(iter(brands)) if brands else "desconhecida"
            if brand_key in brand_map:
                brand_display = BrandDetector.get_combined_display_name({brand_key})
                return (
                    f"Erro de Unicidade: Foram detectados dois arquivos de Grade "
                    f"da mesma marca ({brand_display}).\n\n"
                    f"• {Path(brand_map[brand_key]).name}\n"
                    f"• {Path(path).name}\n\n"
                    f"Por favor, verifique as planilhas de Grade."
                )
            brand_map[brand_key] = path
        return None

    def _on_file_rejected(self, message: str) -> None:
        """Exibe um aviso quando o DropZone rejeita um arquivo (país incorreto ou violação de unicidade)."""
        show_rejection_dialog(self, message, header="Este arquivo foi recusado")

    # ── Filter pills ──────────────────────────────────────────────────────
    def _make_filter_pill(self, text: str, key: str, checked: bool) -> QPushButton:
        btn = QPushButton(text)
        btn.setCheckable(True)
        btn.setChecked(checked)
        btn.setObjectName("btn_ghost")
        btn.setFixedHeight(24)
        btn.setStyleSheet("padding: 2px 10px; font-size: 11px; font-weight: 600;")
        btn.setProperty("filter_key", key)
        btn.clicked.connect(lambda: self._set_brand_filter(key))
        return btn

    def _set_brand_filter(self, key: str):
        self._brand_filter = key
        for btn in (self._btn_all, self._btn_natura, self._btn_avon):
            btn.setChecked(btn.property("filter_key") == key)
        self._refresh_cards()
        self._refresh_table()

    def _make_kit_filter_pill(self, text: str, key: str, checked: bool) -> QPushButton:
        """Gêmea de _make_filter_pill para a seção de kits — aquela está presa ao
        trio do dashboard tradicional e ao _set_brand_filter."""
        btn = QPushButton(text)
        btn.setCheckable(True)
        btn.setChecked(checked)
        btn.setObjectName("btn_ghost")
        btn.setFixedHeight(24)
        btn.setStyleSheet("padding: 2px 10px; font-size: 11px; font-weight: 600;")
        btn.setProperty("filter_key", key)
        btn.clicked.connect(lambda: self._set_kit_brand_filter(key))
        return btn

    def _set_kit_brand_filter_silent(self, key: str) -> None:
        """Reseta o estado/pills de marca da seção de kits sem repintar (usado
        quando ainda não há resultado)."""
        self._kit_brand_filter = key
        for btn in (self._btn_kit_all, self._btn_kit_natura, self._btn_kit_avon):
            btn.setChecked(btn.property("filter_key") == key)

    def _set_kit_brand_filter(self, key: str):
        """Filtra APENAS a seção de kits — não toca no dashboard tradicional."""
        self._set_kit_brand_filter_silent(key)
        self._refresh_kit_cards()
        self._refresh_kit_table()

    # ── Stat cards do painel de kits (BRD-007) ────────────────────────────
    @staticmethod
    def _make_stat_card(value: str, label: str, color: str) -> QWidget:
        card = QWidget()
        card.setStyleSheet(f"background:{color};border-radius:8px;padding:10px;")
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(2)
        val_lbl = QLabel(value)
        val_lbl.setAlignment(Qt.AlignCenter)
        val_lbl.setObjectName("_val")
        val_lbl.setStyleSheet("font-size:28px;font-weight:700;color:white;background:transparent;")
        lay.addWidget(val_lbl)
        name_lbl = QLabel(label)
        name_lbl.setAlignment(Qt.AlignCenter)
        name_lbl.setStyleSheet("font-size:11px;color:white;background:transparent;")
        lay.addWidget(name_lbl)
        return card

    @staticmethod
    def _set_stat(card: QWidget, value: int):
        lbl = card.findChild(QLabel, "_val")
        if lbl:
            lbl.setText(str(value))

    # ── Capítulos recolhíveis (Auditoria Tradicional / Validação de Kits) ──
    @staticmethod
    def _make_collapse_button() -> QPushButton:
        btn = QPushButton("−")
        btn.setFixedSize(26, 26)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setToolTip("Recolher/expandir esta seção")
        # O QSS global define padding/min-height genéricos para todo
        # QPushButton (botões grandes tipo "Executar Auditoria"). Widgets com
        # borda/raio via stylesheet passam a usar o box-model "estilizado" do
        # Qt, onde essas propriedades podem sobrepor setFixedSize — por isso
        # padding/margin/min/max precisam ser zerados/travados aqui também,
        # e não só o setFixedSize, senão o botão "estica" numa pílula.
        btn.setStyleSheet(
            "QPushButton {"
            "  font-size:15px; font-weight:700; color:#3b82f6;"
            "  border:1px solid #3b82f6; border-radius:13px;"
            "  background:transparent; padding:0px; margin:0px;"
            "  min-width:26px; max-width:26px;"
            "  min-height:26px; max-height:26px;"
            "}"
            "QPushButton:hover { background:rgba(59,130,246,30); }"
            "QPushButton:pressed { background:rgba(59,130,246,55); }"
        )
        return btn

    def _toggle_audit_chapter(self) -> None:
        self._audit_collapsed = not self._audit_collapsed
        self._audit_body.setVisible(not self._audit_collapsed)
        self._btn_audit_collapse.setText("+" if self._audit_collapsed else "−")

    def _toggle_kit_chapter(self) -> None:
        self._kit_collapsed = not self._kit_collapsed
        self._kit_panel.setVisible(not self._kit_collapsed)
        self._btn_kit_collapse.setText("+" if self._kit_collapsed else "−")

    # ── Empty-state do capítulo tradicional ────────────────────────────────
    @staticmethod
    def _empty_state_html_ok() -> str:
        """Auditoria tradicional rodou e não encontrou nenhuma divergência."""
        return (
            "<div style='padding:28px 20px;'>"
            "<div style='font-size:42px;line-height:1;'>✅</div>"
            "<div style='font-size:16px;font-weight:700;margin-top:12px;color:#22A06B;'>"
            "Catálogo 100% íntegro"
            "</div>"
            "<div style='font-size:12px;margin-top:6px;color:#888;'>"
            "Nenhuma divergência encontrada nas regras de negócio.<br>"
            "Pricebook, catálogos e planilhas estão alinhados."
            "</div>"
            "</div>"
        )

    @staticmethod
    def _empty_state_html_kit_only() -> str:
        """Modo só-kit (BO sem Pricebook): a auditoria tradicional NÃO rodou —
        "0 divergências" aqui não significa "íntegro", significa "não checado".
        Sem este aviso, o usuário lê o card verde como se preço/categoria/
        visibilidade tivessem sido validados, o que é falso."""
        return (
            "<div style='padding:28px 20px;'>"
            "<div style='font-size:42px;line-height:1;'>ℹ️</div>"
            "<div style='font-size:16px;font-weight:700;margin-top:12px;color:#3b82f6;'>"
            "Auditoria Tradicional não executada"
            "</div>"
            "<div style='font-size:12px;margin-top:6px;color:#888;'>"
            "Nenhum Pricebook foi anexado — rodou apenas a Validação de Kits "
            "(BO), no capítulo abaixo.<br>Preço, categorias e visibilidade "
            "NÃO foram conferidos. Anexe o Pricebook para auditar tudo."
            "</div>"
            "</div>"
        )

    # ── Run ───────────────────────────────────────────────────────────────
    def _run(self):
        excel_paths = self._dz_excel.file_paths
        pb_path     = self._dz_pb.file_path
        cat_paths   = self._dz_cat.file_paths
        bo_path     = self._dz_bo.file_path

        # Modo só-kit (BRD-007): planilha BO anexada SEM Pricebook → valida
        # apenas a composição de kits (catálogo × BO). Exige somente catálogos
        # (1+) e a planilha BO; dispensa Pricebook e a regra dos 3 catálogos.
        # Com Pricebook presente, segue a auditoria completa de sempre.
        kit_only = bool(bo_path) and not pb_path
        self._last_kit_only = kit_only

        missing = []
        if not cat_paths:
            missing.append("Catálogo(s) XML")
        if not kit_only and not pb_path:
            missing.append("Pricebook XML")
        # BRD-007: com BO anexado, a Grade é obrigatória — dela vêm o whitelist
        # e o CM ativo que guiam o cruzamento de kits (SKU+CM). Vale nos modos
        # só-kit e completo+BO.
        if bo_path and not excel_paths:
            missing.append("Grade de Ativação (Excel)")

        if missing:
            QMessageBox.warning(
                self, "Auditor",
                "Selecione os seguintes arquivos antes de executar:\n• " + "\n• ".join(missing)
            )
            return

        # Trava 0: País/loja correta (BRD — incidente Chile). Defesa em
        # profundidade — re-checa mesmo que já validado no drop.
        for path in cat_paths:
            check = BrandDetector.check_catalog_file(path)
            if not check.ok:
                self._on_file_rejected(self._format_country_rejection(path, check))
                return
        if pb_path:
            for check in BrandDetector.check_pricebook_file(pb_path):
                if not check.ok:
                    self._on_file_rejected(self._format_country_rejection(pb_path, check))
                    return

        # Trava 1: Quantidade exata de catálogos — só na auditoria completa.
        # No modo só-kit qualquer quantidade (1+) é aceita.
        if not kit_only and len(cat_paths) != 3:
            QMessageBox.warning(
                self, "Quantidade Inválida",
                f"Quantidade Inválida: A auditoria requer exatamente 3 catálogos "
                f"(Natura, Avon e ML) para garantir a integridade cross-brand.\n\n"
                f"Foram selecionados {len(cat_paths)} arquivo(s)."
            )
            return

        # Trava 2: Unicidade de marca nos catálogos
        cat_brand_map: dict[str, str] = {}
        for path in cat_paths:
            brands = BrandDetector.detect_single(path)
            brand_key = next(iter(brands)) if brands else "desconhecida"
            if brand_key in cat_brand_map:
                brand_display = BrandDetector.get_combined_display_name({brand_key})
                QMessageBox.warning(
                    self, "Erro de Unicidade",
                    f"Erro de Unicidade: Foram detectados dois arquivos da mesma marca "
                    f"({brand_display}).\n\n"
                    f"• {Path(cat_brand_map[brand_key]).name}\n"
                    f"• {Path(path).name}\n\n"
                    f"Por favor, verifique os catálogos."
                )
                return
            cat_brand_map[brand_key] = path

        # Trava 3: Unicidade de marca nas grades Excel (se enviadas)
        if excel_paths:
            excel_brand_map: dict[str, str] = {}
            for path in excel_paths:
                brands = BrandDetector.detect_single(path)
                brand_key = next(iter(brands)) if brands else "desconhecida"
                if brand_key in excel_brand_map:
                    brand_display = BrandDetector.get_combined_display_name({brand_key})
                    QMessageBox.warning(
                        self, "Erro de Unicidade",
                        f"Erro de Unicidade: Foram detectados dois arquivos de Grade "
                        f"da mesma marca ({brand_display}).\n\n"
                        f"• {Path(excel_brand_map[brand_key]).name}\n"
                        f"• {Path(path).name}\n\n"
                        f"Por favor, verifique as planilhas de Grade."
                    )
                    return
                excel_brand_map[brand_key] = path

        self._btn_run.setEnabled(False)
        self._btn_export.setEnabled(False)
        self._btn_webhook.setEnabled(False)
        self._table.setRowCount(0)
        self._ai_browser.clear()
        self._empty_state.hide()
        self._progress_bar.setValue(0)
        self._progress_bar.show()
        self._status_lbl.show()
        self._active_filters.clear()
        self._kit_active_filters.clear()
        self._set_kit_brand_filter_silent("all")

        # Reset cards
        for card in self._error_cards.values():
            card.update_counts(0, 0, 0)
            card.set_selected(False)
        for card in self._kit_error_cards.values():
            card.update_counts(0, 0, 0)
            card.set_selected(False)

        # Cursor de espera (mesmo padrão de main_window.py) — feedback visual
        # imediato de que o processamento está em andamento, evitando a
        # sensação de app travado em máquinas mais lentas. Restaurado em
        # _on_finished/_on_error (só um dos dois dispara por execução).
        QApplication.setOverrideCursor(Qt.WaitCursor)

        self._worker = AuditorWorker(excel_paths, pb_path, cat_paths, bo_path, self)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.error_msg.connect(self._on_error)
        self._worker.start()

    # ── Progress ──────────────────────────────────────────────────────────
    def _on_progress(self, pct: int, msg: str):
        self._progress_bar.setValue(pct)
        self._status_lbl.setText(msg)

    # ── Finished ──────────────────────────────────────────────────────────
    def _on_finished(self, result: AuditResult):
        QApplication.restoreOverrideCursor()
        self._result = result
        self._progress_bar.hide()
        self._status_lbl.hide()
        self._btn_run.setEnabled(True)

        # Exportações e webhook da auditoria tradicional dependem de
        # result.errors/stats, que ficam vazios no modo só-kit (kits não
        # entram em ERROR_META) — habilitá-los aqui geraria um XLSX sem
        # abas (crash) ou um webhook de "0 Divergências / Operação Saudável"
        # falso, escondendo divergências reais de kit. O capítulo de kits
        # tem seus próprios botões de exportação (Relatório/XML de Correção).
        traditional_enabled = not self._last_kit_only
        self._btn_export.setEnabled(traditional_enabled)
        self._btn_export_full.setEnabled(traditional_enabled)
        self._btn_webhook.setEnabled(traditional_enabled)
        tooltip = "" if traditional_enabled else "Indisponível no modo só-kit — use os botões do painel de Kits."
        self._btn_export.setToolTip(tooltip)
        self._btn_export_full.setToolTip(tooltip)
        self._btn_webhook.setToolTip(tooltip)

        total = result.stats.get("total", 0)
        # Certificado Mestre só faz sentido na auditoria completa: atesta
        # conformidade do catálogo inteiro. Numa execução só-kit ele seria
        # enganoso, então fica desabilitado nesse modo.
        if total == 0 and not self._last_kit_only:
            self._btn_master_cert.setEnabled(True)
            self._cert_status_lbl.setObjectName("cert_status_ready")
            self._cert_status_lbl.setText("✅")
            self._cert_status_lbl.setToolTip("Certificado Disponível")
        else:
            self._btn_master_cert.setEnabled(False)
            self._cert_status_lbl.setObjectName("cert_status_waiting")
            self._cert_status_lbl.setText("⏳")
            self._cert_status_lbl.setToolTip("Aguardando Conformidade")
        self._cert_status_lbl.show()
        self._cert_status_lbl.style().unpolish(self._cert_status_lbl)
        self._cert_status_lbl.style().polish(self._cert_status_lbl)

        self._refresh_cards()

        # Default: show all errors (clear selection)
        self._active_filters.clear()
        self._refresh_table()

        # Painel dedicado de kits (BRD-007)
        self._refresh_kit_panel()

        # AI diagnostic
        self._settings.sync()
        theme = str(self._settings.value("theme", "light"))
        if self._last_kit_only:
            # Modo só-kit: as 13 regras tradicionais nunca rodaram (insumos
            # vazios), então stats["by_type"] sempre dá 0 erros em tudo.
            # agent.generate_report() leria isso como "Operação Saudável /
            # 100% íntegro" — a mesma falsa afirmação do empty-state, só que
            # no painel de IA. Não chamamos o agente nesse modo.
            html = (
                '<div style="color:#3b82f6;font-weight:700;">ℹ️ Diagnóstico não '
                'aplicável neste modo</div>'
                '<div style="margin-top:8px;">Sem Pricebook anexado, só a '
                'Validação de Kits (BO) rodou. Preço, categorias e '
                'visibilidade não foram auditados — veja o resultado dos '
                'kits no capítulo abaixo. Anexe o Pricebook para um '
                'diagnóstico completo.</div>'
            )
        else:
            agent = AiAgent()
            html  = agent.generate_report(
                result.stats,
                brands_found=result.brands_found,
                total_excel_skus=result.total_excel_skus,
                theme=theme,
            )
        bg_html = "#fcfdfe" if theme == "light" else "#0e1118"
        fg_html = "#333333" if theme == "light" else "#c0cce0"

        self._ai_browser.setHtml(f"""
        <html><body style="background:{bg_html};color:{fg_html};
                   font-family:'Helvetica Neue', Arial, Helvetica;font-size:12px;
                   padding:8px;line-height:1.6">
        {html}
        </body></html>""")

        color = "#ef5350" if total > 0 else "#66bb6a"
        if p := self.parent():
            if hasattr(p, "show_status"):
                p.show_status(
                    f"Auditoria: {result.total_excel_skus} SKUs · "
                    f"{total} divergências detectadas"
                )

        QTimer.singleShot(0, self._adjust_table_height)

        # Log to History — contagens padronizadas (Tarefa 7): no modo só-kit
        # as 13 regras tradicionais não rodam, então os números vêm do
        # capítulo de kits (evita gravar "0 divergências" enganoso).
        brands = " / ".join(result.brands_found) if result.brands_found else "Desconhecida"
        kd = result.kit_data
        if self._last_kit_only and kd is not None:
            kit_stats = getattr(kd, "stats", {})
            kit_ok = kit_stats.get("ok", 0)
            kit_erro = kit_stats.get("erro", 0)
            kit_total = kit_stats.get("total", 0)
            HistoryEngine.add_entry(
                "Auditor",
                brands,
                f"Validação de Kits concluída: {kit_total} kits, {kit_erro} com erro.",
                ok_count=kit_ok,
                error_count=kit_erro,
                total=kit_total,
                breakdown={"📦 Kits com erro": kit_erro} if kit_erro else None,
            )
        else:
            # Breakdown por tipo de divergência com os rótulos de ERROR_META,
            # gravado já amigável (o dashboard não precisa conhecer os códigos).
            by_type = result.stats.get("by_type", {})
            breakdown = {
                f"{ERROR_META[code]['icon']} {ERROR_META[code]['title']}": data["total"]
                for code, data in by_type.items()
                if code in ERROR_META and data.get("total", 0) > 0
            }
            HistoryEngine.add_entry(
                "Auditor",
                brands,
                f"Auditoria concluída: {result.total_excel_skus} SKUs, {total} divergências.",
                ok_count=len(result.acertos),
                error_count=total,
                total=result.total_excel_skus,
                breakdown=breakdown or None,
            )

    # ── Painel dedicado de kits (BRD-007) ─────────────────────────────────
    def _refresh_kit_panel(self):
        """Popula o painel de Composição de Kits a partir de result.kit_data.
        Reproduz a antiga tela de Validação de Kits (KPIs + tabela + correção),
        mais os cards por subtipo (BRD-007 UX)."""
        kd = getattr(self._result, "kit_data", None) if self._result else None
        if kd is None:
            self._kit_chapter.hide()
            return

        # Resultado novo invalida qualquer seleção de filtro anterior.
        self._kit_active_filters.clear()

        self._refresh_kit_cards(kd)
        self._refresh_kit_table(kd)

        has_errors = kd.stats.get("erro", 0) > 0
        self._btn_kit_report.setEnabled(has_errors)
        self._btn_kit_correction.setEnabled(bool(kd.correction_xmls))
        self._kit_panel.setVisible(not self._kit_collapsed)
        self._kit_chapter.show()

    def _on_kit_card_clicked(self, kind: str) -> None:
        """Toggle de filtro independente do dashboard tradicional — nunca
        toca em self._active_filters nem em self._table."""
        if kind in self._kit_active_filters:
            self._kit_active_filters.remove(kind)
        else:
            self._kit_active_filters.add(kind)

        for k_kind, card in self._kit_error_cards.items():
            card.set_selected(k_kind in self._kit_active_filters)

        self._refresh_kit_table()

    def _kit_data(self):
        return getattr(self._result, "kit_data", None) if self._result else None

    def _kit_rows_da_marca(self, kd) -> list[dict]:
        """Linhas de kit respeitando o filtro de marca da seção (em kd.rows a
        marca vem capitalizada: 'Natura'/'Avon')."""
        if self._kit_brand_filter == "all":
            return list(kd.rows)
        return [r for r in kd.rows
                if str(r.get("brand", "")).lower() == self._kit_brand_filter]

    def _refresh_kit_cards(self, kd=None) -> None:
        """Popula KPIs e cards clicáveis por subtipo (kind), respeitando o
        filtro de marca da seção."""
        if kd is None:
            kd = self._kit_data()
        if kd is None:
            return

        # KPIs: usam o recorte por marca quando há filtro ativo.
        stats = kd.stats or {}
        if self._kit_brand_filter == "all":
            kpi = stats
        else:
            marca = self._kit_brand_filter.capitalize()
            kpi = (stats.get("by_brand") or {}).get(marca, {})
        self._set_stat(self._kit_stat_total, kpi.get("total", 0))
        self._set_stat(self._kit_stat_ok,    kpi.get("ok", 0))
        self._set_stat(self._kit_stat_erro,  kpi.get("erro", 0))

        counts_nat: dict[str, int] = {}
        counts_avn: dict[str, int] = {}
        for row in self._kit_rows_da_marca(kd):
            kind = row.get("kind")
            if kind is None:
                continue
            if str(row.get("brand", "")).lower() == "avon":
                counts_avn[kind] = counts_avn.get(kind, 0) + 1
            else:
                counts_nat[kind] = counts_nat.get(kind, 0) + 1

        for i in reversed(range(self._kit_cards_grid.count())):
            self._kit_cards_grid.itemAt(i).widget().setParent(None)

        visible_cards = []
        for kind, card in self._kit_error_cards.items():
            nat = counts_nat.get(kind, 0)
            avn = counts_avn.get(kind, 0)
            card.update_counts(nat + avn, nat, avn)
            if nat + avn > 0:
                visible_cards.append(card)
                card.show()
            else:
                card.hide()

        for idx, card in enumerate(visible_cards):
            row, col = divmod(idx, 4)
            self._kit_cards_grid.addWidget(card, row, col)

        self._kit_cards_container.setVisible(len(visible_cards) > 0)

    def _refresh_kit_table(self, kd=None) -> None:
        """Popula self._kit_table filtrando por marca e pelos cards de subtipo
        selecionados."""
        if kd is None:
            kd = self._kit_data()
        if kd is None:
            return

        base_rows = self._kit_rows_da_marca(kd)
        rows = base_rows
        if self._kit_active_filters:
            rows = [r for r in rows if r.get("kind") in self._kit_active_filters]

        self._kit_table.setRowCount(0)
        if rows:
            self._kit_table.setRowCount(len(rows))
            for i, e in enumerate(rows):
                brand_val   = str(e.get("brand", ""))
                status_val  = str(e.get("status", ""))
                brand_item  = QTableWidgetItem(brand_val)
                status_item = QTableWidgetItem(status_val)
                # Mesmas cores de marca da tabela do auditor tradicional
                if brand_val.lower() == "natura":
                    brand_item.setForeground(QColor("#FF8050"))
                elif brand_val.lower() == "avon":
                    brand_item.setForeground(QColor("#bb88ff"))
                if "Ausente" in status_val:
                    status_item.setForeground(QColor("#ff7b72"))
                elif "Divergente" in status_val:
                    status_item.setForeground(QColor("#e2b93d"))
                # SKU completo com prefixo (NATBRA-/AVNBRA-), não só o número —
                # independente da coluna Marca, ajuda o auditor a localizar o
                # produto direto no Salesforce/catálogo.
                detail_item = QTableWidgetItem(str(e.get("detail", "")))
                self._kit_table.setItem(i, 0, QTableWidgetItem(str(e.get("sku", e.get("pai", "")))))
                self._kit_table.setItem(i, 1, brand_item)
                self._kit_table.setItem(i, 2, status_item)
                self._kit_table.setItem(i, 3, detail_item)
            # Word wrap muda a altura necessária de cada linha; recalcula
            # depois de popular para não cortar o texto de Detalhe. Adiado via
            # singleShot(0): chamado na hora, a coluna Detalhe (Stretch) ainda
            # não tinha assumido a largura final, e o cálculo de quebra de
            # linha usava uma largura estreita — inflando a altura da linha
            # muito além do necessário (mesmo bug que _adjust_table_height já
            # contorna do mesmo jeito para a tabela tradicional).
            QTimer.singleShot(0, self._kit_table.resizeRowsToContents)
        elif kd.rows:
            # Havia divergências, mas o recorte atual (marca e/ou subtipo) não
            # tem nenhuma — não pode exibir "Sucesso", seria enganoso.
            self._kit_table.setRowCount(1)
            for c in range(3):
                self._kit_table.setItem(0, c, QTableWidgetItem(""))
            self._kit_table.setItem(0, 3, QTableWidgetItem(
                "Nenhuma linha para os filtros selecionados."))
        else:
            self._kit_table.setRowCount(1)
            ok_item = QTableWidgetItem("✅ Sucesso")
            ok_item.setForeground(QColor("#2ecc71"))
            self._kit_table.setItem(0, 0, QTableWidgetItem(""))
            self._kit_table.setItem(0, 1, QTableWidgetItem(""))
            self._kit_table.setItem(0, 2, ok_item)
            self._kit_table.setItem(0, 3, QTableWidgetItem("Todos os kits conferem!"))

    @staticmethod
    def _kit_row_to_export(r: dict) -> dict:
        """Linha de divergência no formato rico do relatório."""
        meta = KIT_ERROR_META.get(r.get("kind", ""), {})
        return {
            "Marca":            r.get("brand", ""),
            "SKU Pai":          r.get("sku", ""),
            "SKU Pai (núm)":    r.get("pai", ""),
            "Tipo":             meta.get("title", r.get("kind", "")),
            "Status":           r.get("status", ""),
            "SKU Filho":        r.get("filho", ""),
            "Qtd SF":           r.get("qtd_sf", ""),
            "Qtd BO":           r.get("qtd_bo", ""),
            "CM (Grade)":       r.get("cm_grade", ""),
            "CM (BO)":          r.get("cm_bo", ""),
            "Detalhe":          r.get("detail", ""),
        }

    @staticmethod
    def _style_kit_resumo_sheet(ws, df: pd.DataFrame, n_kpis: int) -> None:
        """Estiliza a aba Resumo: header em negrito, as 3 linhas de KPI geral
        coloridas (mesma cor dos stat cards da tela), separador antes do
        detalhamento por subtipo, e a coluna Total destacada (negrito + fundo
        cinza-claro) — para não depender de repetir a palavra "TOTAL" como
        valor ambíguo numa coluna Marca."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="0D1F3D")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill  = PatternFill("solid", fgColor="F0F0F0")
        kpi_colors  = {
            "Kits Analisados":  "0072B2",
            "Kits Corretos":    "1E8449",
            "Kits Divergentes": "C0392B",
        }
        thin  = Side(style="thin", color="D9D9D9")
        thick = Side(style="medium", color="888888")

        n_cols = len(df.columns)
        total_col_idx = n_cols  # última coluna é sempre "Total"

        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(2, len(df) + 2):
            indicador = ws.cell(row=row_idx, column=1).value
            is_kpi = (row_idx - 1) <= n_kpis
            # Linha logo abaixo do bloco de KPIs gerais: separador visual
            # antes do detalhamento por subtipo de divergência.
            is_primeira_do_detalhe = (row_idx - 1) == n_kpis + 1
            top_side = thick if is_primeira_do_detalhe else thin
            row_border = Border(left=thin, right=thin, top=top_side, bottom=thin)
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = row_border
                if col == total_col_idx:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            if is_kpi and indicador in kpi_colors:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, color=kpi_colors[indicador])

        for col in range(1, n_cols + 1):
            header_len = len(str(df.columns[col - 1]))
            max_len = max([header_len] + [len(str(v)) for v in df.iloc[:, col - 1]])
            ws.column_dimensions[get_column_letter(col)].width = max(14, max_len + 4)

        ws.freeze_panes = "A2"

    def _export_kit_report(self):
        """Relatório de kits: aba Resumo (matriz Indicador × Marca, com Total
        como coluna — não como valor repetido) + uma aba por marca com as
        divergências do recorte filtrado na tela."""
        kd = self._kit_data()
        if kd is None or not kd.rows:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Relatório de Kits", "Relatorio_Kits.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        path = get_unique_path(path)
        try:
            stats = kd.stats or {}
            by_brand = stats.get("by_brand") or {}
            # Ordem fixa (Natura antes de Avon) — só as marcas com kits nesta auditoria.
            marcas = [m for m in ("Natura", "Avon") if m in by_brand] or ["Natura"]

            def linha_resumo(indicador: str, por_marca: dict, total: int) -> dict:
                row: dict[str, object] = {"Indicador": indicador}
                for m in marcas:
                    row[m] = por_marca.get(m, 0)
                row["Total"] = total
                return row

            # ── Bloco 1: KPIs gerais. Sempre o quadro COMPLETO, ignorando os
            # filtros da tela, para o auditor não perder a visão geral.
            resumo_rows = [
                linha_resumo("Kits Analisados",
                             {m: by_brand.get(m, {}).get("total", 0) for m in marcas},
                             stats.get("total", 0)),
                linha_resumo("Kits Corretos",
                             {m: by_brand.get(m, {}).get("ok", 0) for m in marcas},
                             stats.get("ok", 0)),
                linha_resumo("Kits Divergentes",
                             {m: by_brand.get(m, {}).get("erro", 0) for m in marcas},
                             stats.get("erro", 0)),
            ]
            n_kpis = len(resumo_rows)

            # ── rows_filtradas: recorte por marca + subtipo (cards
            # selecionados na tela). Fonte única de verdade usada tanto pelo
            # Bloco 2 do Resumo quanto pelas abas por marca abaixo (BRD-009).
            rows_filtradas = self._kit_rows_da_marca(kd)
            if self._kit_active_filters:
                rows_filtradas = [r for r in rows_filtradas
                                  if r.get("kind") in self._kit_active_filters]

            # ── Bloco 2: detalhamento por subtipo, respeitando os cards
            # selecionados na tela (só os que tiveram achado no recorte
            # atual). Sem filtro ativo (nenhum card = "Todos"), comportamento
            # idêntico ao anterior (mostra todos os subtipos com achado).
            for kind, meta in KIT_ERROR_META.items():
                total_kind = sum(1 for r in rows_filtradas if r.get("kind") == kind)
                if not total_kind:
                    continue
                por_marca = {m: sum(1 for r in rows_filtradas
                                     if r.get("kind") == kind and r.get("brand") == m)
                             for m in marcas}
                resumo_rows.append(linha_resumo(meta.get("title", kind), por_marca, total_kind))

            # Aviso explícito quando há filtro ativo — quem abrir o Excel sem
            # ver a tela não deve achar que o Bloco 2 é o total geral.
            if self._kit_active_filters:
                titles = [KIT_ERROR_META.get(k, {}).get("title", k)
                          for k in self._kit_active_filters]
                resumo_rows.append({"Indicador": f"⚠ Filtro ativo na exportação: {', '.join(titles)}"})

            df_resumo = pd.DataFrame(resumo_rows)

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
                self._style_kit_resumo_sheet(writer.sheets["Resumo"], df_resumo, n_kpis)
                for m in marcas:
                    linhas = [self._kit_row_to_export(r) for r in rows_filtradas
                              if r.get("brand") == m]
                    if not linhas:
                        continue
                    sheet = m
                    for char in r"/\?*:[]":
                        sheet = sheet.replace(char, "_")
                    pd.DataFrame(linhas).to_excel(writer, sheet_name=sheet[:31], index=False)

            QMessageBox.information(self, "Exportado", f"Relatório salvo em:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Erro ao Exportar", str(exc))

    def _export_kit_correction(self):
        # Nota: assim como no XML combinado antigo, a exportação ignora o
        # filtro de marca da seção de kits (_kit_brand_filter) — esse filtro
        # é só de visualização, o export sempre sai com todas as marcas que
        # tiverem divergência. Mantido igual ao comportamento anterior.
        kd = getattr(self._result, "kit_data", None) if self._result else None
        if kd is None or not kd.correction_xmls:
            return

        marcas = sorted(kd.correction_xmls)
        if len(marcas) == 1:
            path, _ = QFileDialog.getSaveFileName(
                self, "Salvar XML de Correção", "CORRECAO_KITS_BRASIL.xml", "XML (*.xml)"
            )
            if not path:
                return
            try:
                Path(path).write_text(kd.correction_xmls[marcas[0]], encoding="utf-8")
                QMessageBox.information(self, "Exportado", f"XML salvo em:\n{path}")
            except Exception as exc:
                QMessageBox.critical(self, "Erro ao Exportar", str(exc))
            return

        # Mais de uma marca: cada uma tem um catalog-id diferente, então não
        # dá pra combinar num XML só — salva um arquivo por marca na pasta escolhida.
        dir_path = QFileDialog.getExistingDirectory(self, "Escolher pasta para os XMLs de Correção")
        if not dir_path:
            return
        try:
            saved = []
            for marca in marcas:
                file_path = Path(dir_path) / f"CORRECAO_KITS_BRASIL_{marca.lower()}.xml"
                file_path.write_text(kd.correction_xmls[marca], encoding="utf-8")
                saved.append(str(file_path))
            QMessageBox.information(self, "Exportado", "XMLs salvos:\n" + "\n".join(saved))
        except Exception as exc:
            QMessageBox.critical(self, "Erro ao Exportar", str(exc))

    # ── Altura dinâmica da tabela ─────────────────────────────────────────
    def _adjust_table_height(self) -> None:
        """Ajusta a altura da tabela para caber até MAX_TABLE_VISIBLE_ROWS
        linhas sem corte; acima disso a tabela ganha scroll próprio enquanto o
        resto da página continua rolando no QScrollArea externo."""
        rows = self._table.rowCount()

        header_h = self._table.horizontalHeader().height()
        if header_h <= 0:
            header_h = self._table.horizontalHeader().sizeHint().height()

        row_h = self._table.verticalHeader().defaultSectionSize()
        if rows > 0:
            hint = self._table.sizeHintForRow(0)
            if hint > 0:
                row_h = hint

        # Mínimo de linhas mostradas (placeholder agradável quando vazia)
        visible_rows = min(max(rows, MIN_TABLE_VISIBLE_ROWS), MAX_TABLE_VISIBLE_ROWS)
        frame = 2 * self._table.frameWidth()
        self._table.setFixedHeight(header_h + row_h * visible_rows + frame)

    # ── Error ─────────────────────────────────────────────────────────────
    def _on_error(self, msg: str):
        QApplication.restoreOverrideCursor()
        self._btn_run.setEnabled(True)
        self._progress_bar.hide()
        self._status_lbl.hide()
        HistoryEngine.add_entry(
            "Auditor", "N/A", f"Falha na execução: {msg[:200]}", status="falha"
        )
        QMessageBox.critical(self, "Erro — Auditor", msg)

    # ── Card click ────────────────────────────────────────────────────────
    def _on_card_clicked(self, code: str):
        # Toggle filter
        if code in self._active_filters:
            self._active_filters.remove(code)
        else:
            self._active_filters.add(code)

        # Update visuals
        for c_code, card in self._error_cards.items():
            card.set_selected(c_code in self._active_filters)

        self._refresh_table()

    # ── Table refresh ─────────────────────────────────────────────────────
    def _refresh_cards(self):
        if not self._result:
            return

        for i in reversed(range(self._cards_grid.count())):
            self._cards_grid.itemAt(i).widget().setParent(None)

        visible_cards = []
        for code, card in self._error_cards.items():
            bt = self._result.stats.get("by_type", {}).get(code, {})
            
            if self._brand_filter == "natura":
                nat = bt.get("natura", 0)
                card.update_counts(nat, nat, 0)
                if nat > 0: visible_cards.append(card)
            elif self._brand_filter == "avon":
                avn = bt.get("avon", 0)
                card.update_counts(avn, 0, avn)
                if avn > 0: visible_cards.append(card)
            else:
                total = bt.get("total", 0)
                card.update_counts(total, bt.get("natura", 0), bt.get("avon", 0))
                if total > 0: visible_cards.append(card)
                
            if card in visible_cards:
                card.show()
            else:
                card.hide()

        for idx, card in enumerate(visible_cards):
            row, col = divmod(idx, 4)
            self._cards_grid.addWidget(card, row, col)

        self._cards_container.setVisible(len(visible_cards) > 0)

        # Empty-state: modo só-kit NUNCA passa pelas 13 regras tradicionais
        # (rodam com insumos vazios), então total sempre dá 0 mas isso não
        # significa "íntegro" — significa "não checado". Mostrar o card verde
        # de sucesso nesse caso seria uma falsa afirmação de conformidade.
        if self._last_kit_only:
            self._empty_state.setText(self._empty_state_html_kit_only())
            self._empty_state.setVisible(True)
        else:
            # Só aqui, sem modo só-kit, "total == 0" de fato significa que a
            # auditoria rodou e não achou divergência (não é filtro escondendo).
            total_errors = self._result.stats.get("total", 0)
            self._empty_state.setText(self._empty_state_html_ok())
            self._empty_state.setVisible(total_errors == 0)

    def _refresh_table(self):
        if not self._result:
            return

        rows_to_show: list[dict] = []
        errors = self._result.errors

        # Which error types to include (empty set means show all)
        codes = list(self._active_filters) if self._active_filters else list(errors.keys())

        for code in codes:
            df = errors.get(code)
            if df is None or df.empty:
                continue
            meta = ERROR_META.get(code, {})
            for _, row in df.iterrows():
                brand = str(row.get("brand", "")).lower()
                if self._brand_filter != "all" and brand != self._brand_filter:
                    continue
                rows_to_show.append(
                    {
                        "sku":    str(row.get("sku", "")),
                        "brand":  brand,
                        "code":   code,
                        "title":  meta.get("title", code),
                        "detail": str(row.get("detail", "")),
                        "impact": meta.get("impact", ""),
                    }
                )

        # Truncate
        truncated = len(rows_to_show) > MAX_TABLE_ROWS
        visible   = rows_to_show[:MAX_TABLE_ROWS]

        self._table.setRowCount(0)
        self._table.setRowCount(len(visible))

        for r_idx, row in enumerate(visible):
            sku_item = QTableWidgetItem(row["sku"])
            sku_item.setFont(self._table.font())

            brand_item  = QTableWidgetItem(row["brand"].capitalize())
            error_item  = QTableWidgetItem(row["title"])
            detail_item = QTableWidgetItem(row["detail"])
            impact_item = QTableWidgetItem(row["impact"])

            # Brand color hint
            if row["brand"] == "natura":
                brand_item.setForeground(QColor("#FF8050"))
            elif row["brand"] == "avon":
                brand_item.setForeground(QColor("#bb88ff"))

            for c_idx, item in enumerate(
                [sku_item, brand_item, error_item, detail_item, impact_item]
            ):
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self._table.setItem(r_idx, c_idx, item)

        # Update title
        has_filters = bool(self._active_filters) or self._brand_filter != "all"
        filter_indicator = "🔽 " if has_filters else ""

        if not self._active_filters:
            title_text = "Todos os Erros"
        else:
            names = [ERROR_META.get(c, {}).get("title", c) for c in self._active_filters]
            if len(names) > 2:
                title_text = f"{len(names)} filtros ativos"
            else:
                title_text = " + ".join(names)

        brand_suffix = "" if self._brand_filter == "all" else f"  ·  {self._brand_filter.capitalize()}"
        self._table_title.setText(f"{filter_indicator}{title_text}{brand_suffix}")

        # Calculate total errors from selected cards (respecting brand filter)
        cards_total = 0
        for code in (self._active_filters if self._active_filters else errors.keys()):
            by_type = self._result.stats.get("by_type", {}).get(code, {})
            if self._brand_filter == "natura":
                cards_total += by_type.get("natura", 0)
            elif self._brand_filter == "avon":
                cards_total += by_type.get("avon", 0)
            else:
                cards_total += by_type.get("total", 0)

        # Display count from selected cards, with note if table is truncated
        count_text = f"{cards_total} erros encontrados"
        if truncated:
            count_text += f"  •  Exibindo {MAX_TABLE_ROWS} primeiros"

        # Set tooltip to explain that this is the sum of selected cards
        if has_filters:
            tooltip_text = (
                f"Total de {cards_total} erros dos cards selecionados.\n"
                f"Filtros: {title_text}{brand_suffix}\n"
                f"Tabela exibe os primeiros {len(visible)} para melhor visualização."
                if not truncated
                else
                f"Total de {cards_total} erros dos cards selecionados.\n"
                f"Filtros: {title_text}{brand_suffix}\n"
                f"Tabela mostra apenas os {MAX_TABLE_ROWS} primeiros para performance."
            )
        else:
            tooltip_text = (
                f"Total de {cards_total} erros encontrados.\n"
                f"Clique em um card de erro para filtrar por tipo ou use os botões de marca para filtrar."
            )

        self._table_count_lbl.setText(count_text)
        self._table_count_lbl.setToolTip(tooltip_text)

        # Ajusta a altura para caber o conteúdo (até o limite de linhas)
        self._adjust_table_height()

    # ── Export ────────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self._result:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Relatório", "AUDIT_REPORT.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        path = get_unique_path(path)
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # Filter codes to export: active selections OR all non-empty ones
                codes = list(self._active_filters) if self._active_filters else list(self._result.errors.keys())

                for code in codes:
                    df = self._result.errors.get(code)
                    if df is None or df.empty:
                        continue

                    # Apply brand filter if active
                    if self._brand_filter != "all":
                        df = df[df["brand"].str.lower() == self._brand_filter.lower()]
                    
                    if df.empty:
                        continue

                    title = ERROR_META.get(code, {}).get("title", code)
                    # Sanitize sheet name: remove invalid chars / \ ? * : [ ]
                    for char in r"/\?*:[]":
                        title = title.replace(char, "_")
                    sheet = title[:31]
                    df.to_excel(writer, sheet_name=sheet, index=False)

                # Aba de Acertos
                acertos_df = getattr(self._result, "acertos", None)
                if acertos_df is not None and not acertos_df.empty:
                    df_ac = acertos_df.copy()
                    if self._brand_filter != "all":
                        df_ac = df_ac[df_ac["brand"].str.lower() == self._brand_filter.lower()]
                    if not df_ac.empty:
                        df_ac.to_excel(writer, sheet_name="Acertos", index=False)

            QMessageBox.information(self, "Exportado", f"Relatório salvo em:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Erro ao Exportar", str(exc))

    # ── Webhook ───────────────────────────────────────────────────────────
    def _compute_filtered_stats(self) -> dict:
        """Rebuild the stats dict respecting active card filters and brand filter.

        Mirrors the filter logic in `_refresh_table` so the webhook payload
        matches exactly what the user sees on screen. Returns the same shape
        as `AuditResult.stats` so `AiAgent.generate_gchat_report` works unchanged.
        """
        by_type: dict[str, dict] = {}
        by_brand = {"natura": 0, "avon": 0}
        total = 0

        if not self._result:
            return {"total": 0, "by_type": {}, "by_brand": by_brand}

        errors = self._result.errors
        codes = list(self._active_filters) if self._active_filters else list(errors.keys())

        for code in codes:
            df = errors.get(code)
            if df is None or df.empty:
                continue

            if self._brand_filter != "all":
                df = df[df["brand"].str.lower() == self._brand_filter.lower()]

            if df.empty:
                continue

            df_unique = df.drop_duplicates(subset=["sku"])
            brand_series = df_unique["brand"].astype(str).str.lower()
            nat = int((brand_series == "natura").sum())
            avn = int((brand_series == "avon").sum())
            tot = int(len(df_unique))

            by_type[code] = {"total": tot, "natura": nat, "avon": avn}
            by_brand["natura"] += nat
            by_brand["avon"]   += avn
            total += tot

        return {"total": total, "by_type": by_type, "by_brand": by_brand}

    def _send_webhook(self):
        if not self._result:
            return
        if self._last_kit_only:
            # Defesa em profundidade: o botão já fica desabilitado no modo
            # só-kit (_on_finished), mas result.errors/stats vêm vazios nesse
            # modo — enviar aqui reportaria "0 Divergências / Operação
            # Saudável" ao Google Chat mesmo com kits divergentes.
            QMessageBox.warning(
                self, "Google Chat",
                "Webhook indisponível no modo só-kit (sem Pricebook): kits divergentes "
                "não são contabilizados nas estatísticas gerais. Use o painel de Kits."
            )
            return
        url = self._settings.value("gchat_webhook", "")
        if not url:
            QMessageBox.warning(
                self, "Google Chat",
                "Configure a URL do webhook em Configurações antes de enviar."
            )
            return

        import requests
        stats   = self._compute_filtered_stats()
        total   = stats.get("total", 0)
        nat_err = stats.get("by_brand", {}).get("natura", 0)
        avn_err = stats.get("by_brand", {}).get("avon",   0)

        total_global = self._result.stats.get("total", 0)
        if total == 0 and total_global > 0:
            res = QMessageBox.question(
                self, "Google Chat",
                "O recorte atual não possui divergências, mas a auditoria global detectou erros.\n\n"
                "Deseja enviar um relatório de 'Operação Saudável' para este recorte?",
                QMessageBox.Yes | QMessageBox.No
            )
            if res == QMessageBox.No:
                return

        has_filter = bool(self._active_filters) or self._brand_filter != "all"
        if has_filter:
            brand_name = self._brand_filter.capitalize() if self._brand_filter != "all" else "Misto"
            subtitle = f"Relatório Filtrado ({brand_name}) · {total} divergências"
        else:
            subtitle = f"{self._result.total_excel_skus} SKUs auditados"

        agent = AiAgent()
        plain_ai = agent.generate_gchat_report(
            stats,
            brands_found=self._result.brands_found,
            total_excel_skus=self._result.total_excel_skus,
        )

        payload = {
            "cards": [
                {
                    "header": {
                        "title": "SIC — Relatório Estratégico",
                        "subtitle": subtitle,
                    },
                    "sections": [
                        {
                            "widgets": [
                                {"keyValue": {"topLabel": "Total de Divergências",  "content": str(total)}},
                                {"keyValue": {"topLabel": "Erros Natura",           "content": str(nat_err)}},
                                {"keyValue": {"topLabel": "Erros Avon",             "content": str(avn_err)}},
                            ]
                        },
                        {"widgets": [{"textParagraph": {"text": plain_ai[:3000]}}]},
                    ],
                }
            ]
        }

        try:
            resp = requests.post(url, json=payload, timeout=10)
            if resp.status_code == 200:
                QMessageBox.information(self, "Google Chat", "Relatório enviado com sucesso!")
            else:
                QMessageBox.warning(self, "Google Chat", f"Status {resp.status_code}: {resp.text[:200]}")
        except Exception as exc:
            QMessageBox.critical(self, "Erro", str(exc))

    # ── Clear ─────────────────────────────────────────────────────────────
    def _clear(self):
        self._dz_pb.clear()
        self._dz_cat.clear()
        self._dz_excel.clear()
        self._dz_bo.clear()
        self._table.setRowCount(0)
        self._ai_browser.clear()
        self._progress_bar.setValue(0)
        self._progress_bar.hide()
        self._status_lbl.hide()
        self._btn_export.setEnabled(False)
        self._btn_export_full.setEnabled(False)
        self._btn_webhook.setEnabled(False)
        self._btn_master_cert.setEnabled(False)
        self._cert_status_lbl.hide()
        self._result = None
        self._last_kit_only = False
        self._active_filters.clear()
        self._kit_active_filters.clear()
        self._set_kit_brand_filter_silent("all")
        self._brand_filter  = "all"
        self._btn_all.setChecked(True)
        self._btn_natura.setChecked(False)
        self._btn_avon.setChecked(False)
        self._cards_container.hide()
        self._empty_state.hide()
        self._kit_chapter.hide()
        self._kit_table.setRowCount(0)
        self._kit_cards_container.hide()
        for card in self._error_cards.values():
            card.hide()
            card.update_counts(0, 0, 0)
            card.set_selected(False)
        for card in self._kit_error_cards.values():
            card.hide()
            card.update_counts(0, 0, 0)
            card.set_selected(False)
        self._table_title.setText("Selecione um ou mais tipos de erro")
        self._table_count_lbl.setText("")
        # Recolhimento dos capítulos volta ao estado expandido em um novo ciclo.
        self._audit_collapsed = False
        self._kit_collapsed = False
        self._audit_body.setVisible(True)
        self._btn_audit_collapse.setText("−")
        self._btn_kit_collapse.setText("−")
        self._btn_run.setEnabled(True)

    def _export_evidence(self):
        if not self._result or self._result.evidence.empty:
            QMessageBox.information(self, "Evidências", "Nenhuma evidência disponível.")
            return
        
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportar Evidências (Master Report)", "MASTER_AUDIT_REPORT.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        
        path = get_unique_path(path)
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                self._result.evidence.to_excel(writer, sheet_name="EVIDENCIAS_AUDITORIA", index=False)
            QMessageBox.information(self, "Exportado", f"Relatório de Evidências salvo em:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Erro ao Exportar", str(exc))

    def _export_master_audit(self):
        """
        Safety-gated: gera Certificado de Conformidade (PDF) + Excel EVIDENCIAS_MASTER.
        Bloqueado se result.stats["total"] != 0.
        """
        if not self._result:
            return

        if self._last_kit_only:
            # Defesa em profundidade: o botão já fica desabilitado no modo
            # só-kit (_on_finished), mas total==0 trivialmente nesse modo
            # (stats vazios), o que passaria pelo gate abaixo sem essa checagem.
            QMessageBox.critical(
                self,
                "Exportação Bloqueada — Modo Só-Kit",
                "O Certificado Mestre não pode ser emitido a partir de uma execução "
                "só-kit: ele atesta conformidade do catálogo inteiro, que não foi auditado.",
            )
            return

        total = self._result.stats.get("total", 0)
        if total != 0:
            QMessageBox.critical(
                self,
                "Exportação Bloqueada — Conformidade Pendente",
                f"O Certificado de Conformidade não pode ser emitido.\n\n"
                f"A auditoria encontrou {total} divergência(s) pendente(s).\n\n"
                f"Corrija todas as divergências no Salesforce e execute a auditoria "
                f"novamente antes de emitir o certificado.",
            )
            return

        pdf_path, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Certificado de Conformidade (PDF)",
            "CERTIFICADO_CONFORMIDADE.pdf",
            "PDF (*.pdf)",
        )
        if not pdf_path:
            return

        pdf_path = get_unique_path(pdf_path)

        source_files: list[str] = []
        if pb := self._dz_pb.file_path:
            source_files.append(Path(pb).name)
        for p in (self._dz_cat.file_paths or []):
            source_files.append(Path(p).name)
        for p in (self._dz_excel.file_paths or []):
            source_files.append(Path(p).name)

        try:
            CertificateEngine().generate(
                result=self._result,
                output_path=pdf_path,
                source_files=source_files,
            )
        except Exception as exc:
            QMessageBox.critical(self, "Erro — Certificado PDF", str(exc))
            return

        QMessageBox.information(
            self,
            "Certificado Emitido",
            f"Certificado de Conformidade salvo em:\n{pdf_path}",
        )

    def _write_evidence_master(self, path: str) -> None:
        """Gera Excel estilizado com trilha de auditoria completa (aba EVIDENCIAS_MASTER)."""
        from datetime import datetime as _dt
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if self._result is None or self._result.evidence.empty:
            return

        df = self._result.evidence.copy()

        expected_cols = ["SKU", "MARCA", "FONTE", "ATRIBUTO", "VALOR_EXCEL", "VALOR_SALESFORCE", "STATUS"]
        for col in expected_cols:
            if col not in df.columns:
                df[col] = ""
        df["TIMESTAMP"] = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
        df = df[expected_cols + ["TIMESTAMP"]]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EVIDENCIAS_MASTER"

        HEADER_FILL  = PatternFill("solid", fgColor="0D1F3D")
        OK_FILL      = PatternFill("solid", fgColor="D1FAE5")
        ERROR_FILL   = PatternFill("solid", fgColor="FEE2E2")
        ALT_FILL     = PatternFill("solid", fgColor="F5F7FA")

        HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        BODY_FONT    = Font(name="Calibri", size=9)
        OK_FONT      = Font(name="Calibri", size=9, color="065F46", bold=True)
        ERROR_FONT   = Font(name="Calibri", size=9, color="991B1B", bold=True)

        _side = lambda: Side(style="thin", color="CBD5E1")
        THIN_BORDER  = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        CENTER       = Alignment(horizontal="center", vertical="center")
        LEFT         = Alignment(horizontal="left",   vertical="center")

        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 20

        for idx, (_, row) in enumerate(df.iterrows(), start=2):
            ws.append(list(row))
            status  = str(row.get("STATUS", "")).upper()
            is_ok   = any(k in status for k in ("OK", "ACERTO", "CONFORME"))
            is_err  = any(k in status for k in ("ERRO", "ERROR", "DIVERG"))

            for col_idx, cell in enumerate(ws[idx], start=1):
                col_name = df.columns[col_idx - 1]
                cell.border = THIN_BORDER

                if col_name == "STATUS":
                    cell.alignment = CENTER
                    if is_ok:
                        cell.fill, cell.font = OK_FILL, OK_FONT
                    elif is_err:
                        cell.fill, cell.font = ERROR_FILL, ERROR_FONT
                    else:
                        cell.alignment = LEFT
                        cell.font = BODY_FONT
                else:
                    cell.alignment = LEFT
                    if is_ok:
                        cell.fill, cell.font = OK_FILL, BODY_FONT
                    elif is_err:
                        cell.fill, cell.font = ERROR_FILL, BODY_FONT
                    elif idx % 2 == 0:
                        cell.fill, cell.font = ALT_FILL, BODY_FONT
                    else:
                        cell.font = BODY_FONT

        col_widths = {
            "SKU": 20, "MARCA": 12, "FONTE": 18, "ATRIBUTO": 24,
            "VALOR_EXCEL": 18, "VALOR_SALESFORCE": 20, "STATUS": 14, "TIMESTAMP": 20,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

        wb.save(path)

    def refresh_theme(self):
        """Update UI components that have hardcoded theme colors (like HTML panels)."""
        if self._result:
            self._on_finished(self._result)
