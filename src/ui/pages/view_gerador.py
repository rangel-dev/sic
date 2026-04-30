"""
Gerador View – Excel → Pricebook XML generator.
Detecta automaticamente a marca de cada grade inserida (Natura / Avon) e sempre
inclui CB (Minha Loja) no XML gerado.
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QButtonGroup,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QVBoxLayout,
    QWidget,
)

from src.ui.components.base_widgets import Divider, DropZone, SectionHeader, StatPill
from src.workers.worker_gerador import GeradorWorker
from src.core.history_engine import HistoryEngine
from src.core.utils import get_unique_path


# ── Brand palette ────────────────────────────────────────────────────────────
BRAND_COLORS: dict[str, str] = {
    "natura": "#f59e0b",
    "avon":   "#c4b5fd",
    "ml":     "#60a5fa",
}
BRAND_LABELS: dict[str, str] = {
    "natura": "Natura",
    "avon":   "Avon",
    "ml":     "CB (Minha Loja)",
}


# ── Fast brand-sniff (reads only first 300 rows) ─────────────────────────────
def _sniff_brand(path: str) -> str | None:
    """
    Quickly detects whether the Excel file is Natura or Avon.
    Returns "natura", "avon", or None.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        target_sheet = None
        for name in wb.sheetnames:
            if "GRADE" in name.upper() and "ATIVA" in name.upper():
                target_sheet = name
                break
        if target_sheet is None:
            wb.close()
            return None

        ws = wb[target_sheet]
        nat = avn = 0
        for row in ws.iter_rows(max_row=300, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    v = cell.strip().upper()
                    if v.startswith("NATBRA-"):
                        nat += 1
                    elif v.startswith("AVNBRA-"):
                        avn += 1
            if nat + avn >= 5:
                break
        wb.close()

        if nat == 0 and avn == 0:
            return None
        return "natura" if nat >= avn else "avon"
    except Exception:
        return None


# ── Main view ────────────────────────────────────────────────────────────────
class GeradorView(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._worker: GeradorWorker | None = None
        self._result: dict | None = None

        # Tracks brand for each loaded file path: {path: "natura"|"avon"|None}
        self._file_brands: dict[str, str | None] = {}

        self._setup_ui()

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        outer.addWidget(SectionHeader(
            "⊕  Gerador de Pricebook",
            "Insira até duas grades Excel (uma Natura e uma Avon). "
            "A marca é detectada automaticamente — CB (Minha Loja) é sempre incluído."
        ))
        outer.addWidget(Divider())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        outer.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(20)

        # ── Grades de Entrada ──────────────────────────────────────────────
        files_box = QGroupBox("Grades de Ativação")
        files_layout = QVBoxLayout(files_box)
        files_layout.setContentsMargins(16, 18, 16, 14)
        files_layout.setSpacing(12)

        lbl_excel = QLabel(
            "Insira até <b>duas grades</b> — uma "
            "<b style='color:#f59e0b'>Natura</b> e uma "
            "<b style='color:#c4b5fd'>Avon</b>. "
            "Arraste ou clique para selecionar."
        )
        lbl_excel.setObjectName("label_section")
        lbl_excel.setWordWrap(True)
        files_layout.addWidget(lbl_excel)

        self._dz_excel = DropZone(
            "Arraste as grades aqui  —  GRADE DE ATIVAÇÃO\n"
            "(até 2 arquivos: Natura e/ou Avon)",
            "Excel (*.xlsx *.xlsm *.xls)",
            multiple=True,
        )
        self._dz_excel.setToolTip(
            "Insira até 2 grades. Não é permitido inserir duas grades da mesma marca."
        )
        self._dz_excel.files_selected.connect(self._on_files_selected)
        files_layout.addWidget(self._dz_excel)

        # Badges de marca detectada (aparecem abaixo do drop)
        self._badges_row = QHBoxLayout()
        self._badges_row.setSpacing(10)
        self._badge_a = QLabel("")
        self._badge_b = QLabel("")
        for badge in (self._badge_a, self._badge_b):
            badge.setFixedHeight(28)
            badge.setStyleSheet(
                "font-size:12px; font-weight:700; border-radius:5px; "
                "padding:0 12px; background:transparent;"
            )
            badge.hide()
            self._badges_row.addWidget(badge)
        self._badges_row.addStretch()
        files_layout.addLayout(self._badges_row)

        # Preview do XML que será gerado
        self._preview_lbl = QLabel("")
        self._preview_lbl.setObjectName("label_muted")
        self._preview_lbl.setWordWrap(True)
        self._preview_lbl.hide()
        files_layout.addWidget(self._preview_lbl)

        layout.addWidget(files_box)

        # ── Modo de Geração ────────────────────────────────────────────────
        mode_box = QGroupBox("Modo de Geração")
        mode_layout = QHBoxLayout(mode_box)
        mode_layout.setContentsMargins(16, 18, 16, 14)
        mode_layout.setSpacing(24)

        self._bg = QButtonGroup(self)
        self._radio_full  = QRadioButton("Full Grade  (todos os produtos)")
        self._radio_delta = QRadioButton("Delta Ajustes  (somente alterados)")
        self._radio_full.setChecked(True)
        self._bg.addButton(self._radio_full,  0)
        self._bg.addButton(self._radio_delta, 1)
        self._bg.buttonClicked.connect(self._on_mode_changed)

        mode_layout.addWidget(self._radio_full)
        mode_layout.addWidget(self._radio_delta)
        mode_layout.addStretch()
        layout.addWidget(mode_box)

        # ── Delta base XML (oculto até modo delta) ─────────────────────────
        self._delta_widget = QWidget()
        delta_inner = QVBoxLayout(self._delta_widget)
        delta_inner.setContentsMargins(0, 0, 0, 0)
        delta_inner.setSpacing(6)
        lbl_base = QLabel("Pricebook XML Base (referência para o Delta)")
        lbl_base.setObjectName("label_section")
        delta_inner.addWidget(lbl_base)
        self._dz_base = DropZone(
            "Arraste o Pricebook XML atual (base de comparação)",
            "XML (*.xml)",
        )
        delta_inner.addWidget(self._dz_base)
        self._delta_widget.hide()
        layout.addWidget(self._delta_widget)

        # ── Action bar ────────────────────────────────────────────────────
        action_row = QHBoxLayout()
        action_row.setSpacing(12)

        self._btn_run = QPushButton("⊕  Gerar Pricebook XML")
        self._btn_run.setObjectName("btn_primary")
        self._btn_run.setFixedWidth(220)
        self._btn_run.clicked.connect(self._run)
        action_row.addWidget(self._btn_run)

        self._btn_clear = QPushButton("Limpar")
        self._btn_clear.setObjectName("btn_ghost")
        self._btn_clear.clicked.connect(self._clear)
        action_row.addWidget(self._btn_clear)

        action_row.addStretch()
        layout.addLayout(action_row)

        # ── Progress ──────────────────────────────────────────────────────
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, 100)
        self._progress_bar.setValue(0)
        self._progress_bar.hide()
        layout.addWidget(self._progress_bar)

        self._status_lbl = QLabel("")
        self._status_lbl.setObjectName("label_muted")
        self._status_lbl.hide()
        layout.addWidget(self._status_lbl)

        # ── Stats row — pricebooks gerados ────────────────────────────────
        self._stats_widget = QWidget()
        stats_row = QHBoxLayout(self._stats_widget)
        stats_row.setContentsMargins(0, 0, 0, 0)
        stats_row.setSpacing(12)

        self._stat_total  = StatPill("Total SKUs",         "—",  "#f0f0f0")
        self._stat_pb_nat = StatPill("Pricebook Natura",   "—",  BRAND_COLORS["natura"])
        self._stat_pb_avn = StatPill("Pricebook Avon",     "—",  BRAND_COLORS["avon"])
        self._stat_pb_ml  = StatPill("Pricebook Minha Loja", "—", BRAND_COLORS["ml"])
        self._stat_mode   = StatPill("Modo",               "—",  "#888888")

        for w in (self._stat_total, self._stat_pb_nat, self._stat_pb_avn,
                  self._stat_pb_ml, self._stat_mode):
            stats_row.addWidget(w)
        stats_row.addStretch()
        self._stats_widget.hide()
        layout.addWidget(self._stats_widget)

        # ── Download ──────────────────────────────────────────────────────
        self._btn_download = QPushButton("⬇  Salvar XML")
        self._btn_download.setObjectName("btn_secondary")
        self._btn_download.setFixedWidth(160)
        self._btn_download.clicked.connect(self._save_xml)
        self._btn_download.hide()
        layout.addWidget(self._btn_download)

        layout.addStretch()

    # ── File selection handler ────────────────────────────────────────────
    def _on_files_selected(self, paths: list[str]) -> None:
        """Called by DropZone whenever the file list changes."""
        # Hard limit: max 2 files
        if len(paths) > 2:
            QMessageBox.warning(
                self, "Gerador",
                "Insira no máximo <b>duas grades</b>: uma Natura e uma Avon."
            )
            self._dz_excel.clear()
            self._reset_badges()
            return

        # Sniff brand for any new paths
        for p in paths:
            if p not in self._file_brands:
                self._file_brands[p] = _sniff_brand(p)

        # Remove brands for paths no longer selected
        current = set(paths)
        for old in list(self._file_brands):
            if old not in current:
                del self._file_brands[old]

        # Validate: no duplicate brands
        brands = [b for b in self._file_brands.values() if b]
        if len(brands) == 2 and brands[0] == brands[1]:
            brand_name = BRAND_LABELS.get(brands[0], brands[0].capitalize())
            QMessageBox.warning(
                self, "Grades incompatíveis",
                f"Ambas as grades pertencem à marca <b>{brand_name}</b>.<br><br>"
                "Insira uma grade <b>Natura</b> e uma grade <b>Avon</b>.",
            )
            self._dz_excel.clear()
            self._file_brands.clear()
            self._reset_badges()
            return

        self._update_badges(paths)
        self._update_preview()

    def _update_badges(self, paths: list[str]) -> None:
        badges = [self._badge_a, self._badge_b]
        for i, badge in enumerate(badges):
            if i < len(paths):
                brand = self._file_brands.get(paths[i])
                self._apply_badge(badge, brand)
                badge.show()
            else:
                badge.hide()

    def _apply_badge(self, badge: QLabel, brand: str | None) -> None:
        if brand == "natura":
            color, text = BRAND_COLORS["natura"], "🟧  Natura detectada ✓"
        elif brand == "avon":
            color, text = BRAND_COLORS["avon"],   "🟪  Avon detectada ✓"
        else:
            color, text = "#888888", "❓  Marca não identificada"
        badge.setText(text)
        badge.setStyleSheet(
            f"font-size:12px; font-weight:700; color:{color}; "
            "border-radius:5px; padding:0 12px; background:transparent;"
        )

    def _reset_badges(self) -> None:
        self._badge_a.hide()
        self._badge_b.hide()
        self._preview_lbl.hide()

    def _update_preview(self) -> None:
        active = [b for b in self._file_brands.values() if b]
        if not active:
            self._preview_lbl.hide()
            return

        parts: list[str] = []
        if "natura" in active:
            parts.append(f"<b style='color:{BRAND_COLORS['natura']}'>Pricebook Natura</b>")
        if "avon" in active:
            parts.append(f"<b style='color:{BRAND_COLORS['avon']}'>Pricebook Avon</b>")
        parts.append(f"<b style='color:{BRAND_COLORS['ml']}'>Pricebook Minha Loja</b>")

        self._preview_lbl.setText("📋  Será gerado:  " + "  +  ".join(parts))
        self._preview_lbl.show()

    # ── Slots ─────────────────────────────────────────────────────────────
    def _on_mode_changed(self):
        self._delta_widget.setVisible(self._radio_delta.isChecked())

    def _run(self):
        excel_paths = [
            p for p, b in self._file_brands.items() if b in ("natura", "avon")
        ]
        if not excel_paths:
            QMessageBox.warning(
                self, "Gerador",
                "Selecione ao menos uma grade Excel válida.\n\n"
                "A planilha deve conter a aba 'GRADE DE ATIVAÇÃO' com SKUs "
                "NATBRA- (Natura) ou AVNBRA- (Avon).",
            )
            return

        mode     = "delta" if self._radio_delta.isChecked() else "full"
        base_xml = self._dz_base.file_path if mode == "delta" else None

        self._btn_run.setEnabled(False)
        self._btn_download.hide()
        self._stats_widget.hide()
        self._progress_bar.setValue(0)
        self._progress_bar.show()
        self._status_lbl.show()

        self._worker = GeradorWorker(excel_paths, mode, base_xml, self)
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.error_msg.connect(self._on_error)
        self._worker.start()

    def _on_progress(self, pct: int, msg: str):
        self._progress_bar.setValue(pct)
        self._status_lbl.setText(msg)

    def _on_finished(self, result: dict):
        self._result = result
        self._btn_run.setEnabled(True)

        s      = result.get("stats", {})
        brands = s.get("brands", [])   # sorted list e.g. ["natura"] or ["avon", "natura"]

        # ── Total SKUs ────────────────────────────────────────────────────
        self._stat_total.set_value(str(s.get("total", 0)))

        # ── Pricebook Natura ──────────────────────────────────────────────
        if "natura" in brands:
            self._stat_pb_nat.set_value("✓", BRAND_COLORS["natura"])
        else:
            self._stat_pb_nat.set_value("—", "#444")

        # ── Pricebook Avon ────────────────────────────────────────────────
        if "avon" in brands:
            self._stat_pb_avn.set_value("✓", BRAND_COLORS["avon"])
        else:
            self._stat_pb_avn.set_value("—", "#444")

        # ── Pricebook Minha Loja — sempre ─────────────────────────────────
        self._stat_pb_ml.set_value("✓", BRAND_COLORS["ml"])

        # ── Modo ──────────────────────────────────────────────────────────
        self._stat_mode.set_value(s.get("mode", "—").upper(), "#888")

        self._stats_widget.show()
        self._btn_download.show()

        # ── Status bar ────────────────────────────────────────────────────
        brands_display = (
            " + ".join(BRAND_LABELS.get(b, b) for b in brands)
            + " + CB (Minha Loja)"
        )
        if parent := self.parent():
            if hasattr(parent, "show_status"):
                parent.show_status(
                    f"Gerador: {s.get('total', 0)} SKUs → {brands_display} "
                    f"({s.get('mode', '')})"
                )

        # ── History ───────────────────────────────────────────────────────
        HistoryEngine.add_entry(
            "Gerador",
            result.get("brand", "Desconhecida"),
            f"Pricebook gerado ({s.get('mode', 'full').upper()}) → "
            f"{brands_display}: {s.get('total', 0)} SKUs.",
        )

    def _on_error(self, msg: str):
        self._btn_run.setEnabled(True)
        self._progress_bar.hide()
        self._status_lbl.hide()
        QMessageBox.critical(self, "Erro — Gerador", msg)

    def _save_xml(self):
        if not self._result or not self._result.get("xml_content"):
            return

        from datetime import date
        s      = self._result.get("stats", {})
        brands = s.get("brands", [])

        today      = date.today()
        dd         = today.strftime("%d")
        mm         = today.strftime("%m")
        
        if brands:
            brands_str = "_".join(b.upper() for b in brands) + "+CB"
        else:
            brands_str = "CB"
            
        default_name = f"---{dd}.{mm}-{brands_str}-PRICEBOOK-AJUSTE.xml"

        path, _ = QFileDialog.getSaveFileName(
            self, "Salvar Pricebook XML", default_name, "XML (*.xml)"
        )
        if path:
            path = get_unique_path(path)
            with open(path, "wb") as f:
                f.write(self._result["xml_content"])
            QMessageBox.information(self, "Salvo", f"XML salvo em:\n{path}")

    def _clear(self):
        self._dz_excel.clear()
        self._dz_base.clear()
        self._file_brands.clear()
        self._reset_badges()
        self._progress_bar.setValue(0)
        self._progress_bar.hide()
        self._status_lbl.hide()
        self._stats_widget.hide()
        self._btn_download.hide()
        self._result = None
        self._radio_full.setChecked(True)
        self._delta_widget.hide()
        self._btn_run.setEnabled(True)
